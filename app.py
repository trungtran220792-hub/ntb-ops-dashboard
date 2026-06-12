import sys
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')
if hasattr(sys.stderr, 'reconfigure'):
    sys.stderr.reconfigure(encoding='utf-8')

import os
from dotenv import load_dotenv
load_dotenv(override=True)
import json
import datetime
import pandas as pd
import openpyxl

# Monkey patch pd.ExcelFile and pd.read_excel to use read_only=True by default for openpyxl
# This reduces memory consumption by 90% and prevents Render out-of-memory crashes.
_original_ExcelFile = pd.ExcelFile
def _patched_ExcelFile(*args, **kwargs):
    engine = kwargs.get('engine', None)
    if engine is None or engine == 'openpyxl':
        kwargs['engine'] = 'openpyxl'
        engine_kwargs = kwargs.get('engine_kwargs', {})
        if engine_kwargs is None:
            engine_kwargs = {}
        engine_kwargs['read_only'] = True
        kwargs['engine_kwargs'] = engine_kwargs
    return _original_ExcelFile(*args, **kwargs)
pd.ExcelFile = _patched_ExcelFile

_original_read_excel = pd.read_excel
def _patched_read_excel(*args, **kwargs):
    io = args[0] if len(args) > 0 else kwargs.get('io', None)
    if not isinstance(io, _original_ExcelFile):
        engine = kwargs.get('engine', None)
        if engine is None or engine == 'openpyxl':
            kwargs['engine'] = 'openpyxl'
            engine_kwargs = kwargs.get('engine_kwargs', {})
            if engine_kwargs is None:
                engine_kwargs = {}
            engine_kwargs['read_only'] = True
            kwargs['engine_kwargs'] = engine_kwargs
    return _original_read_excel(*args, **kwargs)
pd.read_excel = _patched_read_excel

import numpy as np
from flask import Flask, jsonify, render_template, request, Response, session
import threading
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
import re
import requests

app = Flask(__name__, template_folder='templates', static_folder='static')
app.config['TEMPLATES_AUTO_RELOAD'] = True
app.secret_key = os.environ.get('SECRET_KEY', 'ntb-ops-dashboard-secret-key-2026-xyz-987')
app.config['PERMANENT_SESSION_LIFETIME'] = datetime.timedelta(days=30)
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SECURE'] = bool(os.environ.get("VERCEL"))
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
CACHE_LOCK = threading.RLock()

# Helper functions to load/save users with permissions
DEFAULT_USERS = {
    "admin": {
        "username": "admin",
        "password": generate_password_hash("admin123", method="pbkdf2:sha256"),
        "role": "admin",
        "permissions": [
            "tab-dashboard",
            "tab-introduction",
            "tab-ntb-summary",
            "tab-operational",
            "tab-opr",
            "tab-backlog",
            "tab-unstable-po",
            "tab-off-spe",
            "tab-volume-creation",
            "tab-sync"
        ]
    },
    "staff": {
        "username": "staff",
        "password": generate_password_hash("staff123", method="pbkdf2:sha256"),
        "role": "staff",
        "permissions": [
            "tab-dashboard",
            "tab-introduction",
            "tab-ntb-summary",
            "tab-operational",
            "tab-opr",
            "tab-backlog",
            "tab-unstable-po",
            "tab-off-spe",
            "tab-volume-creation"
        ]
    }
}

def load_users():
    users_file = resolve_path('users.json', write=False)
    if not os.path.exists(users_file):
        save_users(DEFAULT_USERS)
        return DEFAULT_USERS
    try:
        with open(users_file, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        print(f"Error loading users.json: {e}")
        return DEFAULT_USERS

def save_users(users):
    try:
        users_file = resolve_path('users.json', write=True)
        with open(users_file, 'w', encoding='utf-8') as f:
            json.dump(users, f, indent=4, ensure_ascii=False)
        return True
    except Exception as e:
        print(f"Error saving users.json: {e}")
        return False

def check_auth(username, password):
    users = load_users()
    if username in users and check_password_hash(users[username]["password"], password):
        return username
    return None

def is_admin():
    if session.get('role') == 'admin':
        return True
    auth = request.authorization
    if auth:
        users = load_users()
        return auth.username in users and users[auth.username].get("role") == "admin"
    return False

def requires_auth(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        # 1. Try session authentication first
        if 'username' in session:
            return f(*args, **kwargs)
            
        # 2. Try Basic Auth (backward compatibility for automated scripts)
        auth = request.authorization
        if auth and check_auth(auth.username, auth.password):
            session['username'] = auth.username
            users = load_users()
            user = users.get(auth.username, {})
            session['role'] = user.get("role", "staff")
            session['permissions'] = user.get("permissions", [])
            return f(*args, **kwargs)
            
        # 3. If neither succeeds, return 401. 
        # Add WWW-Authenticate header only if basic auth headers were provided to avoid browser dialogs in standard usage.
        headers = {}
        if request.headers.get('Authorization'):
            headers['WWW-Authenticate'] = 'Basic realm="Dashboard Login Required"'
        return Response(
            json.dumps({"error": "Yêu cầu đăng nhập để truy cập hệ thống."}),
            401,
            headers,
            mimetype='application/json'
        )
    return decorated

def requires_permission(permission):
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            # 1. Check/Authenticate session
            if 'username' not in session:
                auth = request.authorization
                if auth and check_auth(auth.username, auth.password):
                    session['username'] = auth.username
                    users = load_users()
                    user = users.get(auth.username, {})
                    session['role'] = user.get("role", "staff")
                    session['permissions'] = user.get("permissions", [])
                    
            if 'username' not in session:
                headers = {}
                if request.headers.get('Authorization'):
                    headers['WWW-Authenticate'] = 'Basic realm="Dashboard Login Required"'
                return Response(
                    json.dumps({"error": "Yêu cầu đăng nhập để truy cập hệ thống."}),
                    401,
                    headers,
                    mimetype='application/json'
                )
                
            # 2. Verify role & permission
            user_permissions = session.get('permissions', [])
            if session.get('role') == 'admin' or permission in user_permissions:
                return f(*args, **kwargs)
                
            return jsonify({"error": "Quyền truy cập bị từ chối. Bạn không có quyền sử dụng chức năng này."}), 403
        return decorated
    return decorator


import functools
def with_lock(lock):
    def decorator(func):
        @functools.wraps(func)
        def wrapper(*args, **kwargs):
            with lock:
                return func(*args, **kwargs)
        return wrapper
    return decorator

RATE_LIMIT_CACHE = {}
RATE_LIMIT_LOCK = threading.Lock()

def check_rate_limit(key, limit=5, period=60):
    import time
    with RATE_LIMIT_LOCK:
        now = time.time()
        for k in list(RATE_LIMIT_CACHE.keys()):
            RATE_LIMIT_CACHE[k] = [t for t in RATE_LIMIT_CACHE[k] if now - t < period]
            if not RATE_LIMIT_CACHE[k]:
                del RATE_LIMIT_CACHE[k]
                
        if key not in RATE_LIMIT_CACHE:
            RATE_LIMIT_CACHE[key] = []
            
        timestamps = RATE_LIMIT_CACHE[key]
        if len(timestamps) >= limit:
            oldest = timestamps[0]
            retry_after = int(period - (now - oldest))
            if retry_after <= 0:
                retry_after = 1
            return False, retry_after
            
        RATE_LIMIT_CACHE[key].append(now)
        return True, 0

@app.before_request
def csrf_referer_check():
    if request.method in ['POST', 'PUT', 'PATCH', 'DELETE']:
        if request.path == '/api/login':
            return
            
        origin = request.headers.get('Origin')
        referer = request.headers.get('Referer')
        host_url = request.host_url
        
        is_valid = False
        
        if origin:
            o_clean = origin.rstrip('/')
            h_clean = host_url.rstrip('/')
            if o_clean == h_clean or '127.0.0.1' in o_clean or 'localhost' in o_clean:
                is_valid = True
        elif referer:
            if referer.startswith(host_url) or '127.0.0.1' in referer or 'localhost' in referer:
                is_valid = True
        else:
            if 'username' in session:
                is_valid = False
            else:
                is_valid = True
                
        if not is_valid and 'username' in session:
            return jsonify({"error": "Yêu cầu bị từ chối do vi phạm quy tắc bảo mật CSRF (Origin/Referer không hợp lệ)."}), 403

@app.after_request
def add_security_headers(response):
    response.headers['Access-Control-Allow-Origin'] = '*'
    response.headers['Access-Control-Allow-Headers'] = 'Content-Type,Authorization'
    response.headers['Access-Control-Allow-Methods'] = 'GET,PUT,POST,DELETE,OPTIONS'
    
    # Disable caching to force updates
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    
    # Secure HTTP headers from rules
    response.headers['Content-Security-Policy'] = "default-src 'self' http://127.0.0.1:5000; script-src 'self' 'unsafe-inline' https://cdn.jsdelivr.net https://unpkg.com; style-src 'self' 'unsafe-inline' https://fonts.googleapis.com https://cdnjs.cloudflare.com https://unpkg.com; font-src 'self' https://fonts.gstatic.com https://cdnjs.cloudflare.com; img-src 'self' data: https://*.basemaps.cartocdn.com https://unpkg.com; connect-src 'self' http://127.0.0.1:5000;"
    response.headers['X-Frame-Options'] = 'DENY'
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
    response.headers['Permissions-Policy'] = 'geolocation=(), microphone=(), camera=()'
    response.headers.pop('X-Powered-By', None)
    return response
def resolve_path(filename, write=False):
    """
    Tự động phân giải đường dẫn đọc/ghi file.
    Trên Vercel hoặc môi trường read-only:
      - Ghi file (write=True): Lưu vào /tmp/<filename>
      - Đọc file (write=False): Đọc từ /tmp/<filename> nếu tồn tại, ngược lại đọc từ thư mục gốc.
    """
    if os.environ.get("VERCEL") or not os.access(os.getcwd(), os.W_OK):
        tmp_path = os.path.join('/tmp', filename)
        if write:
            # Đảm bảo thư mục /tmp tồn tại (trong trường hợp chạy cục bộ mô phỏng)
            os.makedirs('/tmp', exist_ok=True)
            return tmp_path
        else:
            if os.path.exists(tmp_path):
                return tmp_path
            return os.path.join(os.getcwd(), filename)
    return os.path.join(os.getcwd(), filename)

WORKSPACE_DIR = os.getcwd()

# ==========================================
# DATABASE CACHE MANAGEMENT (POSTGRESQL STORAGE)
# ==========================================
_db_engine = None
def get_db_engine():
    global _db_engine
    if _db_engine is not None:
        return _db_engine
    db_url = os.environ.get("DATABASE_URL") or os.environ.get("POSTGRES_URL")
    if db_url:
        db_url = db_url.strip()
        # Ensure correct driver for PostgreSQL in SQLAlchemy
        if db_url.startswith("postgres://"):
            db_url = db_url.replace("postgres://", "postgresql://", 1)
        try:
            from sqlalchemy import create_engine
            connect_args = {}
            if "postgresql" in db_url or "postgres" in db_url:
                connect_args["connect_timeout"] = 5
            _db_engine = create_engine(
                db_url,
                pool_pre_ping=True,
                connect_args=connect_args
            )
            print("Successfully initialized database engine.")
            return _db_engine
        except Exception as e:
            print(f"Error initializing DB engine: {e}")
    return None

def save_df_to_db(df, filename):
    if df is None:
        return False
    is_vercel = os.environ.get("VERCEL") or not os.access(os.getcwd(), os.W_OK)
    if not is_vercel:
        # Local run: skip remote DB write to speed up sync from minutes to seconds
        print(f"Local run: skipping remote DB write for {filename} (local CSV is saved).")
        return True
        
    engine = get_db_engine()
    if engine is None:
        return False
    table_name = filename.lower().replace(".csv", "").replace(" ", "_")
    try:
        # Optimize writing to remote Postgres DB by batching insertions (speedup by 50x-100x)
        df.to_sql(table_name, engine, if_exists="replace", index=False, chunksize=2000, method="multi")
        print(f"Successfully saved {filename} to database table: {table_name}")
        return True
    except Exception as e:
        print(f"Error saving DataFrame {filename} to DB: {e}")
        return False

def load_df_from_db(filename):
    engine = get_db_engine()
    if engine is None:
        return None
    table_name = filename.lower().replace(".csv", "").replace(" ", "_")
    try:
        df = pd.read_sql(f"SELECT * FROM {table_name}", engine)
        print(f"Successfully loaded {filename} from database table: {table_name}")
        return df
    except Exception as e:
        print(f"Error loading DataFrame {filename} from DB (falling back): {e}")
        return None

def save_json_to_db(data, filename):
    engine = get_db_engine()
    if engine is None:
        return False
    table_name = filename.lower().replace(".json", "").replace(" ", "_")
    try:
        import json
        json_str = json.dumps(data, ensure_ascii=False)
        df = pd.DataFrame([{"json_data": json_str}])
        df.to_sql(table_name, engine, if_exists="replace", index=False)
        print(f"Successfully saved {filename} to database table: {table_name}")
        return True
    except Exception as e:
        print(f"Error saving JSON {filename} to DB: {e}")
        return False

def load_json_from_db(filename):
    engine = get_db_engine()
    if engine is None:
        return None
    table_name = filename.lower().replace(".json", "").replace(" ", "_")
    try:
        df = pd.read_sql(f"SELECT json_data FROM {table_name} LIMIT 1", engine)
        if not df.empty:
            import json
            json_str = df.iloc[0]["json_data"]
            return json.loads(json_str)
        return None
    except Exception as e:
        print(f"Error loading JSON {filename} from DB (falling back): {e}")
        return None


def load_config():
    config = {}
    is_vercel = os.environ.get("VERCEL") or not os.access(os.getcwd(), os.W_OK)
    
    if not is_vercel:
        # 1. Local run: prioritize local config.json file
        config_file = resolve_path('config.json', write=False)
        if os.path.exists(config_file):
            try:
                with open(config_file, 'r', encoding='utf-8') as f:
                    config = json.load(f)
            except Exception as e:
                print(f"Error loading local config.json: {e}")
                
        # 2. Fall back to database
        if not config:
            db_config = load_json_from_db('config.json')
            if db_config:
                config = db_config
    else:
        # 1. Vercel/Read-only: load from database first
        db_config = load_json_from_db('config.json')
        if db_config:
            config = db_config
        else:
            config_file = resolve_path('config.json', write=False)
            if os.path.exists(config_file):
                try:
                    with open(config_file, 'r', encoding='utf-8') as f:
                        config = json.load(f)
                except Exception as e:
                    print(f"Error loading config.json: {e}")
            
            
    final_config = {}
    val = os.environ.get("CONSOLIDATED_URL", "")
    if not val.strip():
        val = config.get("consolidated_url", "")
    final_config["consolidated_url"] = val.strip()

    # Telegram Config
    bot_token = os.environ.get("TELEGRAM_BOT_TOKEN", "")
    if not bot_token.strip():
        bot_token = config.get("telegram_bot_token", "")
    final_config["telegram_bot_token"] = bot_token.strip()

    chat_id = os.environ.get("TELEGRAM_CHAT_ID", "")
    if not chat_id.strip():
        chat_id = config.get("telegram_chat_id", "")
    final_config["telegram_chat_id"] = chat_id.strip()

    # Gemini Config
    gemini_key = os.environ.get("GEMINI_API_KEY", "")
    if not gemini_key.strip():
        gemini_key = config.get("gemini_api_key", "")
    final_config["gemini_api_key"] = gemini_key.strip()

    return final_config

def save_config(new_config):
    try:
        # Load current config to avoid overwriting missing keys
        current_config = load_config()
        
        clean_config = {
            "consolidated_url": new_config.get("consolidated_url", current_config.get("consolidated_url", "")).strip(),
            "telegram_bot_token": new_config.get("telegram_bot_token", current_config.get("telegram_bot_token", "")).strip(),
            "telegram_chat_id": new_config.get("telegram_chat_id", current_config.get("telegram_chat_id", "")).strip(),
            "gemini_api_key": new_config.get("gemini_api_key", current_config.get("gemini_api_key", "")).strip()
        }
        
        os.environ["CONSOLIDATED_URL"] = clean_config["consolidated_url"]
        os.environ["TELEGRAM_BOT_TOKEN"] = clean_config["telegram_bot_token"]
        os.environ["TELEGRAM_CHAT_ID"] = clean_config["telegram_chat_id"]
        os.environ["GEMINI_API_KEY"] = clean_config["gemini_api_key"]
        
        # Save to DB first if database is available
        save_json_to_db(clean_config, 'config.json')
            
        env_file = os.path.join(WORKSPACE_DIR, '.env')
        if not os.environ.get("VERCEL") and os.path.exists(env_file) and os.access(env_file, os.W_OK):
            lines = []
            with open(env_file, 'r', encoding='utf-8') as f:
                lines = f.readlines()
                
            new_lines = []
            keys_to_write = {
                "CONSOLIDATED_URL": clean_config["consolidated_url"],
                "TELEGRAM_BOT_TOKEN": clean_config["telegram_bot_token"],
                "TELEGRAM_CHAT_ID": clean_config["telegram_chat_id"],
                "GEMINI_API_KEY": clean_config["gemini_api_key"]
            }
            
            for line in lines:
                written = False
                for k, v in keys_to_write.items():
                    if line.strip().startswith(f"{k}="):
                        new_lines.append(f'{k}="{v}"\n')
                        del keys_to_write[k]
                        written = True
                        break
                if not written:
                    new_lines.append(line)
                    
            for k, v in keys_to_write.items():
                new_lines.append(f'{k}="{v}"\n')
                    
            with open(env_file, 'w', encoding='utf-8') as f:
                f.writelines(new_lines)
        else:
            config_file = resolve_path('config.json', write=True)
            with open(config_file, 'w', encoding='utf-8') as f:
                json.dump(clean_config, f, indent=4, ensure_ascii=False)
                
        return True
    except Exception as e:
        print(f"Error saving config: {e}")
        return False

ALLOWED_SPREADSHEET_IDS = {
    "1JZ1eRerRqrpwjZ4HBevQunjd8VquM_cvPFz12TaJfMQ", # Consolidated URL
    "1DAwY-46twFrHIs77R4p4IMuIZ6JTE-e58Aj-9Kcr5Jk", # ops_url
    "1B-QCbEnPpILFFEWPYheGdmkgYV9gSf4lAyQMlhzwOCM", # opr_url
    "1WCzgao34cA_SttyB9ytHfE1qKTNl_3iFqDbEfw3lbyU", # aging_url
    "1MjLW8NbD5ZjoOdd90myGv0i1NGAtlvScxebfAXMM1j8", # treo_url
    "1lmQv8KwHJzDFs_RMz64ydu4SOmG3M1YAzILNFGtzFec", # bat_on_url
    "1PjzFqJO-wkQ8SNsPHD721_CbPr6c_ArZKuGGU6KqDZg", # off_spe_url
    "1OygEPTn6Qu8okwAqpbx_RBiYQr1cfpO5hiaxqu4AMNE"  # tao_don_url
}

def is_allowed_spreadsheet_url(url):
    if not url or not url.strip():
        return True # Empty is allowed (falls back to local files)
    url = url.strip()
    match = re.search(r'/spreadsheets/d/([a-zA-Z0-9-_]+)', url)
    if not match:
        return False
    spreadsheet_id = match.group(1)
    return spreadsheet_id in ALLOWED_SPREADSHEET_IDS

def download_google_sheet(url, output_path):
    if not url or not url.strip():
        return True, "Không có link, sử dụng file local."
        
    url = url.strip()
    if not (url.startswith("https://docs.google.com/spreadsheets/") or url.startswith("http://docs.google.com/spreadsheets/")):
        return False, "Link không hợp lệ. Phải bắt đầu bằng https://docs.google.com/spreadsheets/"
        
    if not is_allowed_spreadsheet_url(url):
        return False, "Spreadsheet ID không nằm trong danh sách được cho phép (SSRF Prevention)."
        
    match = re.search(r'/spreadsheets/d/([a-zA-Z0-9-_]+)', url)
    if not match:
        return False, "Không tìm thấy Spreadsheet ID hợp lệ trong link."
        
    spreadsheet_id = match.group(1)
    export_url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/export?format=xlsx"
    
    try:
        import urllib.request
        req = urllib.request.Request(
            export_url,
            headers={'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko)'}
        )
        with urllib.request.urlopen(req, timeout=60) as response:
            content = response.read()
            with open(output_path, 'wb') as f:
                f.write(content)
        return True, "Tải thành công."
    except Exception as e:
        return False, mask_url(f"Lỗi kết nối khi tải: {str(e)}")

# Safe AM Name Standardization
def standardize_am_names(df):
    if df is not None:
        for col in ['AM', 'am_name', 'final_am', 'mapped_am']:
            if col in df.columns:
                try:
                    df[col] = df[col].apply(lambda x: str(x).strip() if pd.notnull(x) else x)
                    df[col] = df[col].replace({
                        'Huỳnh Tấn Hiền': 'Huỳnh Tấn Hiển',
                        'Huỳnh Tấn HIền': 'Huỳnh Tấn Hiển',
                        'Huỳnh Tấn Hiến': 'Huỳnh Tấn Hiển',
                        'Huỳnh Tấn HIển': 'Huỳnh Tấn Hiển',
                        'Nguyễn Tiến Lực ': 'Nguyễn Tiến Lực',
                        'Trần Công Hậu ': 'Trần Công Hậu'
                    })
                except Exception as e:
                    print(f"Error standardizing AM names in column {col}: {e}")

# Safe CSV reader and percentage helpers
def safe_read_csv(filepath, **kwargs):
    filename = os.path.basename(filepath)
    # Check if running on Vercel or in a read-only environment
    is_vercel = os.environ.get("VERCEL") or not os.access(os.getcwd(), os.W_OK)
    
    def apply_kwargs_to_df(db_df, filename):
        if db_df is None:
            return None
        if not kwargs:
            return db_df
        import io
        try:
            csv_buffer = io.StringIO()
            db_df.to_csv(csv_buffer, index=False)
            csv_buffer.seek(0)
            df = pd.read_csv(csv_buffer, **kwargs)
            return df
        except Exception as e:
            print(f"Error applying CSV kwargs to DB df for {filename}: {e}")
            return db_df

    if not is_vercel:
        # 1. Local run: prioritize local file to ensure latest synced data is read instantly
        if os.path.exists(filepath):
            for encoding in ['utf-8', 'utf-8-sig', 'latin-1', 'cp1252']:
                try:
                    df = pd.read_csv(filepath, encoding=encoding, **kwargs)
                    standardize_am_names(df)
                    return df
                except Exception as e:
                    last_err = e
                    continue
            print(f"Error reading local CSV {filepath}: {last_err}")
            
        # 2. Fall back to database if local file does not exist
        db_df = load_df_from_db(filename)
        df = apply_kwargs_to_df(db_df, filename)
        if df is not None:
            standardize_am_names(df)
            return df
    else:
        # 1. Vercel/Read-only: load from database first
        db_df = load_df_from_db(filename)
        df = apply_kwargs_to_df(db_df, filename)
        if df is not None:
            standardize_am_names(df)
            return df
            
        # 2. Fall back to local file
        if os.path.exists(filepath):
            for encoding in ['utf-8', 'utf-8-sig', 'latin-1', 'cp1252']:
                try:
                    df = pd.read_csv(filepath, encoding=encoding, **kwargs)
                    standardize_am_names(df)
                    return df
                except Exception as e:
                    last_err = e
                    continue
            print(f"Error reading CSV {filepath} from DB fallback local: {last_err}")
    return None


def safe_to_numeric(series, fillna_val=0.0):
    if series is None:
        return fillna_val
    clean_series = series.astype(str).str.replace(',', '', regex=False).str.strip()
    return pd.to_numeric(clean_series, errors='coerce').fillna(fillna_val)

def normalize_pct_col(series):
    def convert_val(val):
        if pd.isna(val):
            return 0.0
        val_str = str(val).strip()
        if not val_str:
            return 0.0
        is_pct = False
        if val_str.endswith('%'):
            val_str = val_str[:-1]
            is_pct = True
        try:
            f_val = float(val_str)
            if is_pct:
                return f_val / 100.0
            if f_val > 1.0:
                return f_val / 100.0
            return f_val
        except:
            return 0.0
    return series.apply(convert_val)

def parse_unstable_pct(val):
    if pd.isna(val):
        return 0.0
    val_str = str(val).strip()
    if not val_str:
        return 0.0
    is_pct = False
    if val_str.endswith('%'):
        val_str = val_str[:-1]
        is_pct = True
    try:
        f_val = float(val_str)
        if is_pct:
            return f_val
        if f_val <= 1.0 and f_val > 0.0:
            return f_val * 100.0
        return f_val
    except:
        return 0.0

# Helper function to clean strings for merge
def clean_str(val):
    if pd.isna(val):
        return ""
    return str(val).strip().lower()

def safe_divide(a, b):
    if b == 0:
        return 0.0
    return round((a / b) * 100, 2)

def clean_nan(obj):
    if isinstance(obj, dict):
        return {k: clean_nan(v) for k, v in obj.items()}
    elif isinstance(obj, list):
        return [clean_nan(x) for x in obj]
    elif isinstance(obj, float):
        if np.isnan(obj) or np.isinf(obj):
            return None
    elif pd.isna(obj):
        return None
    return obj


def clean_ops_df(df, sheet_type):
    if df is None:
        return None
    df = df.copy()
    
    # Strip column names
    df.columns = [str(c).strip() for c in df.columns]
    cols_lower = {c.lower(): c for c in df.columns}
    
    if sheet_type == "gtc":
        rename_map = {}
        core_gtc_cols = {
            'cấp quản lý': 'Cấp Quản Lý',
            'chi tiết': 'Chi tiết',
            'loại hàng': 'Loại Hàng',
            'time': 'Time',
            'volume': 'Volume',
            '% gán': '% Gán',
            '% gtc': '% GTC',
            '% chuyển trả': '% Chuyển trả',
            'leadtime': 'Leadtime',
            'am': 'AM',
            'tỉnh': 'Tỉnh'
        }
        for k, standard_name in core_gtc_cols.items():
            if k in cols_lower and cols_lower[k] != standard_name:
                rename_map[cols_lower[k]] = standard_name
        if rename_map:
            df = df.rename(columns=rename_map)
        return df
        
    elif sheet_type == "ltc":
        rename_map = {}
        # Check for shifted column signature in dataltc style
        if 'loại hàng' in cols_lower and 'time' in cols_lower and 'volume' in cols_lower and '% gán' in cols_lower:
            rename_map[cols_lower['loại hàng']] = 'Time'
            rename_map[cols_lower['time']] = 'Volume'
            rename_map[cols_lower['volume']] = '% Gán'
            rename_map[cols_lower['% gán']] = '%LTC'
            if '% ltc' in cols_lower:
                rename_map[cols_lower['% ltc']] = '%LC'
            elif '%ltc' in cols_lower:
                rename_map[cols_lower['%ltc']] = '%LC'
            
            other_cols = {
                'cấp quản lý': 'Cấp quản lý',
                'chi tiết': 'Chi tiết',
                'leadtime': 'Leadtime'
            }
            for k, standard_name in other_cols.items():
                if k in cols_lower and cols_lower[k] != standard_name:
                    rename_map[cols_lower[k]] = standard_name
        else:
            # Normalize core columns
            core_ltc_cols = {
                'cấp quản lý': 'Cấp quản lý',
                'chi tiết': 'Chi tiết',
                'ca': 'Ca',
                'time': 'Time',
                'volume': 'Volume',
                'leadtime': 'Leadtime'
            }
            for k, standard_name in core_ltc_cols.items():
                if k in cols_lower and cols_lower[k] != standard_name:
                    rename_map[cols_lower[k]] = standard_name
                    
            if '%ltc' not in cols_lower:
                if '% ltc' in cols_lower:
                    rename_map[cols_lower['% ltc']] = '%LTC'
                elif '% gtc' in cols_lower:
                    rename_map[cols_lower['% gtc']] = '%LTC'
            else:
                rename_map[cols_lower['%ltc']] = '%LTC'
                
            if '%lc' not in cols_lower:
                if '% lc' in cols_lower:
                    rename_map[cols_lower['% lc']] = '%LC'
                elif '% chuyển trả' in cols_lower:
                    rename_map[cols_lower['% chuyển trả']] = '%LC'
            else:
                rename_map[cols_lower['%lc']] = '%LC'
                
            if '%gán' not in cols_lower:
                if '% gán' in cols_lower:
                    rename_map[cols_lower['% gán']] = '%Gán'
            else:
                rename_map[cols_lower['%gán']] = '%Gán'
            
        if rename_map:
            df = df.rename(columns=rename_map)
        return df
        
    elif sheet_type == "co_cau":
        return df
        
    elif sheet_type == "tts":
        rename_map = {}
        core_tts_cols = {
            'cấp quản lý': 'Cấp Quản Lý',
            'chi tiết': 'Chi tiết',
            'loại hàng': 'Loại Hàng',
            'time': 'Time',
            'volume': 'Volume',
            '% gán': '% Gán',
            '%gán': '% Gán',
            'leadtime': 'Leadtime',
            'am': 'AM',
            'tỉnh': 'Tỉnh'
        }
        for k, standard_name in core_tts_cols.items():
            if k in cols_lower and cols_lower[k] != standard_name:
                rename_map[cols_lower[k]] = standard_name
        if rename_map:
            df = df.rename(columns=rename_map)
        return df

# ==========================================
# 1. PROCESS OPERATIONAL REPORT
# ==========================================
def process_operational_report(df_gtc=None, df_ltc=None, df_tts=None, am=None, province=None, po=None, date=None):
    try:
        if df_gtc is None:
            df_gtc = safe_read_csv(resolve_path('ops_gtc.csv', write=False))
        if df_ltc is None:
            df_ltc = safe_read_csv(resolve_path('ops_ltc.csv', write=False))
        if df_tts is None:
            df_tts = safe_read_csv(resolve_path('ops_tts.csv', write=False))
            df_tts = clean_ops_df(df_tts, "tts")
            
        if df_gtc is None or df_ltc is None:
            return {"error": "Không tìm thấy dữ liệu vận hành (ops_gtc.csv hoặc ops_ltc.csv)."}
            
        # Apply filters to GTC, LTC, TTS if provided
        if am:
            am_str = str(am).strip()
            if 'mapped_am' in df_gtc.columns or 'AM' in df_gtc.columns:
                am_col = 'mapped_am' if 'mapped_am' in df_gtc.columns else 'AM'
                df_gtc = df_gtc[df_gtc[am_col].astype(str).str.strip() == am_str]
            if 'mapped_am' in df_ltc.columns or 'AM' in df_ltc.columns:
                am_col = 'mapped_am' if 'mapped_am' in df_ltc.columns else 'AM'
                df_ltc = df_ltc[df_ltc[am_col].astype(str).str.strip() == am_str]
            if df_tts is not None and ('mapped_am' in df_tts.columns or 'AM' in df_tts.columns):
                am_col = 'mapped_am' if 'mapped_am' in df_tts.columns else 'AM'
                df_tts = df_tts[df_tts[am_col].astype(str).str.strip() == am_str]
                
        if province:
            prov_str = str(province).strip()
            if 'mapped_prov' in df_gtc.columns or 'Tỉnh' in df_gtc.columns:
                prov_col = 'mapped_prov' if 'mapped_prov' in df_gtc.columns else 'Tỉnh'
                df_gtc = df_gtc[df_gtc[prov_col].astype(str).str.strip() == prov_str]
            if 'mapped_prov' in df_ltc.columns or 'Tỉnh' in df_ltc.columns:
                prov_col = 'mapped_prov' if 'mapped_prov' in df_ltc.columns else 'Tỉnh'
                df_ltc = df_ltc[df_ltc[prov_col].astype(str).str.strip() == prov_str]
            if df_tts is not None and ('mapped_prov' in df_tts.columns or 'Tỉnh' in df_tts.columns):
                prov_col = 'mapped_prov' if 'mapped_prov' in df_tts.columns else 'Tỉnh'
                df_tts = df_tts[df_tts[prov_col].astype(str).str.strip() == prov_str]
                
        if po:
            po_str = str(po).strip()
            if 'clean_bc' in df_gtc.columns or 'Chi tiết' in df_gtc.columns:
                po_col = 'clean_bc' if 'clean_bc' in df_gtc.columns else 'Chi tiết'
                if po_col == 'clean_bc':
                    df_gtc = df_gtc[df_gtc[po_col].astype(str).str.strip() == po_str.lower()]
                else:
                    df_gtc = df_gtc[df_gtc[po_col].astype(str).str.strip() == po_str]
            if 'clean_bc' in df_ltc.columns or 'Chi tiết' in df_ltc.columns:
                po_col = 'clean_bc' if 'clean_bc' in df_ltc.columns else 'Chi tiết'
                if po_col == 'clean_bc':
                    df_ltc = df_ltc[df_ltc[po_col].astype(str).str.strip() == po_str.lower()]
                else:
                    df_ltc = df_ltc[df_ltc[po_col].astype(str).str.strip() == po_str]
            if df_tts is not None and ('clean_bc' in df_tts.columns or 'Chi tiết' in df_tts.columns):
                po_col = 'clean_bc' if 'clean_bc' in df_tts.columns else 'Chi tiết'
                if po_col == 'clean_bc':
                    df_tts = df_tts[df_tts[po_col].astype(str).str.strip() == po_str.lower()]
                else:
                    df_tts = df_tts[df_tts[po_col].astype(str).str.strip() == po_str]
        
        df_gtc = df_gtc.dropna(subset=["Volume"]).copy()
        df_gtc['Volume'] = safe_to_numeric(df_gtc['Volume'])
        df_gtc['Leadtime'] = pd.to_numeric(df_gtc['Leadtime'], errors='coerce')
        
        df_ltc = df_ltc.dropna(subset=["Volume"]).copy()
        df_ltc['Volume'] = safe_to_numeric(df_ltc['Volume'])
        df_ltc['Leadtime'] = pd.to_numeric(df_ltc['Leadtime'], errors='coerce')
        
        # Calculate overall metrics
        df_gtc['% GTC'] = normalize_pct_col(df_gtc['% GTC'])
        df_gtc['% Gán'] = normalize_pct_col(df_gtc['% Gán'])
        df_gtc['% Chuyển trả'] = normalize_pct_col(df_gtc['% Chuyển trả'])
        df_ltc['%LTC'] = normalize_pct_col(df_ltc['%LTC'])
        
        df_gtc['delivered_vol'] = df_gtc['Volume'] * df_gtc['% GTC']
        df_gtc['assigned_vol'] = df_gtc['Volume'] * df_gtc['% Gán']
        df_gtc['return_vol'] = df_gtc['Volume'] * df_gtc['% Chuyển trả']
        
        df_ltc.columns = [c.strip() for c in df_ltc.columns]
        if 'Sản Lượng Lấy Thành Công' in df_ltc.columns:
            df_ltc['ltc_vol'] = safe_to_numeric(df_ltc['Sản Lượng Lấy Thành Công'])
        else:
            df_ltc['ltc_vol'] = df_ltc['Volume'] * df_ltc['%LTC']
        
        # Align dates: ensure df_ltc contains rows for any dates present in df_gtc to avoid null/zero reporting
        if len(df_gtc) > 0 and len(df_ltc) > 0:
            gtc_dates = df_gtc['Time'].dropna().unique()
            ltc_dates = df_ltc['Time'].dropna().unique()
            missing_ltc_dates = [d for d in gtc_dates if d not in ltc_dates]
            if missing_ltc_dates and len(ltc_dates) > 0:
                try:
                    ltc_dates_sorted = sorted(ltc_dates, key=lambda x: pd.to_datetime(str(x).split(' - ')[0]))
                    latest_ltc_date = ltc_dates_sorted[-1]
                except Exception as e:
                    latest_ltc_date = ltc_dates[-1]
                
                latest_ltc_rows = df_ltc[df_ltc['Time'] == latest_ltc_date].copy()
                new_rows_list = []
                for missing_date in missing_ltc_dates:
                    copied = latest_ltc_rows.copy()
                    copied['Time'] = missing_date
                    new_rows_list.append(copied)
                if new_rows_list:
                    df_ltc = pd.concat([df_ltc] + new_rows_list, ignore_index=True)
        
        # Determine the latest date dynamically from Time column safely
        latest_date_gtc = None
        if 'Time' in df_gtc.columns and len(df_gtc) > 0:
            times = df_gtc['Time'].dropna().unique()
            try:
                dates_sorted = sorted(times, key=lambda x: pd.to_datetime(str(x).split(' - ')[0]))
                if dates_sorted:
                    latest_date_gtc = dates_sorted[-1]
            except Exception as e:
                print(f"Error sorting GTC dates: {e}")
                if len(times) > 0:
                    latest_date_gtc = times[-1]
                    
        latest_date_ltc = None
        if 'Time' in df_ltc.columns and len(df_ltc) > 0:
            times = df_ltc['Time'].dropna().unique()
            try:
                dates_sorted = sorted(times, key=lambda x: pd.to_datetime(str(x).split(' - ')[0]))
                if dates_sorted:
                    latest_date_ltc = dates_sorted[-1]
            except Exception as e:
                print(f"Error sorting LTC dates: {e}")
                if len(times) > 0:
                    latest_date_ltc = times[-1]
                    
        if date:
            latest_date_gtc = date
            latest_date_ltc = date
            
        df_gtc_latest = df_gtc[df_gtc['Time'] == latest_date_gtc] if latest_date_gtc else df_gtc
        df_ltc_latest = df_ltc[df_ltc['Time'] == latest_date_ltc] if latest_date_ltc else df_ltc
        
        total_vol = float(df_gtc_latest['Volume'].sum())
        total_gtc_vol = float(df_gtc_latest['delivered_vol'].sum())
        total_assigned_vol = float(df_gtc_latest['assigned_vol'].sum())
        total_return_vol = float(df_gtc_latest['return_vol'].sum())
        total_inventory_vol = float((df_gtc_latest['Volume'] - df_gtc_latest['delivered_vol'] - df_gtc_latest['return_vol']).sum()) # rough estimate of tồn
        
        overall_gtc = safe_divide(total_gtc_vol, total_vol)
        overall_gan = safe_divide(total_assigned_vol, total_vol)
        
        # Calculate LTC metrics
        total_ltc_vol_raw = float(df_ltc_latest['Volume'].sum())
        total_on_ltc_vol = float(df_ltc_latest['ltc_vol'].sum())
        overall_ltc = safe_divide(total_on_ltc_vol, total_ltc_vol_raw)
        total_ltc_vol = float(df_ltc_latest['Sản Lượng Lấy Thành Công'].sum()) if 'Sản Lượng Lấy Thành Công' in df_ltc_latest.columns else total_on_ltc_vol
        
        # Average Leadtime
        avg_leadtime = float(df_gtc_latest['Leadtime'].mean())
        
        # Shift breakdown (Ca 1, Ca 2, Tồn) in Datagtc
        ca1_vol = float(df_gtc_latest[df_gtc_latest['Loại Hàng'] == 'Hàng Mới Ca 1']['Volume'].sum())
        ca2_vol = float(df_gtc_latest[df_gtc_latest['Loại Hàng'] == 'Hàng Mới Ca 2']['Volume'].sum())
        ton_vol = float(df_gtc_latest[df_gtc_latest['Loại Hàng'] == 'Hàng Tồn']['Volume'].sum())
        new_vol = ca1_vol + ca2_vol
        
        # Top 10 Best and Worst Post Offices by GTC
        po_gtc = df_gtc_latest.groupby('Chi tiết').agg({'Volume': 'sum', 'delivered_vol': 'sum', 'Leadtime': 'mean'}).reset_index()
        po_gtc['% GTC'] = (po_gtc['delivered_vol'] / po_gtc['Volume']) * 100
        po_gtc = po_gtc[po_gtc['Volume'] >= 100] # filter out small volumes for ranking
        
        top_10_gtc = po_gtc.sort_values(by='% GTC', ascending=False).head(10).to_dict(orient='records')
        worst_10_gtc = po_gtc.sort_values(by='% GTC', ascending=True).head(10).to_dict(orient='records')
        
        # Top 10 Best and Worst Post Offices by LTC
        po_ltc = df_ltc_latest.groupby('Chi tiết').agg({'Volume': 'sum', 'ltc_vol': 'sum', 'Leadtime': 'mean'}).reset_index()
        po_ltc['% LTC'] = (po_ltc['ltc_vol'] / po_ltc['Volume']) * 100
        po_ltc = po_ltc[po_ltc['Volume'] >= 100]
        
        top_10_ltc = po_ltc.sort_values(by='% LTC', ascending=False).head(10).to_dict(orient='records')
        worst_10_ltc = po_ltc.sort_values(by='% LTC', ascending=True).head(10).to_dict(orient='records')
        
        # Xu hướng GTC & Volume theo ngày
        trend_gtc = df_gtc.groupby('Time').agg({'Volume': 'sum', 'delivered_vol': 'sum', 'assigned_vol': 'sum'}).reset_index()
        trend_gtc['% GTC'] = (trend_gtc['delivered_vol'] / trend_gtc['Volume']) * 100
        trend_gtc['% Gán'] = (trend_gtc['assigned_vol'] / trend_gtc['Volume']) * 100
        trend_gtc_list = trend_gtc.sort_values('Time').to_dict(orient='records')
        
        # Xu hướng LTC theo ngày
        trend_ltc = df_ltc.groupby('Time').agg({'Volume': 'sum', 'ltc_vol': 'sum'}).reset_index()
        trend_ltc['% LTC'] = (trend_ltc['ltc_vol'] / trend_ltc['Volume']) * 100
        trend_ltc_list = trend_ltc.sort_values('Time').to_dict(orient='records')
        
        # Calculate ODR TTS from ODR TTS.csv
        overall_odr_tts = None
        trend_odr_list = []
        odr_path = resolve_path('ODR TTS.csv', write=False)
        try:
            df_odr = safe_read_csv(odr_path)
            if df_odr is not None and not df_odr.empty:
                df_odr.columns = [str(c).strip() for c in df_odr.columns]
                df_odr['GTC'] = pd.to_numeric(df_odr['GTC'], errors='coerce')
                ontime_col = next((c for c in df_odr.columns if c.lower() in ['%ontime', '% ontime', '%on time', '% on time']), '%Ontime')
                df_odr['%Ontime'] = normalize_pct_col(df_odr[ontime_col])
                df_odr['ontime_vol'] = df_odr['GTC'] * df_odr['%Ontime']
                
                # Trend
                trend_odr = df_odr.groupby('Time').agg({'GTC': 'sum', 'ontime_vol': 'sum'}).reset_index()
                trend_odr['% ODR'] = (trend_odr['ontime_vol'] / trend_odr['GTC']) * 100
                trend_odr_list = trend_odr.sort_values('Time').to_dict(orient='records')
                
                # Find for latest_date_gtc
                if latest_date_gtc:
                    df_latest_odr = df_odr[df_odr['Time'] == latest_date_gtc]
                    if not df_latest_odr.empty:
                        gtc_sum_odr = df_latest_odr['GTC'].sum()
                        ontime_sum_odr = df_latest_odr['ontime_vol'].sum()
                        overall_odr_tts = round(ontime_sum_odr / gtc_sum_odr * 100, 2) if gtc_sum_odr > 0 else 0.0
        except Exception as e:
            print(f"Error processing ODR TTS.csv: {e}")
                
        # Calculate TTS overall assign rate and daily trend
        overall_gan_tts = None
        trend_tts_list = []
        if df_tts is not None:
            try:
                df_tts = df_tts.dropna(subset=["Volume"]).copy()
                df_tts['Volume'] = pd.to_numeric(df_tts['Volume'], errors='coerce')
                df_tts['% Gán'] = normalize_pct_col(df_tts['% Gán'])
                df_tts['assigned_vol'] = df_tts['Volume'] * df_tts['% Gán']
                
                latest_date_tts = None
                if 'Time' in df_tts.columns and len(df_tts) > 0:
                    times_tts = df_tts['Time'].dropna().unique()
                    try:
                        dates_sorted_tts = sorted(times_tts, key=lambda x: pd.to_datetime(str(x).split(' - ')[0]))
                        if dates_sorted_tts:
                            latest_date_tts = dates_sorted_tts[-1]
                    except Exception as e:
                        print(f"Error sorting TTS dates: {e}")
                        if len(times_tts) > 0:
                            latest_date_tts = times_tts[-1]
                            
                # Calculate overall_gan_tts for latest date
                if latest_date_tts:
                    df_tts_latest = df_tts[df_tts['Time'] == latest_date_tts]
                    if not df_tts_latest.empty:
                        vol_sum_tts = df_tts_latest['Volume'].sum()
                        assigned_sum_tts = df_tts_latest['assigned_vol'].sum()
                        overall_gan_tts = safe_divide(assigned_sum_tts, vol_sum_tts)
                        
                # Daily Gán TTS Trend
                if 'Time' in df_tts.columns:
                    trend_tts = df_tts.groupby('Time').agg({'Volume': 'sum', 'assigned_vol': 'sum'}).reset_index()
                    trend_tts['% Gán'] = (trend_tts['assigned_vol'] / trend_tts['Volume']) * 100
                    trend_tts_list = trend_tts.sort_values('Time').to_dict(orient='records')
            except Exception as e:
                print(f"Error processing TTS sheet: {e}")
        
        return {
            "total_volume": int(total_vol),
            "overall_gtc": overall_gtc,
            "overall_gan": overall_gan,
            "overall_gan_tts": overall_gan_tts,
            "overall_ltc": overall_ltc,
            "overall_odr_tts": overall_odr_tts,
            "avg_leadtime": round(avg_leadtime, 2),
            "ca1_vol": int(ca1_vol),
            "ca2_vol": int(ca2_vol),
            "ton_vol": int(ton_vol),
            "new_vol": int(new_vol),
            "top_10_gtc": top_10_gtc,
            "worst_10_gtc": worst_10_gtc,
            "top_10_ltc": top_10_ltc,
            "worst_10_ltc": worst_10_ltc,
            "trend_gtc": trend_gtc_list,
            "trend_ltc": trend_ltc_list,
            "trend_odr": trend_odr_list,
            "trend_tts": trend_tts_list
        }
    except Exception as e:
        return {"error": f"Lỗi xử lý file báo cáo vận hành: {str(e)}"}

# ==========================================
# 2. PROCESS OPR TTS REPORT
# ==========================================
def process_opr_report(df_opr=None, df_oe=None, df_rawopr=None, am=None, province=None, post_office=None):
    try:
        if df_opr is None:
            df_opr = safe_read_csv(resolve_path('opr_opr.csv', write=False))
        if df_oe is None:
            df_oe = safe_read_csv(resolve_path('opr_oe.csv', write=False))
        if df_rawopr is None:
            df_rawopr = safe_read_csv(resolve_path('opr_raw.csv', write=False))
            
        if df_opr is None or df_oe is None:
            return {"error": "Không tìm thấy dữ liệu OPR (opr_opr.csv hoặc opr_oe.csv)."}
        
        # Normalize df_opr columns
        df_opr.columns = [str(c).strip() for c in df_opr.columns]
        opr_cols_lower = {c.lower(): c for c in df_opr.columns}
        opr_rename = {}
        if 'ngayltc' in opr_cols_lower:
            opr_rename[opr_cols_lower['ngayltc']] = 'NgayLTC'
        if 'vol_ltc' in opr_cols_lower:
            opr_rename[opr_cols_lower['vol_ltc']] = 'vol_ltc'
        if 'ot' in opr_cols_lower:
            opr_rename[opr_cols_lower['ot']] = 'ot'
        if 'am' in opr_cols_lower:
            opr_rename[opr_cols_lower['am']] = 'AM'
        if 'ly_do_tre_12h' in opr_cols_lower:
            opr_rename[opr_cols_lower['ly_do_tre_12h']] = 'ly_do_tre_12h'
        if opr_rename:
            df_opr = df_opr.rename(columns=opr_rename)
            
        for col in ['NgayLTC', 'vol_ltc', 'ot', 'AM']:
            if col not in df_opr.columns:
                df_opr[col] = np.nan if col in ['NgayLTC', 'AM'] else 0.0

        # Normalize df_oe columns (from RAW n-1)
        df_oe.columns = [str(c).strip() for c in df_oe.columns]
        oe_cols_lower = {c.lower(): c for c in df_oe.columns}
        oe_rename = {}
        if 'bc lấy' in oe_cols_lower:
            oe_rename[oe_cols_lower['bc lấy']] = 'kholay'
        elif 'bc lay' in oe_cols_lower:
            oe_rename[oe_cols_lower['bc lay']] = 'kholay'
        elif 'kholay' in oe_cols_lower:
            oe_rename[oe_cols_lower['kholay']] = 'kholay'
            
        if 'thời gian tạo' in oe_cols_lower:
            oe_rename[oe_cols_lower['thời gian tạo']] = 'khung_gio_tao_don'
        elif 'thoi gian tao' in oe_cols_lower:
            oe_rename[oe_cols_lower['thoi gian tao']] = 'khung_gio_tao_don'
        elif 'khung_gio_tao_don' in oe_cols_lower:
            oe_rename[oe_cols_lower['khung_gio_tao_don']] = 'khung_gio_tao_don'
            
        if 'ly_do_tre_12h' in oe_cols_lower:
            oe_rename[oe_cols_lower['ly_do_tre_12h']] = 'ly_do_tre_12h'
            
        if 'sellername' in oe_cols_lower:
            oe_rename[oe_cols_lower['sellername']] = 'sellername'
            
        if 'madh' in oe_cols_lower:
            oe_rename[oe_cols_lower['madh']] = 'madh'
            
        if 'am' in oe_cols_lower:
            oe_rename[oe_cols_lower['am']] = 'AM'
            
        if 'vung' in oe_cols_lower:
            oe_rename[oe_cols_lower['vung']] = 'vung'
            
        if oe_rename:
            df_oe = df_oe.rename(columns=oe_rename)
            
        required_oe_cols = ['tutinh', 'kholay', 'sellername', 'khung_gio_tao_don', 'ly_do_tre_12h', 'madh', 'AM']
        for col in required_oe_cols:
            if col not in df_oe.columns:
                if col == 'tutinh' and 'vung' in df_oe.columns:
                    df_oe['tutinh'] = df_oe['vung']
                else:
                    df_oe[col] = np.nan

        # Normalize df_rawopr columns
        if df_rawopr is not None and not df_rawopr.empty:
            df_rawopr.columns = [str(c).strip() for c in df_rawopr.columns]
            rawopr_cols_lower = {c.lower(): c for c in df_rawopr.columns}
            rawopr_rename = {}
            if 'ngayltc' in rawopr_cols_lower:
                rawopr_rename[rawopr_cols_lower['ngayltc']] = 'NgayLTC'
            if 'vol_ltc' in rawopr_cols_lower:
                rawopr_rename[rawopr_cols_lower['vol_ltc']] = 'vol_ltc'
            if 'ot' in rawopr_cols_lower:
                rawopr_rename[rawopr_cols_lower['ot']] = 'ot'
            if 'am' in rawopr_cols_lower:
                rawopr_rename[rawopr_cols_lower['am']] = 'AM'
            if rawopr_rename:
                df_rawopr = df_rawopr.rename(columns=rawopr_rename)
                
            for col in ['NgayLTC', 'vol_ltc', 'ot', 'AM']:
                if col not in df_rawopr.columns:
                    df_rawopr[col] = np.nan if col in ['NgayLTC', 'AM'] else 0.0

        # Standardize AM spelling mismatches
        for df in [df_opr, df_oe, df_rawopr]:
            if df is not None and 'AM' in df.columns:
                df['AM'] = df['AM'].replace({'Huỳnh Tấn Hiền': 'Huỳnh Tấn Hiển'})

        # Filter OPR data to NTB region only
        if 'vung' in df_opr.columns:
            df_opr = df_opr[df_opr['vung'].astype(str).str.strip() == 'NTB'].copy()
        elif 'tutinh' in df_opr.columns:
            ntb_provinces = ['Bình Thuận', 'Ninh Thuận', 'Lâm Đồng', 'Khánh Hòa', 'Đắk Nông']
            df_opr = df_opr[df_opr['tutinh'].astype(str).str.strip().isin(ntb_provinces)].copy()

        # Build dynamic mapping of kholay -> tutinh from df_opr
        po_to_prov = {}
        if 'kholay' in df_opr.columns and 'tutinh' in df_opr.columns:
            tmp_df = df_opr.dropna(subset=['kholay', 'tutinh'])
            po_to_prov = dict(zip(tmp_df['kholay'].astype(str).str.strip(), tmp_df['tutinh'].astype(str).str.strip()))
            
        if df_oe is not None and not df_oe.empty:
            if 'vung' in df_oe.columns:
                df_oe = df_oe[df_oe['vung'].astype(str).str.strip() == 'NTB'].copy()
            elif 'tutinh' in df_oe.columns:
                ntb_provinces = ['Bình Thuận', 'Ninh Thuận', 'Lâm Đồng', 'Khánh Hòa', 'Đắk Nông']
                df_oe = df_oe[df_oe['tutinh'].astype(str).str.strip().isin(ntb_provinces)].copy()
            
            if 'kholay' in df_oe.columns:
                df_oe['tutinh_mapped'] = df_oe['kholay'].astype(str).str.strip().map(po_to_prov)
                if 'tutinh' in df_oe.columns:
                    df_oe['tutinh'] = df_oe['tutinh_mapped'].fillna(df_oe['tutinh'])
                else:
                    df_oe['tutinh'] = df_oe['tutinh_mapped']
                if 'tutinh_mapped' in df_oe.columns:
                    df_oe = df_oe.drop(columns=['tutinh_mapped'])

        if df_rawopr is not None and not df_rawopr.empty:
            if 'vung' in df_rawopr.columns:
                df_rawopr = df_rawopr[df_rawopr['vung'].astype(str).str.strip() == 'NTB'].copy()
            elif 'tutinh' in df_rawopr.columns:
                ntb_provinces = ['Bình Thuận', 'Ninh Thuận', 'Lâm Đồng', 'Khánh Hòa', 'Đắk Nông']
                df_rawopr = df_rawopr[df_rawopr['tutinh'].astype(str).str.strip().isin(ntb_provinces)].copy()

        # Apply active filters dynamically
        if am and am.strip():
            df_opr = df_opr[df_opr['AM'].astype(str).str.strip() == am.strip()]
            if df_oe is not None and not df_oe.empty:
                df_oe = df_oe[df_oe['AM'].astype(str).str.strip() == am.strip()]
            if df_rawopr is not None and not df_rawopr.empty:
                df_rawopr = df_rawopr[df_rawopr['AM'].astype(str).str.strip() == am.strip()]
                
        if province and province.strip():
            df_opr = df_opr[df_opr['tutinh'].astype(str).str.strip() == province.strip()]
            if df_oe is not None and not df_oe.empty:
                df_oe = df_oe[df_oe['tutinh'].astype(str).str.strip() == province.strip()]
            if df_rawopr is not None and not df_rawopr.empty:
                df_rawopr = df_rawopr[df_rawopr['tutinh'].astype(str).str.strip() == province.strip()]
                
        if post_office and post_office.strip():
            df_opr = df_opr[df_opr['kholay'].astype(str).str.strip() == post_office.strip()]
            if df_oe is not None and not df_oe.empty:
                df_oe = df_oe[df_oe['kholay'].astype(str).str.strip() == post_office.strip()]
            if df_rawopr is not None and not df_rawopr.empty:
                df_rawopr = df_rawopr[df_rawopr['kholay'].astype(str).str.strip() == post_office.strip()]

        df_opr = df_opr.dropna(subset=["vol_ltc"]).copy()
        
        # Calculate OPR Overall
        total_vol = float(df_opr['vol_ltc'].sum())
        total_ot = float(df_opr['ot'].sum())
        overall_opr = safe_divide(total_ot, total_vol)
        
        # Calculate OPR comparisons (n-1 and cùng kỳ thứ)
        opr_n1 = None
        opr_wk = None
        delta_n1 = None
        delta_wk = None
        
        if df_rawopr is not None and not df_rawopr.empty:
            try:
                # Normalize dates to YYYY-MM-DD
                df_opr_dates = pd.to_datetime(df_opr['NgayLTC'], errors='coerce').dt.strftime('%Y-%m-%d')
                df_rawopr_dates = pd.to_datetime(df_rawopr['NgayLTC'], errors='coerce').dt.strftime('%Y-%m-%d')
                
                valid_dates = df_opr_dates.dropna().unique()
                if len(valid_dates) > 0:
                    report_date_str = max(valid_dates)
                    report_dt = pd.to_datetime(report_date_str)
                    
                    n1_date_str = (report_dt - pd.Timedelta(days=1)).strftime('%Y-%m-%d')
                    wk_date_str = (report_dt - pd.Timedelta(days=7)).strftime('%Y-%m-%d')
                    
                    # Create temporary date columns
                    df_opr_temp = df_opr.copy()
                    df_rawopr_temp = df_rawopr.copy()
                    df_opr_temp['NgayLTC_str'] = df_opr_dates
                    df_rawopr_temp['NgayLTC_str'] = df_rawopr_dates
                    
                    cols = ['NgayLTC_str', 'vol_ltc', 'ot']
                    df_comb = pd.concat([df_opr_temp[cols], df_rawopr_temp[cols]], ignore_index=True)
                    df_comb = df_comb.dropna(subset=['NgayLTC_str', 'vol_ltc', 'ot'])
                    
                    def get_opr_for_date(df, date_str):
                        df_date = df[df['NgayLTC_str'] == date_str]
                        if len(df_date) == 0:
                            return None
                        t_vol = float(df_date['vol_ltc'].sum())
                        t_ot = float(df_date['ot'].sum())
                        return safe_divide(t_ot, t_vol)
                    
                    opr_n1 = get_opr_for_date(df_comb, n1_date_str)
                    opr_wk = get_opr_for_date(df_comb, wk_date_str)
                    
                    if opr_n1 is not None:
                        delta_n1 = round(overall_opr - opr_n1, 2)
                    if opr_wk is not None:
                        delta_wk = round(overall_opr - opr_wk, 2)
            except Exception as ex:
                print(f"Lỗi tính toán OPR so sánh: {ex}")
        
        # Calculate OPR by Creation Time Slot (Khung giờ tạo)
        time_slot_col = 'Khung giờ tạo' if 'Khung giờ tạo' in df_opr.columns else 'khung_gio_tao_don'
        slot_opr = df_opr.groupby(time_slot_col).agg({'vol_ltc': 'sum', 'ot': 'sum'}).reset_index()
        slot_opr['% OPR'] = (slot_opr['ot'] / slot_opr['vol_ltc']) * 100
        slot_opr_list = slot_opr.to_dict(orient='records')
        
        # Calculate Error Reasons (Sắp xếp theo tỷ trọng lỗi)
        reason_col = 'ly_do_tre_12h'
        df_late = df_opr[df_opr[reason_col] != '0.Ontime OPR']
        df_late = df_late.dropna(subset=[reason_col])
        
        # Late orders column
        if 'Đơn trễ' in df_late.columns:
            df_late['late_orders'] = df_late['Đơn trễ']
        else:
            df_late['late_orders'] = df_late['vol_ltc'] - df_late['ot']
            
        total_late = float(df_late['late_orders'].sum())
        
        errors = df_late.groupby(reason_col).agg({'late_orders': 'sum'}).reset_index()
        errors['weight'] = (errors['late_orders'] / total_late) * 100 if total_late > 0 else 0
        errors = errors.sort_values(by='late_orders', ascending=False)
        errors_list = errors.to_dict(orient='records')
        
        # Detail N-1 error orders
        df_oe = df_oe.dropna(subset=["madh"]).copy()
        
        # Clean details
        df_oe = df_oe.where(pd.notnull(df_oe), None)
        oe_details = df_oe[['tutinh', 'kholay', 'sellername', 'khung_gio_tao_don', 'ly_do_tre_12h', 'madh', 'AM']].to_dict(orient='records')
        
        # Calculate OPR performance by AM
        am_opr = df_opr.groupby('AM').agg({'vol_ltc': 'sum', 'ot': 'sum'}).reset_index()
        am_opr['% OPR'] = (am_opr['ot'] / am_opr['vol_ltc']) * 100
        am_opr = am_opr.sort_values(by='% OPR', ascending=True).to_dict(orient='records')
        
        return {
            "overall_opr": overall_opr,
            "total_volume": int(total_vol),
            "total_ontime": int(total_ot),
            "total_late": int(total_late),
            "time_slots": slot_opr_list,
            "error_reasons": errors_list,
            "am_performance": am_opr,
            "oe_details": oe_details,
            "opr_n1": opr_n1,
            "opr_wk": opr_wk,
            "delta_n1": delta_n1,
            "delta_wk": delta_wk
        }
    except Exception as e:
        return {"error": f"Lỗi xử lý file OPR: {str(e)}"}

# ==========================================
# 3. PROCESS BACKLOG AGING
# ==========================================
def process_aging_backlog(df_raw=None, df_co_cau=None):
    try:
        if df_raw is None:
            df_raw = safe_read_csv(resolve_path('aging_raw.csv', write=False))
        if df_co_cau is None:
            df_co_cau = safe_read_csv(resolve_path('ops_co_cau.csv', write=False))
            
        if df_raw is None or df_co_cau is None:
            return {"error": "Không tìm thấy dữ liệu aging (aging_raw.csv hoặc ops_co_cau.csv)."}
        
        df_raw = df_raw.dropna(subset=["Mã đơn"]).copy()
        df_co_cau = df_co_cau.dropna(subset=["Bưu cục"]).copy()
        
        df_raw = df_raw.dropna(subset=["Mã đơn"])
        df_co_cau = df_co_cau.dropna(subset=["Bưu cục"])
        
        # Map AM dynamically for rows where it is NaN
        # Clean warehouse strings for merge
        df_raw['bc_clean'] = df_raw['BC'].astype(str).str.strip().str.lower()
        df_co_cau['bc_clean'] = df_co_cau['Bưu cục'].astype(str).str.strip().str.lower()
        
        bc_to_am = dict(zip(df_co_cau['bc_clean'], df_co_cau['AM']))
        bc_to_province = dict(zip(df_co_cau['bc_clean'], df_co_cau['Tỉnh']))
        
        df_raw['mapped_am'] = df_raw['bc_clean'].map(bc_to_am)
        df_raw['mapped_province'] = df_raw['bc_clean'].map(bc_to_province)
        
        # Fallback to existing columns if map fails
        fallback_am = df_raw['am_name'] if 'am_name' in df_raw.columns else pd.Series("Không xác định", index=df_raw.index)
        df_raw['final_am'] = df_raw['mapped_am'].fillna(fallback_am).fillna("Không xác định")
        
        fallback_prov = df_raw['Tỉnh'] if 'Tỉnh' in df_raw.columns else pd.Series("Không xác định", index=df_raw.index)
        df_raw['final_province'] = df_raw['mapped_province'].fillna(fallback_prov).fillna("Không xác định")
        
        if 'Trạng thái' not in df_raw.columns:
            if 'Nhóm BL' in df_raw.columns:
                df_raw['Trạng thái'] = df_raw['Nhóm BL']
            else:
                df_raw['Trạng thái'] = "Chưa giao"
        
        # Categorize into aging brackets
        def get_aging_bracket(days):
            try:
                d = float(days)
                if d >= 5 and d < 8:
                    return "5 - 8 ngày"
                elif d >= 8 and d < 15:
                    return "8 - 15 ngày"
                elif d >= 15:
                    return "Trên 15 ngày"
            except:
                pass
            return "Khác"
            
        df_raw['aging_bracket'] = df_raw['Số ngày đã nhập BC'].apply(get_aging_bracket)
        df_raw = df_raw.where(pd.notnull(df_raw), None)
        
        # Summarize by AM and Bracket
        am_summary = df_raw.groupby(['final_am', 'aging_bracket']).size().unstack(fill_value=0).reset_index()
        # Ensure all columns exist
        for col in ["5 - 8 ngày", "8 - 15 ngày", "Trên 15 ngày"]:
            if col not in am_summary.columns:
                am_summary[col] = 0
        am_summary['Total'] = am_summary["5 - 8 ngày"] + am_summary["8 - 15 ngày"] + am_summary["Trên 15 ngày"]
        am_summary_list = am_summary.to_dict(orient='records')
        
        # Summarize by Post Office (BC)
        po_summary = df_raw.groupby(['BC', 'final_am', 'final_province', 'aging_bracket']).size().unstack(fill_value=0).reset_index()
        for col in ["5 - 8 ngày", "8 - 15 ngày", "Trên 15 ngày"]:
            if col not in po_summary.columns:
                po_summary[col] = 0
        po_summary['Total'] = po_summary["5 - 8 ngày"] + po_summary["8 - 15 ngày"] + po_summary["Trên 15 ngày"]
        po_summary_list = po_summary.to_dict(orient='records')
        
        # Raw records
        raw_records = df_raw[['final_province', 'BC', 'Mã đơn', 'Tệp khách', 'Số ngày đã nhập BC', 'Số lần giao', 'aging_bracket', 'final_am', 'Trạng thái']].to_dict(orient='records')
        
        return {
            "am_summary": am_summary_list,
            "po_summary": po_summary_list,
            "raw_records": raw_records,
            "total_backlog": len(df_raw)
        }
    except Exception as e:
        return {"error": f"Lỗi xử lý file aging backlog: {str(e)}"}

# ==========================================
# 4. PROCESS PENDING TRANSIT (TREO LUÂN CHUYỂN)
# ==========================================
def process_treo_backlog(df_raw=None, df_co_cau=None):
    try:
        if df_raw is None:
            df_raw = safe_read_csv(resolve_path('treo_stuck.csv', write=False))
        if df_co_cau is None:
            df_co_cau = safe_read_csv(resolve_path('ops_co_cau.csv', write=False))
            
        if df_raw is None or df_co_cau is None:
            return {"error": "Không tìm thấy dữ liệu treo luân chuyển (treo_stuck.csv hoặc ops_co_cau.csv)."}
                    
        df_raw = df_raw.dropna(subset=["Mã đơn hàng"]).copy()
        df_co_cau = df_co_cau.dropna(subset=["Bưu cục"]).copy()
        
        # Clean
        df_raw = df_raw.dropna(subset=["Mã đơn hàng"])
        df_co_cau = df_co_cau.dropna(subset=["Bưu cục"])
        
        # Apply filter: Trạng thái must be 'Chưa đóng kiện' or 'Không cần đóng kiện'
        allowed_statuses = ['Chưa đóng kiện', 'Không cần đóng kiện']
        df_raw = df_raw[df_raw['Trạng thái'].isin(allowed_statuses)]
        
        # Map AM dynamically via warehouse_name
        df_raw['wh_clean'] = df_raw['warehouse_name'].astype(str).str.strip().str.lower()
        df_co_cau['bc_clean'] = df_co_cau['Bưu cục'].astype(str).str.strip().str.lower()
        
        bc_to_am = dict(zip(df_co_cau['bc_clean'], df_co_cau['AM']))
        bc_to_province = dict(zip(df_co_cau['bc_clean'], df_co_cau['Tỉnh']))
        
        df_raw['mapped_am'] = df_raw['wh_clean'].map(bc_to_am)
        df_raw['mapped_province'] = df_raw['wh_clean'].map(bc_to_province)
        
        fallback_am = df_raw['am_name'] if 'am_name' in df_raw.columns else pd.Series(np.nan, index=df_raw.index)
        fallback_prov = df_raw['province_name'] if 'province_name' in df_raw.columns else pd.Series(np.nan, index=df_raw.index)
        df_raw['final_am'] = df_raw['mapped_am'].fillna(fallback_am).fillna("Không xác định")
        df_raw['final_province'] = df_raw['mapped_province'].fillna(fallback_prov).fillna("Không xác định")
        
        # Define delay brackets based on 'Thời gian tồn đọng'
        def get_treo_bracket(t):
            if not isinstance(t, str):
                return None
            t = t.strip()
            if t in ["36_48", "48_72"]:
                return "36 - 72h"
            elif t in ["72_96", "96_120"]:
                return "72 - 120h"
            elif t in ["120_192"]:
                return "120 - 192h"
            elif t in ["192", "192h+"]:
                return "192h+"
            return None

        df_raw['treo_bracket'] = df_raw['Thời gian tồn đọng'].apply(get_treo_bracket)
        df_raw = df_raw.dropna(subset=["treo_bracket"]) # Filter to delay > 36h
        
        df_raw = df_raw.where(pd.notnull(df_raw), None)
        
        # Split into Delivery (Giao) and Return (Trả)
        # Loại đơn has 'Luân chuyển giao' or 'Luân chuyển trả'
        df_giao = df_raw[df_raw['Loại đơn'] == 'Luân chuyển giao']
        df_tra = df_raw[df_raw['Loại đơn'] == 'Luân chuyển trả']
        
        # Helper to summarize df
        def get_summary_data(df_sub):
            if len(df_sub) == 0:
                return [], []
            
            # AM Summary
            am_sum = df_sub.groupby(['final_am', 'treo_bracket']).size().unstack(fill_value=0).reset_index()
            for col in ["36 - 72h", "72 - 120h", "120 - 192h", "192h+"]:
                if col not in am_sum.columns:
                    am_sum[col] = 0
            am_sum['Total'] = am_sum["36 - 72h"] + am_sum["72 - 120h"] + am_sum["120 - 192h"] + am_sum["192h+"]
            
            # PO Summary
            po_sum = df_sub.groupby(['warehouse_name', 'final_am', 'final_province', 'treo_bracket']).size().unstack(fill_value=0).reset_index()
            for col in ["36 - 72h", "72 - 120h", "120 - 192h", "192h+"]:
                if col not in po_sum.columns:
                    po_sum[col] = 0
            po_sum['Total'] = po_sum["36 - 72h"] + po_sum["72 - 120h"] + po_sum["120 - 192h"] + po_sum["192h+"]
            
            return am_sum.to_dict(orient='records'), po_sum.to_dict(orient='records')

        giao_am_sum, giao_po_sum = get_summary_data(df_giao)
        tra_am_sum, tra_po_sum = get_summary_data(df_tra)
        
        raw_records = df_raw[['final_province', 'warehouse_name', 'Mã đơn hàng', 'Loại đơn', 'Thời gian tồn đọng', 'treo_bracket', 'final_am', 'Trạng thái']].to_dict(orient='records')
        
        return {
            "giao_am_summary": giao_am_sum,
            "giao_po_summary": giao_po_sum,
            "tra_am_summary": tra_am_sum,
            "tra_po_summary": tra_po_sum,
            "raw_records": raw_records,
            "total_backlog": len(df_raw),
            "total_giao": len(df_giao),
            "total_tra": len(df_tra)
        }
    except Exception as e:
        return {"error": f"Lỗi xử lý file treo luân chuyển: {str(e)}"}

# ==========================================
# 4b. PROCESS UNSTABLE POST OFFICES
# ==========================================
def clean_po_name(name):
    if not name:
        return ""
    name = str(name).lower().strip()
    prefixes = ["bưu cục ", "kho giao hàng ", "kho giao ", "kho hàng ", "điểm xử lý ", "giao hàng nhanh ", "ghn "]
    for p in prefixes:
        if name.startswith(p):
            name = name[len(p):]
    replacements = {
        "khánh hoà": "khánh hòa",
        "đắc nông": "đắk nông",
        "bình phước": "lâm đồng",
        "-": " ",
        ".": " ",
        ",": " "
    }
    for k, v in replacements.items():
        name = name.replace(k, v)
    name = " ".join(name.split())
    return name

def map_po_to_am_prov(po_id, po_name, id_to_am, id_to_prov, name_to_am, name_to_prov, default_am="Không xác định", default_prov="Không xác định"):
    def clean_val(val, default):
        if pd.isna(val) or not str(val).strip() or str(val).lower() == 'nan':
            return default
        return str(val).strip()
    
    default_am = clean_val(default_am, "Không xác định")
    default_prov = clean_val(default_prov, "Không xác định")

    try:
        pid = int(float(po_id))
        if pid in id_to_am:
            return clean_val(id_to_am[pid], default_am), clean_val(id_to_prov[pid], default_prov)
    except:
        pass

    name_clean = clean_po_name(po_name)
    if name_clean in name_to_am:
        return clean_val(name_to_am[name_clean], default_am), clean_val(name_to_prov[name_clean], default_prov)

    for k in name_to_am:
        if k in name_clean or name_clean in k:
            return clean_val(name_to_am[k], default_am), clean_val(name_to_prov[k], default_prov)

    return default_am, default_prov

def process_unstable_po():
    file_path = resolve_path('buu_cuc_bat_on.csv', write=False)
    try:
        df_raw = safe_read_csv(file_path, header=None)
        if df_raw is None or df_raw.empty:
            return {"error": "Không tìm thấy dữ liệu bưu cục bất ổn (buu_cuc_bat_on.csv)."}
            
        update_time = None
        total_warning = None
        
        # Search metadata in the first 15 rows
        for r_idx in range(min(15, len(df_raw))):
            for c_idx in range(len(df_raw.columns)):
                cell_val = str(df_raw.iloc[r_idx, c_idx])
                if "Thời gian cập nhật" in cell_val:
                    for offset in range(1, 4):
                        if c_idx + offset < len(df_raw.columns):
                            val = df_raw.iloc[r_idx, c_idx + offset]
                            if pd.notna(val) and str(val).strip() != "":
                                update_time = str(val).strip()
                                break
                elif "bưu cục cảnh báo" in cell_val or "bưu cục bất ổn" in cell_val or "lượng bưu cục" in cell_val:
                    for offset in range(1, 4):
                        if c_idx + offset < len(df_raw.columns):
                            val = df_raw.iloc[r_idx, c_idx + offset]
                            if pd.notna(val) and str(val).strip() != "":
                                try:
                                    total_warning = int(float(val))
                                except:
                                    total_warning = str(val).strip()
                                break
        
        # Find the table headers row
        header_row_idx = None
        for r_idx in range(len(df_raw)):
            row_vals = [str(x).lower().strip() for x in df_raw.iloc[r_idx].values]
            if any("tổng số lượng" in x or "thời gian cập nhật" in x for x in row_vals):
                continue
            if any(x == "bưu cục" or x == "chi tiết" or x == "tên bc" or "tên bưu cục" in x or "kho_giao_id" in x or "kho_giao_name" in x or "tinh_giao" in x for x in row_vals):
                header_row_idx = r_idx
                break
                
        if header_row_idx is None:
            header_row_idx = 4 if len(df_raw) > 4 else 0
            
        # Load the table skipping rows before the header
        df_table = safe_read_csv(file_path, skiprows=header_row_idx)
        if df_table is None:
            return {"error": "Lỗi đọc bảng bưu cục bất ổn."}
        df_table.columns = [str(c).strip() for c in df_table.columns]
        
        # Remove completely empty rows
        id_col = next((c for c in df_table.columns if "id" in c.lower() or "kho_giao_id" in c.lower()), df_table.columns[0])
        name_col = next((c for c in df_table.columns if "name" in c.lower() or "bưu cục" in c.lower() or "kho_giao_name" in c.lower()), df_table.columns[1] if len(df_table.columns) > 1 else df_table.columns[0])
        
        df_table = df_table.dropna(subset=[id_col, name_col], how='all')
        df_table = df_table[df_table[id_col].astype(str).str.strip() != ""]
        
        # Ensure we have clean mappings from co_cau_ntb.csv
        id_to_am = {}
        id_to_prov = {}
        name_to_am = {}
        name_to_prov = {}
        co_cau_path = resolve_path('co_cau_ntb.csv', write=False)
        if os.path.exists(co_cau_path):
            try:
                df_cc = safe_read_csv(co_cau_path)
                for _, r in df_cc.iterrows():
                    bc_id = r.get('warehouse_id')
                    bc_name = str(r.get('Bưu cục', '')).strip()
                    am = str(r.get('AM', '')).strip()
                    prov = str(r.get('Tỉnh', '')).strip()
                    if prov == 'Khánh Hoà':
                        prov = 'Khánh Hòa'
                    if prov == 'Bình Phước':
                        prov = 'Lâm Đồng'
                    
                    if pd.notna(bc_id):
                        try:
                            id_to_am[int(bc_id)] = am
                            id_to_prov[int(bc_id)] = prov
                        except:
                            pass
                    if bc_name:
                        name_clean = clean_po_name(bc_name)
                        name_to_am[name_clean] = am
                        name_to_prov[name_clean] = prov
            except Exception as e:
                print(f"Error reading co_cau_ntb.csv in process_unstable_po: {e}")

        # Now map each row in the df_table
        processed_records = []
        for _, r in df_table.iterrows():
            po_id = r.get(id_col)
            po_name = r.get(name_col)
            
            # Map AM/Prov
            mapped_am, mapped_prov = map_po_to_am_prov(po_id, po_name, id_to_am, id_to_prov, name_to_am, name_to_prov, r.get('AM', 'Không xác định'), r.get('tinh_giao', 'Không xác định'))
            
            # Map Column R: du_kien_clear_ton (18th column)
            days_col = next((c for c in df_table.columns if "du_kien_clear_ton" in c.lower() or "clear_ton" in c.lower() or c == "du_kien_clear_ton"), None)
            days_val = 0
            if days_col and days_col in df_table.columns:
                days_val = r[days_col]
            elif len(df_table.columns) > 17:
                days_val = r.iloc[17]
                
            try:
                days_val = int(float(days_val)) if pd.notna(days_val) else 0
            except:
                days_val = 0
                
            # Map Reason
            reason_col = next((c for c in df_table.columns if "ly_do" in c.lower() or "reason" in c.lower() or "ly_do_bat_on" in c.lower()), None)
            reason_val = ""
            if reason_col:
                reason_val = str(r[reason_col]).strip() if pd.notna(r[reason_col]) else ""
            elif len(df_table.columns) > 18:
                reason_val = str(r.iloc[18]).strip() if pd.notna(r.iloc[18]) else ""
                
            # Map Status
            status_col = next((c for c in df_table.columns if "trạng thái" in c.lower() or "status" in c.lower() or "trang_thai" in c.lower()), None)
            status_val = "Bình thường"
            if status_col:
                status_val = str(r[status_col]).strip() if pd.notna(r[status_col]) else "Bình thường"
            elif len(df_table.columns) > 19:
                status_val = str(r.iloc[19]).strip() if pd.notna(r.iloc[19]) else "Bình thường"
                
            try:
                po_id_clean = int(float(po_id)) if pd.notna(po_id) else None
            except:
                # Skip header/footer description rows
                continue
                
            def safe_int(val):
                try:
                    return int(float(val)) if pd.notna(val) else 0
                except:
                    return 0

            record = {
                "id": po_id_clean,
                "name": str(po_name).strip() if pd.notna(po_name) else "",
                "am": mapped_am,
                "province": mapped_prov,
                "ton_lm": safe_int(r.get('BL LM', 0)),
                "ton_lm_5n": safe_int(r.get('BL LM >5 ngay', 0)),
                "pct_lm_5n": round(parse_unstable_pct(r.get('%BL LM >5 ngay', 0)), 2),
                "ton_ktc": safe_int(r.get('BL KTC', 0)),
                "ton_ktc_cung_tinh": safe_int(r.get('BL KTC cung tinh %', r.get('BL KTC cung tinh', 0))),
                "pct_ktc_cung_tinh": round(parse_unstable_pct(r.get('%BL KTC cung tinh', 0)), 2),
                "days_unstable": days_val,
                "reason": reason_val,
                "status": status_val
            }
            processed_records.append(record)
            
        # Group warning/unstable post offices by AM (only Bất ổn)
        unstable_by_am = {}
        for rec in processed_records:
            if rec["status"] == "Bất ổn":
                am_name = rec["am"]
                if am_name not in unstable_by_am:
                    unstable_by_am[am_name] = []
                unstable_by_am[am_name].append(rec["name"])
                
        # Sort AMs by the number of warning post offices (descending)
        am_deepdive = []
        for am, pos in unstable_by_am.items():
            am_deepdive.append({
                "am": am,
                "count": len(pos),
                "pos": pos
            })
        am_deepdive = sorted(am_deepdive, key=lambda x: x["count"], reverse=True)
        
        # Calculate actual total warnings (only Bất ổn)
        actual_total_warnings = sum(1 for rec in processed_records if rec["status"] == "Bất ổn")
        
        return {
            "update_time": update_time,
            "total_warning": actual_total_warnings,
            "records": clean_nan(processed_records),
            "am_deepdive": clean_nan(am_deepdive)
        }
    except Exception as e:
        import traceback
        traceback.print_exc()
        return {"error": f"Lỗi xử lý bưu cục bất ổn: {str(e)}"}

def process_off_spe():
    file_path = resolve_path('off_tuyen_spe.csv', write=False)
    try:
        df_raw = safe_read_csv(file_path)
        if df_raw is None or df_raw.empty:
            return {"error": "Không tìm thấy dữ liệu OFF tuyến SPE (off_tuyen_spe.csv)."}
        
        # Remove empty rows or rows that have all NaN
        df_raw = df_raw.dropna(how='all')
        
        # Find columns dynamically
        col_tinh = next((c for c in df_raw.columns if "tỉnh" in c.lower() or "tinh" in c.lower() or "province" in c.lower()), None)
        col_quan = next((c for c in df_raw.columns if "quận" in c.lower() or "huyện" in c.lower() or "quan" in c.lower() or "huyen" in c.lower() or "district" in c.lower()), None)
        col_phuong = next((c for c in df_raw.columns if "phường" in c.lower() or "xã" in c.lower() or "phuong" in c.lower() or "xa" in c.lower() or "ward" in c.lower()), None)
        col_bc = next((c for c in df_raw.columns if "bưu cục" in c.lower() or "buu cuc" in c.lower() or "post" in c.lower() or "bc" in c.lower()), None)
        col_ketqua = next((c for c in df_raw.columns if "kết quả" in c.lower() or "ket qua" in c.lower() or "update" in c.lower()), None)
        col_capdown = next((c for c in df_raw.columns if "cap" in c.lower() or "down" in c.lower() or "%" in c.lower()), None)
        col_time_off = next((c for c in df_raw.columns if "thời gian tắt" in c.lower() or "thoi gian tat" in c.lower() or "tg tắt" in c.lower() or "tg tat" in c.lower()), None)
        col_time_on = next((c for c in df_raw.columns if "thời gian mở" in c.lower() or "thoi gian mo" in c.lower() or "tg mở" in c.lower() or "tg mo" in c.lower()), None)
        col_note = next((c for c in df_raw.columns if "note" in c.lower() or "ghi chú" in c.lower() or "ghi chu" in c.lower()), None)
        
        # Fallback to indices if columns not found dynamically
        cols = list(df_raw.columns)
        if not col_tinh and len(cols) > 0: col_tinh = cols[0]
        if not col_quan and len(cols) > 1: col_quan = cols[1]
        if not col_phuong and len(cols) > 2: col_phuong = cols[2]
        if not col_bc and len(cols) > 3: col_bc = cols[3]
        if not col_ketqua and len(cols) > 4: col_ketqua = cols[4]
        if not col_capdown and len(cols) > 5: col_capdown = cols[5]
        if not col_time_off and len(cols) > 6: col_time_off = cols[6]
        if not col_time_on and len(cols) > 7: col_time_on = cols[7]
        if not col_note and len(cols) > 8: col_note = cols[8]
        
        df_table = df_raw.dropna(subset=[col_tinh, col_bc], how='all')
        
        processed_records = []
        total_off = 0
        total_pending = 0
        
        # Map standard dates if they look like timestamp objects
        def clean_date_str(val):
            if pd.isna(val):
                return ""
            if isinstance(val, (datetime.datetime, pd.Timestamp)):
                return val.strftime("%d-%m-%Y")
            val_str = str(val).strip()
            # If timestamp string, format it
            if "00:00:00" in val_str:
                val_str = val_str.replace(" 00:00:00", "")
            return val_str
            
        for _, r in df_table.iterrows():
            tinh_val = str(r[col_tinh]).strip() if col_tinh and pd.notna(r[col_tinh]) else "Không xác định"
            quan_val = str(r[col_quan]).strip() if col_quan and pd.notna(r[col_quan]) else ""
            phuong_val = str(r[col_phuong]).strip() if col_phuong and pd.notna(r[col_phuong]) else ""
            bc_val = str(r[col_bc]).strip() if col_bc and pd.notna(r[col_bc]) else ""
            
            # Status:
            # - Column E has "duyệt" => Đang OFF
            # - Column E empty => Đang chờ duyệt
            kq_raw = r[col_ketqua] if col_ketqua else ""
            kq_val = str(kq_raw).strip().lower() if pd.notna(kq_raw) else ""
            
            if pd.isna(kq_raw) or not kq_val:
                status_val = "Đang chờ duyệt"
                total_pending += 1
            elif "duyệt" in kq_val:
                status_val = "Đang OFF"
                total_off += 1
            else:
                status_val = "Đang chờ duyệt"
                total_pending += 1
                
            capdown_val = ""
            if col_capdown and pd.notna(r[col_capdown]):
                try:
                    val_float = float(r[col_capdown])
                    if val_float <= 1.0:
                        capdown_val = f"{int(val_float * 100)}%"
                    else:
                        capdown_val = f"{int(val_float)}%"
                except:
                    capdown_val = str(r[col_capdown]).strip()
                    
            time_off_val = clean_date_str(r[col_time_off]) if col_time_off else ""
            time_on_val = clean_date_str(r[col_time_on]) if col_time_on else ""
            note_val = str(r[col_note]).strip() if col_note and pd.notna(r[col_note]) else ""
            
            processed_records.append({
                "province": tinh_val,
                "district": quan_val,
                "ward": phuong_val,
                "post_office": bc_val,
                "status": status_val,
                "pct_cap_down": capdown_val,
                "off_time": time_off_val,
                "on_time": time_on_val,
                "note": note_val
            })
            
        if os.path.exists(file_path):
            mtime = os.path.getmtime(file_path)
            update_time = datetime.datetime.fromtimestamp(mtime).strftime("%d-%m-%Y %H:%M:%S")
        else:
            update_time = datetime.datetime.now().strftime("%d-%m-%Y %H:%M:%S")
        
        return {
            "update_time": update_time,
            "total_off": total_off,
            "total_pending": total_pending,
            "records": clean_nan(processed_records)
        }
    except Exception as e:
        import traceback
        traceback.print_exc()
        return {"error": f"Lỗi xử lý file off_tuyen_spe.xlsx: {str(e)}"}

# ==========================================
# 5. SYNC & HISTORY CACHE MANAGEMENT
# ==========================================
def load_history():
    db_history = load_json_from_db('backlog_history.json')
    if db_history is not None:
        return db_history
        
    history_file = resolve_path('backlog_history.json', write=False)
    if not os.path.exists(history_file):
        return []
    try:
        with open(history_file, 'r', encoding='utf-8') as f:
            return json.load(f)
    except:
        return []

def save_history(history):
    save_json_to_db(history, 'backlog_history.json')
    try:
        history_file = resolve_path('backlog_history.json', write=True)
        with open(history_file, 'w', encoding='utf-8') as f:
            json.dump(history, f, indent=4, ensure_ascii=False)
    except Exception as e:
        print(f"Error saving history: {e}")

# Save current run to history
def add_to_history(aging_data, treo_data):
    history = load_history()
    
    # Format date and time
    now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    # Save simple metrics map for comparison
    aging_am_map = {}
    for r in aging_data.get("am_summary", []):
        aging_am_map[r["final_am"]] = int(r["Total"])
        
    aging_po_map = {}
    for r in aging_data.get("po_summary", []):
        aging_po_map[r["BC"]] = int(r["Total"])
        
    treo_giao_am_map = {}
    for r in treo_data.get("giao_am_summary", []):
        treo_giao_am_map[r["final_am"]] = int(r["Total"])
        
    treo_giao_po_map = {}
    for r in treo_data.get("giao_po_summary", []):
        treo_giao_po_map[r["warehouse_name"]] = int(r["Total"])
        
    treo_tra_am_map = {}
    for r in treo_data.get("tra_am_summary", []):
        treo_tra_am_map[r["final_am"]] = int(r["Total"])
        
    treo_tra_po_map = {}
    for r in treo_data.get("tra_po_summary", []):
        treo_tra_po_map[r["warehouse_name"]] = int(r["Total"])
        
    new_entry = {
        "timestamp": now_str,
        "aging_total": aging_data.get("total_backlog", 0),
        "treo_total": treo_data.get("total_backlog", 0),
        "treo_giao_total": treo_data.get("total_giao", 0),
        "treo_tra_total": treo_data.get("total_tra", 0),
        "aging_am_map": aging_am_map,
        "aging_po_map": aging_po_map,
        "treo_giao_am_map": treo_giao_am_map,
        "treo_giao_po_map": treo_giao_po_map,
        "treo_tra_am_map": treo_tra_am_map,
        "treo_tra_po_map": treo_tra_po_map
    }
    
    history.append(new_entry)
    # Keep last 20 entries
    if len(history) > 20:
        history = history[-20:]
    save_history(history)
    return now_str

# Helper to calculate chênh lệch (+/-) based on history entry
def calculate_trend(current_data, baseline_entry):
    if not baseline_entry:
        return current_data
        
    # Process Aging Backlog
    aging_am_map = baseline_entry.get("aging_am_map", {})
    aging_po_map = baseline_entry.get("aging_po_map", {})
    
    for r in current_data["aging"].get("am_summary", []):
        prev = aging_am_map.get(r["final_am"], 0)
        r["diff"] = int(r["Total"]) - prev
        
    for r in current_data["aging"].get("po_summary", []):
        prev = aging_po_map.get(r["BC"], 0)
        r["diff"] = int(r["Total"]) - prev
        
    current_data["aging"]["diff_total"] = current_data["aging"].get("total_backlog", 0) - baseline_entry.get("aging_total", 0)
    
    # Process Treo luân chuyển
    tg_am_map = baseline_entry.get("treo_giao_am_map", {})
    tg_po_map = baseline_entry.get("treo_giao_po_map", {})
    tt_am_map = baseline_entry.get("treo_tra_am_map", {})
    tt_po_map = baseline_entry.get("treo_tra_po_map", {})
    
    for r in current_data["treo"].get("giao_am_summary", []):
        prev = tg_am_map.get(r["final_am"], 0)
        r["diff"] = int(r["Total"]) - prev
        
    for r in current_data["treo"].get("giao_po_summary", []):
        prev = tg_po_map.get(r["warehouse_name"], 0)
        r["diff"] = int(r["Total"]) - prev
        
    for r in current_data["treo"].get("tra_am_summary", []):
        prev = tt_am_map.get(r["final_am"], 0)
        r["diff"] = int(r["Total"]) - prev
        
    for r in current_data["treo"].get("tra_po_summary", []):
        prev = tt_po_map.get(r["warehouse_name"], 0)
        r["diff"] = int(r["Total"]) - prev
        
    current_data["treo"]["diff_total"] = current_data["treo"].get("total_backlog", 0) - baseline_entry.get("treo_total", 0)
    current_data["treo"]["diff_giao"] = current_data["treo"].get("total_giao", 0) - baseline_entry.get("treo_giao_total", 0)
    current_data["treo"]["diff_tra"] = current_data["treo"].get("total_tra", 0) - baseline_entry.get("treo_tra_total", 0)
    
    return current_data

# ==========================================
# 6. IN-MEMORY CACHE DECLARATION & INITIALIZATION
# ==========================================
OPERATIONAL_CACHE = None
OPR_CACHE = None
BACKLOG_CACHE_RAW = None
UNSTABLE_PO_CACHE = None
OFF_SPE_CACHE = None

# Dataframe caches for NTB Summary Dashboard
DF_GTC_CACHE = None
DF_LTC_CACHE = None
DF_CO_CAU_CACHE = None
DF_AGING_CACHE = None
DF_TREO_CACHE = None
DF_TTS_CACHE = None
DF_TAO_DON_CACHE = None
DF_BUU_CUC_TYPE_MAP = None

def load_buu_cuc_type_map():
    file_path = resolve_path('buu_cuc_bat_on.csv', write=False)
    try:
        df_raw = safe_read_csv(file_path, header=None)
        if df_raw is None or df_raw.empty:
            return {}
        header_row_idx = None
        for r_idx in range(len(df_raw)):
            row_vals = [str(x).lower().strip() for x in df_raw.iloc[r_idx].values]
            if any("tổng số lượng" in x or "thời gian cập nhật" in x for x in row_vals):
                continue
            if any(x == "bưu cục" or x == "chi tiết" or x == "tên bc" or "tên bưu cục" in x or "kho_giao_id" in x or "kho_giao_name" in x or "tinh_giao" in x for x in row_vals):
                header_row_idx = r_idx
                break
        if header_row_idx is None:
            header_row_idx = 4 if len(df_raw) > 4 else 0
        df_table = safe_read_csv(file_path, skiprows=header_row_idx)
        if df_table is None or df_table.empty:
            return {}
        df_table.columns = [str(c).strip() for c in df_table.columns]
        
        id_col = next((c for c in df_table.columns if "id" in c.lower() or "kho_giao_id" in c.lower()), None)
        type_col = next((c for c in df_table.columns if "type" in c.lower() or "warehouse_type" in c.lower() or "loại" in c.lower()), None)
        
        if id_col and type_col:
            df_table = df_table.dropna(subset=[id_col, type_col])
            type_map = {}
            for _, row in df_table.iterrows():
                try:
                    p_id = int(float(row[id_col]))
                    p_type = str(row[type_col]).strip()
                    type_map[p_id] = p_type
                except:
                    pass
            return type_map
    except Exception as e:
        print(f"Error loading buu_cuc_type_map: {e}")
    return {}

def load_vols_tao_don_df():
    file_path = resolve_path('vols_tao_don.csv', write=False)
    try:
        df = safe_read_csv(file_path)
        if df is None or df.empty:
            return None
        df.columns = [str(c).strip() for c in df.columns]
        df['Date'] = pd.to_datetime(df['Date'])
        df = df[df['bat_on'].fillna('').str.strip() != 'BC Cũ/Không thuộc ĐCL'].copy()
        return df
    except Exception as e:
        print(f"Error loading vols_tao_don.csv: {e}")
    return None

def clean_str_key(val):
    if pd.isna(val):
        return "unknown"
    val = str(val).strip().lower()
    replacements = {
        "lâm đồng": "lam_dong",
        "bình thuận": "binh_thuan",
        "khánh hòa": "khanh_hoa",
        "ninh thuận": "ninh_thuan",
        "đắk nông": "dak_nong",
        "đắc nông": "dak_nong"
    }
    return replacements.get(val, val.replace(" ", "_"))

def get_dataframes(force=False, raw_gtc=None, raw_ltc=None, raw_co_cau=None, raw_aging=None, raw_co_cau_aging=None, raw_treo=None, raw_co_cau_treo=None, raw_tts=None):
    global DF_GTC_CACHE, DF_LTC_CACHE, DF_CO_CAU_CACHE, DF_AGING_CACHE, DF_TREO_CACHE, DF_TTS_CACHE
    import gc
    if force or DF_GTC_CACHE is None or DF_LTC_CACHE is None or DF_CO_CAU_CACHE is None or DF_AGING_CACHE is None or DF_TREO_CACHE is None or DF_TTS_CACHE is None:
        if not force:
            print("Cache is empty. Initializing cache synchronously...")
            try:
                update_all_caches()
            except Exception as e:
                print(f"Error loading cache synchronously: {e}")
                raise RuntimeError("Dữ liệu đang được tải vào bộ nhớ đệm hoặc chưa được đồng bộ. Vui lòng đợi vài giây hoặc nhấn nút Đồng bộ dữ liệu.")
            if DF_GTC_CACHE is None or DF_LTC_CACHE is None or DF_CO_CAU_CACHE is None or DF_AGING_CACHE is None or DF_TREO_CACHE is None or DF_TTS_CACHE is None:
                raise RuntimeError("Dữ liệu đang được tải vào bộ nhớ đệm hoặc chưa được đồng bộ. Vui lòng đợi vài giây hoặc nhấn nút Đồng bộ dữ liệu.")
        bc_to_am = {}
        bc_to_prov = {}
        
        # 1. Load and process Báo cáo Vận hành
        if raw_gtc is None:
            raw_gtc = safe_read_csv(resolve_path('ops_gtc.csv', write=False))
        raw_gtc = clean_ops_df(raw_gtc, "gtc")
        
        if raw_ltc is None:
            raw_ltc = safe_read_csv(resolve_path('ops_ltc.csv', write=False))
        raw_ltc = clean_ops_df(raw_ltc, "ltc")
        
        if raw_co_cau is None:
            raw_co_cau = safe_read_csv(resolve_path('ops_co_cau.csv', write=False))

        if raw_tts is None:
            raw_tts = safe_read_csv(resolve_path('ops_tts.csv', write=False))
        raw_tts = clean_ops_df(raw_tts, "tts")
            
        if raw_gtc is None or raw_ltc is None or raw_co_cau is None:
            raise FileNotFoundError("Không tìm thấy dữ liệu vận hành CSV (ops_gtc.csv, ops_ltc.csv hoặc ops_co_cau.csv).")
            
        df_gtc = raw_gtc[raw_gtc['Cấp Quản Lý'] != 'Grand Total'].dropna(subset=["Volume"]).copy()
        df_gtc['Leadtime'] = pd.to_numeric(df_gtc['Leadtime'], errors='coerce')
        
        df_ltc = raw_ltc[raw_ltc['Cấp quản lý'] != 'Grand Total'].dropna(subset=["Volume"]).copy()
        df_ltc['Leadtime'] = pd.to_numeric(df_ltc['Leadtime'], errors='coerce')
        
        df_tts = None
        if raw_tts is not None:
            df_tts = raw_tts.dropna(subset=["Volume"]).copy()
            if 'Cấp Quản Lý' in df_tts.columns:
                df_tts = df_tts[df_tts['Cấp Quản Lý'] != 'Grand Total'].copy()
            leadtime_col = next((c for c in df_tts.columns if c.lower() == 'leadtime'), 'Leadtime')
            df_tts['Leadtime'] = pd.to_numeric(df_tts.get(leadtime_col, 0), errors='coerce')
            
        df_co_cau = raw_co_cau.copy()
        
        # Map co_cau
        for _, r in df_co_cau.iterrows():
            bc = str(r.get('BC', '')).strip().lower()
            buucuc = str(r.get('Bưu cục', '')).strip().lower()
            am = str(r.get('AM', r.get('Am', r.get('ID - Họ Tên Am', '')))).strip()
            prov = str(r.get('Tỉnh', '')).strip()
            if prov == 'Bình Phước':
                prov = 'Lâm Đồng'
            if bc and bc != 'nan':
                bc_to_am[bc] = am
                bc_to_prov[bc] = prov
            if buucuc and buucuc != 'nan':
                bc_to_am[buucuc] = am
                bc_to_prov[buucuc] = prov
                
        # Delete raw and clean memory
        del raw_gtc, raw_ltc, raw_co_cau
        if raw_tts is not None:
            del raw_tts
        gc.collect()
        
        # 2. Load and process Aging
        if raw_aging is None:
            raw_aging = safe_read_csv(resolve_path('aging_raw.csv', write=False))
        if raw_co_cau_aging is None:
            raw_co_cau_aging = safe_read_csv(resolve_path('ops_co_cau.csv', write=False))
            
        if raw_aging is None or raw_co_cau_aging is None:
            raise FileNotFoundError("Không tìm thấy dữ liệu aging CSV (aging_raw.csv hoặc ops_co_cau.csv).")
            
        df_aging = raw_aging.copy()
        if 'am_name' not in df_aging.columns:
            df_aging['am_name'] = np.nan
        if 'Trạng thái' not in df_aging.columns:
            df_aging['Trạng thái'] = np.nan
        df_co_cau_aging = raw_co_cau_aging.copy()
        
        for _, r in df_co_cau_aging.iterrows():
            bc = str(r.get('BC', '')).strip().lower()
            buucuc = str(r.get('Bưu cục', '')).strip().lower()
            am = str(r.get('AM', r.get('Am', r.get('ID - Họ Tên Am', '')))).strip()
            prov = str(r.get('Tỉnh', '')).strip()
            if prov == 'Bình Phước':
                prov = 'Lâm Đồng'
            if bc and bc != 'nan':
                bc_to_am[bc] = am
                bc_to_prov[bc] = prov
            if buucuc and buucuc != 'nan':
                bc_to_am[buucuc] = am
                bc_to_prov[buucuc] = prov
                
        del raw_aging, raw_co_cau_aging
        gc.collect()
        
        # 3. Load and process Treo
        if raw_treo is None:
            raw_treo = safe_read_csv(resolve_path('treo_stuck.csv', write=False))
        if raw_co_cau_treo is None:
            raw_co_cau_treo = safe_read_csv(resolve_path('ops_co_cau.csv', write=False))
            
        if raw_treo is None or raw_co_cau_treo is None:
            raise FileNotFoundError("Không tìm thấy dữ liệu treo luân chuyển CSV (treo_stuck.csv hoặc ops_co_cau.csv).")
            
        df_treo = raw_treo.copy()
        if 'am_name' not in df_treo.columns:
            df_treo['am_name'] = np.nan
        if 'province_name' not in df_treo.columns:
            df_treo['province_name'] = np.nan
        df_co_cau_treo = raw_co_cau_treo.copy()
        
        for _, r in df_co_cau_treo.iterrows():
            bc = str(r.get('BC', '')).strip().lower()
            buucuc = str(r.get('Bưu cục', '')).strip().lower()
            am = str(r.get('AM', r.get('Am', r.get('ID - Họ Tên Am', '')))).strip()
            prov = str(r.get('Tỉnh', '')).strip()
            if prov == 'Bình Phước':
                prov = 'Lâm Đồng'
            if bc and bc != 'nan':
                bc_to_am[bc] = am
                bc_to_prov[bc] = prov
            if buucuc and buucuc != 'nan':
                bc_to_am[buucuc] = am
                bc_to_prov[buucuc] = prov
                
        del raw_treo, raw_co_cau_treo
        gc.collect()
        
        # 4. Final Processing & Mapping
        df_gtc['clean_bc'] = df_gtc['Chi tiết'].apply(clean_str)
        df_gtc['mapped_prov'] = df_gtc['clean_bc'].map(bc_to_prov).fillna(df_gtc['Tỉnh']).fillna("Không xác định")
        df_gtc['mapped_am'] = df_gtc['clean_bc'].map(bc_to_am).fillna(df_gtc['AM']).fillna("Không xác định")
        
        df_gtc['Volume'] = safe_to_numeric(df_gtc['Volume'])
        df_gtc['% GTC'] = normalize_pct_col(df_gtc['% GTC'])
        df_gtc['% Chuyển trả'] = normalize_pct_col(df_gtc['% Chuyển trả'])
        
        df_gtc['delivered_vol'] = df_gtc['Volume'] * df_gtc['% GTC']
        df_gtc['return_vol'] = df_gtc['Volume'] * df_gtc['% Chuyển trả']
        df_gtc['ttc_vol'] = df_gtc['delivered_vol'] + df_gtc['return_vol']
        
        df_ltc['clean_bc'] = df_ltc['Chi tiết'].apply(clean_str)
        prov_col = next((c for c in df_ltc.columns if c.lower() == 'tỉnh'), None)
        am_col = next((c for c in df_ltc.columns if c.lower() == 'am'), None)
        
        df_ltc['mapped_prov'] = df_ltc['clean_bc'].map(bc_to_prov)
        if prov_col:
            df_ltc['mapped_prov'] = df_ltc['mapped_prov'].fillna(df_ltc[prov_col])
        df_ltc['mapped_prov'] = df_ltc['mapped_prov'].fillna("Không xác định")
        
        df_ltc['mapped_am'] = df_ltc['clean_bc'].map(bc_to_am)
        if am_col:
            df_ltc['mapped_am'] = df_ltc['mapped_am'].fillna(df_ltc[am_col])
        df_ltc['mapped_am'] = df_ltc['mapped_am'].fillna("Không xác định")
        
        # Normalize columns and map ltc_vol to Column K (Sản Lượng Lấy Thành Công)
        df_ltc['Volume'] = safe_to_numeric(df_ltc['Volume'])
        df_ltc['%LTC'] = normalize_pct_col(df_ltc['%LTC'])
        df_ltc.columns = [c.strip() for c in df_ltc.columns]
        if 'Sản Lượng Lấy Thành Công' in df_ltc.columns:
            df_ltc['ltc_vol'] = safe_to_numeric(df_ltc['Sản Lượng Lấy Thành Công'])
        else:
            df_ltc['ltc_vol'] = df_ltc['Volume'] * df_ltc['%LTC']
            
        # Align dates: ensure df_ltc contains rows for any dates present in df_gtc to avoid null/zero reporting
        if len(df_gtc) > 0 and len(df_ltc) > 0:
            gtc_dates = df_gtc['Time'].dropna().unique()
            ltc_dates = df_ltc['Time'].dropna().unique()
            missing_ltc_dates = [d for d in gtc_dates if d not in ltc_dates]
            if missing_ltc_dates and len(ltc_dates) > 0:
                try:
                    ltc_dates_sorted = sorted(ltc_dates, key=lambda x: pd.to_datetime(str(x).split(' - ')[0]))
                    latest_ltc_date = ltc_dates_sorted[-1]
                except Exception as e:
                    latest_ltc_date = ltc_dates[-1]
                
                latest_ltc_rows = df_ltc[df_ltc['Time'] == latest_ltc_date].copy()
                new_rows_list = []
                for missing_date in missing_ltc_dates:
                    copied = latest_ltc_rows.copy()
                    copied['Time'] = missing_date
                    new_rows_list.append(copied)
                if new_rows_list:
                    df_ltc = pd.concat([df_ltc] + new_rows_list, ignore_index=True)
        
        df_aging['clean_bc'] = df_aging['BC'].apply(clean_str)
        df_aging['mapped_prov'] = df_aging['clean_bc'].map(bc_to_prov).fillna(df_aging['Tỉnh']).fillna("Không xác định")
        df_aging['mapped_am'] = df_aging['clean_bc'].map(bc_to_am).fillna(df_aging['am_name']).fillna("Không xác định")
        
        df_treo['clean_bc'] = df_treo['warehouse_name'].apply(clean_str)
        df_treo['mapped_prov'] = df_treo['clean_bc'].map(bc_to_prov).fillna(df_treo['province_name']).fillna("Không xác định")
        df_treo['mapped_am'] = df_treo['clean_bc'].map(bc_to_am).fillna(df_treo['am_name']).fillna("Không xác định")
        
        allowed_statuses = ['Chưa đóng kiện', 'Không cần đóng kiện']
        df_treo_filtered = df_treo[df_treo['Trạng thái'].isin(allowed_statuses)].copy()
        
        def get_treo_bracket_local(t):
            if not isinstance(t, str):
                return None
            t = t.strip()
            if t in ["36_48", "48_72"]:
                return "36 - 72h"
            elif t in ["72_96", "96_120"]:
                return "72 - 120h"
            elif t in ["120_192"]:
                return "120 - 192h"
            elif t in ["192", "192h+"]:
                return "192h+"
            return None

        df_treo_filtered['treo_bracket'] = df_treo_filtered['Thời gian tồn đọng'].apply(get_treo_bracket_local)
        df_treo_filtered = df_treo_filtered.dropna(subset=["treo_bracket"])
        
        DF_GTC_CACHE = df_gtc
        DF_LTC_CACHE = df_ltc
        DF_CO_CAU_CACHE = df_co_cau
        DF_AGING_CACHE = df_aging
        DF_TREO_CACHE = df_treo_filtered
        DF_TTS_CACHE = df_tts
        
        gc.collect()
        
    return DF_GTC_CACHE, DF_LTC_CACHE, DF_AGING_CACHE, DF_TREO_CACHE

def apply_filters(df, am=None, province=None, post_office=None):
    df_filtered = df.copy()
    if am:
        df_filtered = df_filtered[df_filtered['mapped_am'] == am]
    if province:
        df_filtered = df_filtered[df_filtered['mapped_prov'] == province]
    if post_office:
        df_filtered = df_filtered[df_filtered['clean_bc'] == post_office.strip().lower()]
    return df_filtered

def update_all_caches():
    global OPERATIONAL_CACHE, OPR_CACHE, BACKLOG_CACHE_RAW, UNSTABLE_PO_CACHE, OFF_SPE_CACHE, DF_TAO_DON_CACHE, DF_BUU_CUC_TYPE_MAP
    import gc
    
    print("--------------------------------------------------")
    print("STARTING MEMORY-EFFICIENT CACHE LOAD...")
    print("--------------------------------------------------")
    
    # 1. First trigger get_dataframes(force=True) to build DF_GTC_CACHE, DF_LTC_CACHE, DF_AGING_CACHE, DF_TREO_CACHE
    # sequentially and efficiently!
    try:
        get_dataframes(force=True)
        print("-> Pandas DataFrames loaded sequentially.")
    except Exception as e:
        err_msg = f"Loi doc file Excel: {str(e)}"
        print(err_msg)
        OPERATIONAL_CACHE = {"error": err_msg}
        OPR_CACHE = {"error": err_msg}
        BACKLOG_CACHE_RAW = {"aging": {"error": err_msg}, "treo": {"error": err_msg}}
        UNSTABLE_PO_CACHE = {"error": err_msg}
        OFF_SPE_CACHE = {"error": err_msg}
        return

    # 2. Now process operational report using cached DataFrames
    try:
        print("Parsing Operational report...")
        # Note: process_operational_report reads file_path inside if df_gtc is None
        # We pass the cached dataframes so it doesn't need to load the file again!
        ops = process_operational_report(df_gtc=DF_GTC_CACHE, df_ltc=DF_LTC_CACHE, df_tts=DF_TTS_CACHE)
        if "error" not in ops:
            OPERATIONAL_CACHE = ops
            print("-> Operational report processed and cached.")
        else:
            OPERATIONAL_CACHE = ops
            print("-> Error processing operational report.")
    except Exception as e:
        OPERATIONAL_CACHE = {"error": f"Lỗi xử lý báo cáo vận hành: {e}"}
        

    # 3. Process OPR report
    try:
        print("Parsing OPR report...")
        raw_opr = safe_read_csv(resolve_path('opr_opr.csv', write=False))
        raw_oe = safe_read_csv(resolve_path('opr_oe.csv', write=False))
        raw_rawopr = safe_read_csv(resolve_path('opr_raw.csv', write=False))
        if raw_opr is not None and raw_oe is not None:
            opr = process_opr_report(df_opr=raw_opr, df_oe=raw_oe, df_rawopr=raw_rawopr)
            OPR_CACHE = opr
            del raw_opr, raw_oe, raw_rawopr
            print("-> OPR report processed and cached.")
        else:
            OPR_CACHE = {"error": "Không tìm thấy dữ liệu OPR CSV."}
        gc.collect()
    except Exception as e:
        OPR_CACHE = {"error": f"Lỗi xử lý OPR: {e}"}
        
    # 4. Process Backlog reports (Aging and Treo)
    try:
        print("Parsing Backlog reports...")
        # Fallback to ops_co_cau.csv automatically by passing df_co_cau=None
        aging = process_aging_backlog(df_raw=DF_AGING_CACHE, df_co_cau=None)
        treo = process_treo_backlog(df_raw=DF_TREO_CACHE, df_co_cau=None)
        
        BACKLOG_CACHE_RAW = {
            "aging": aging,
            "treo": treo
        }
        print("-> Backlog reports processed and cached.")
        gc.collect()
    except Exception as e:
        BACKLOG_CACHE_RAW = {"aging": {"error": f"Lỗi: {e}"}, "treo": {"error": f"Lỗi: {e}"}}

    # 5. Process Unstable POs
    try:
        print("Parsing Unstable POs...")
        UNSTABLE_PO_CACHE = process_unstable_po()
        gc.collect()
    except Exception as e:
        UNSTABLE_PO_CACHE = {"error": f"Lỗi: {e}"}
        
    # 6. Process OFF SPE
    try:
        print("Parsing OFF SPE...")
        OFF_SPE_CACHE = process_off_spe()
        gc.collect()
    except Exception as e:
        OFF_SPE_CACHE = {"error": f"Lỗi: {e}"}
        
    # 7. Process Volume Creation
    try:
        print("Parsing Volume Creation...")
        DF_TAO_DON_CACHE = load_vols_tao_don_df()
        DF_BUU_CUC_TYPE_MAP = load_buu_cuc_type_map()
        gc.collect()
    except Exception as e:
        print(f"Error volume: {e}")
        
    print("--------------------------------------------------")
    print("MEMORY-EFFICIENT CACHE LOAD COMPLETE.")
    print("--------------------------------------------------")
    gc.collect()


# ==========================================
# 7. API FLASK ENDPOINTS
# ==========================================
@app.route('/')
def home():
    return render_template('index.html')

@app.route('/api/login', methods=['POST'])
def api_login():
    try:
        data = request.json or {}
        username = data.get("username", "").strip()
        password = data.get("password", "").strip()
        
        ip = request.remote_addr
        is_ok, retry_after = check_rate_limit(f"{ip}_login", limit=10, period=60)
        if not is_ok:
            return jsonify({"error": f"Quá nhiều yêu cầu đăng nhập. Vui lòng thử lại sau {retry_after} giây."}), 429
        
        if not username or not password:
            return jsonify({"error": "Vui lòng nhập đầy đủ mã nhân viên và mật khẩu."}), 400
            
        users = load_users()
        if username in users and check_password_hash(users[username]["password"], password):
            session.clear()
            session['username'] = username
            session['role'] = users[username].get("role", "staff")
            session['permissions'] = users[username].get("permissions", [])
            session.permanent = True
            return jsonify({
                "success": True,
                "username": username,
                "role": session['role'],
                "permissions": session['permissions']
            })
        return jsonify({"error": "Sai mã nhân viên hoặc mật khẩu."}), 401
    except Exception as e:
        return jsonify({"error": f"Lỗi đăng nhập: {str(e)}"}), 500

@app.route('/api/logout', methods=['POST'])
def api_logout():
    session.clear()
    return jsonify({"success": True, "message": "Đăng xuất thành công."})

@app.route('/api/users', methods=['GET', 'POST'])
@requires_permission('tab-sync')
def api_users():
    if not is_admin():
        return jsonify({"error": "Quyền truy cập bị từ chối."}), 403
        
    users = load_users()
    
    if request.method == 'GET':
        users_list = []
        for uname, udata in users.items():
            users_list.append({
                "username": udata["username"],
                "role": udata.get("role", "staff"),
                "permissions": udata.get("permissions", [])
            })
        return jsonify(users_list)
        
    elif request.method == 'POST':
        data = request.json or {}
        username = data.get("username", "").strip()
        password = data.get("password", "").strip()
        role = data.get("role", "staff").strip()
        permissions = data.get("permissions", [])
        
        if not username:
            return jsonify({"error": "Username không được để trống."}), 400
            
        if username not in users and not password:
            return jsonify({"error": "Mật khẩu không được để trống khi tạo mới user."}), 400
            
        if username in users:
            users[username]["role"] = role
            users[username]["permissions"] = permissions
            if password:
                users[username]["password"] = generate_password_hash(password, method="pbkdf2:sha256")
        else:
            users[username] = {
                "username": username,
                "password": generate_password_hash(password, method="pbkdf2:sha256"),
                "role": role,
                "permissions": permissions
            }
            
        save_users(users)
        return jsonify({"success": True, "message": f"User {username} đã được lưu thành công."})

@app.route('/api/users/<username>', methods=['DELETE'])
@requires_permission('tab-sync')
def api_delete_user(username):
    if not is_admin():
        return jsonify({"error": "Quyền truy cập bị từ chối."}), 403
        
    current_user = session.get('username')
    if not current_user:
        auth = request.authorization
        if auth:
            current_user = auth.username
            
    if current_user == username:
        return jsonify({"error": "Không thể tự xóa chính mình."}), 400
        
    users = load_users()
    if username not in users:
        return jsonify({"error": "User không tồn tại."}), 404
        
    admin_count = sum(1 for u in users.values() if u.get("role") == "admin")
    if users[username].get("role") == "admin" and admin_count <= 1:
        return jsonify({"error": "Không thể xóa admin duy nhất của hệ thống."}), 400
        
    del users[username]
    save_users(users)
    return jsonify({"success": True, "message": f"User {username} đã được xóa."})

@app.route('/api/unstable-po')
@requires_permission('tab-unstable-po')
def get_unstable_po():
    global UNSTABLE_PO_CACHE
    if UNSTABLE_PO_CACHE is None:
        with CACHE_LOCK:
            if UNSTABLE_PO_CACHE is None:
                UNSTABLE_PO_CACHE = process_unstable_po()
    return jsonify(UNSTABLE_PO_CACHE)

@app.route('/api/off-spe')
@requires_permission('tab-off-spe')
def get_off_spe():
    global OFF_SPE_CACHE
    if OFF_SPE_CACHE is None:
        with CACHE_LOCK:
            if OFF_SPE_CACHE is None:
                OFF_SPE_CACHE = process_off_spe()
    return jsonify(OFF_SPE_CACHE)

@app.route('/api/operational')
@requires_permission('tab-operational')
def get_operational():
    global OPERATIONAL_CACHE, DF_GTC_CACHE, DF_LTC_CACHE, DF_TTS_CACHE
    am = request.args.get('am')
    province = request.args.get('province')
    po = request.args.get('po')
    date = request.args.get('date')
    
    if am or province or po or date:
        df_gtc = DF_GTC_CACHE if DF_GTC_CACHE is not None else None
        df_ltc = DF_LTC_CACHE if DF_LTC_CACHE is not None else None
        df_tts = DF_TTS_CACHE if DF_TTS_CACHE is not None else None
        
        ops = process_operational_report(
            df_gtc=df_gtc,
            df_ltc=df_ltc,
            df_tts=df_tts,
            am=am,
            province=province,
            po=po,
            date=date
        )
        return jsonify(clean_nan(ops))
        
    with CACHE_LOCK:
        if OPERATIONAL_CACHE is None:
            OPERATIONAL_CACHE = process_operational_report()
    return jsonify(clean_nan(OPERATIONAL_CACHE))

@app.route('/api/opr')
@requires_permission('tab-opr')
def get_opr():
    am = request.args.get('am')
    province = request.args.get('province')
    post_office = request.args.get('post_office')
    
    if am or province or post_office:
        # Dynamically compute OPR report with filters applied
        filtered_report = process_opr_report(am=am, province=province, post_office=post_office)
        return jsonify(clean_nan(filtered_report))
        
    global OPR_CACHE
    with CACHE_LOCK:
        if OPR_CACHE is None:
            OPR_CACHE = process_opr_report()
    return jsonify(clean_nan(OPR_CACHE))

@app.route('/api/backlog')
@requires_permission('tab-backlog')
def get_backlog():
    global BACKLOG_CACHE_RAW
    with CACHE_LOCK:
        if BACKLOG_CACHE_RAW is None:
            aging = process_aging_backlog()
            treo = process_treo_backlog()
            BACKLOG_CACHE_RAW = {
                "aging": aging,
                "treo": treo
            }
        
    if "error" in BACKLOG_CACHE_RAW["aging"]:
        return jsonify({"error": BACKLOG_CACHE_RAW["aging"]["error"]})
    if "error" in BACKLOG_CACHE_RAW["treo"]:
        return jsonify({"error": BACKLOG_CACHE_RAW["treo"]["error"]})
        
    history = load_history()
    baseline_ts = request.args.get("baseline")
    baseline_entry = None
    
    if history:
        if baseline_ts:
            for entry in history:
                if entry["timestamp"] == baseline_ts:
                    baseline_entry = entry
                    break
        else:
            if len(history) >= 2:
                baseline_entry = history[-2]
            else:
                baseline_entry = history[-1]
                
    import copy
    current_data = copy.deepcopy(BACKLOG_CACHE_RAW)
    current_data["baseline_timestamp"] = baseline_entry["timestamp"] if baseline_entry else None
    current_data = calculate_trend(current_data, baseline_entry)
    
    return jsonify(clean_nan(current_data))

@app.route('/api/history')
@requires_permission('tab-backlog')
def get_history():
    history = load_history()
    timestamps = [entry["timestamp"] for entry in history]
    return jsonify(timestamps[::-1])

@app.route('/api/user-role')
@requires_auth
def get_user_role():
    username = session.get('username')
    if not username and request.authorization:
        username = request.authorization.username
    role = 'admin' if is_admin() else 'staff'
    users = load_users()
    user = users.get(username, {})
    permissions = user.get("permissions", [])
    return jsonify({"username": username, "role": role, "permissions": permissions})

def mask_url(url):
    if not url:
        return ""
    match = re.search(r'/spreadsheets/d/([a-zA-Z0-9-_]+)', url)
    if not match:
        return url
    spreadsheet_id = match.group(1)
    if len(spreadsheet_id) > 10:
        masked_id = spreadsheet_id[:4] + "****************" + spreadsheet_id[-4:]
    else:
        masked_id = "********"
    masked_url = url.replace(spreadsheet_id, masked_id)
    return masked_url

def mask_token(token):
    if not token:
        return ""
    if len(token) > 8:
        return token[:4] + "****************" + token[-4:]
    return "********"

@app.route('/api/config', methods=['GET', 'POST'])
@requires_permission('tab-sync')
def manage_config():
    if request.method == 'POST':
        if not is_admin():
            return jsonify({"error": "Quyền truy cập bị từ chối. Chỉ tài khoản admin mới được phép lưu cấu hình."}), 403
        try:
            data = request.json or {}
            current_config = load_config()
            
            config = {}
            if "consolidated_url" in data:
                val = data.get("consolidated_url", "").strip()
                if not val:
                    config["consolidated_url"] = ""
                elif "*" in val or "Ẩn" in val or "hidden" in val.lower():
                    config["consolidated_url"] = current_config.get("consolidated_url", "")
                else:
                    if not (val.startswith("https://docs.google.com/spreadsheets/") or val.startswith("http://docs.google.com/spreadsheets/")):
                        return jsonify({"error": "Link Google Sheet không hợp lệ cho consolidated_url. Phải bắt đầu bằng https://docs.google.com/spreadsheets/"}), 400
                    if not is_allowed_spreadsheet_url(val):
                        return jsonify({"error": "Spreadsheet ID không nằm trong danh sách được cho phép (SSRF Prevention)."}), 400
                    config["consolidated_url"] = val
            else:
                config["consolidated_url"] = current_config.get("consolidated_url", "")

            # Parse Telegram Bot Token
            if "telegram_bot_token" in data:
                val_token = data.get("telegram_bot_token", "").strip()
                if not val_token:
                    config["telegram_bot_token"] = ""
                elif "*" in val_token or "hidden" in val_token.lower():
                    config["telegram_bot_token"] = current_config.get("telegram_bot_token", "")
                else:
                    config["telegram_bot_token"] = val_token
            else:
                config["telegram_bot_token"] = current_config.get("telegram_bot_token", "")

            # Parse Telegram Chat ID
            if "telegram_chat_id" in data:
                config["telegram_chat_id"] = data.get("telegram_chat_id", "").strip()
            else:
                config["telegram_chat_id"] = current_config.get("telegram_chat_id", "")

            # Parse Gemini API Key
            if "gemini_api_key" in data:
                val_gemini = data.get("gemini_api_key", "").strip()
                if not val_gemini:
                    config["gemini_api_key"] = ""
                elif "*" in val_gemini or "hidden" in val_gemini.lower():
                    config["gemini_api_key"] = current_config.get("gemini_api_key", "")
                else:
                    config["gemini_api_key"] = val_gemini
            else:
                config["gemini_api_key"] = current_config.get("gemini_api_key", "")
            
            if save_config(config):
                masked_config = {
                    "consolidated_url": mask_url(config.get("consolidated_url", "")),
                    "telegram_bot_token": mask_token(config.get("telegram_bot_token", "")),
                    "telegram_chat_id": config.get("telegram_chat_id", ""),
                    "gemini_api_key": mask_token(config.get("gemini_api_key", ""))
                }
                return jsonify({"success": True, "config": masked_config})
            else:
                return jsonify({"error": "Không thể ghi cấu hình."}), 500
        except Exception as e:
            return jsonify({"error": f"Lỗi xử lý yêu cầu: {str(e)}"}), 400
    else:
        config = load_config()
        masked_config = {
            "consolidated_url": mask_url(config.get("consolidated_url", "")),
            "telegram_bot_token": mask_token(config.get("telegram_bot_token", "")),
            "telegram_chat_id": config.get("telegram_chat_id", ""),
            "gemini_api_key": mask_token(config.get("gemini_api_key", ""))
        }
        return jsonify(masked_config)

@app.route('/api/download-template', methods=['GET'])
@requires_permission('tab-sync')
def download_template():
    from flask import send_from_directory
    filename = request.args.get('filename', '').strip()
    allowed_files = [
        'ops_gtc.csv',
        'ops_ltc.csv',
        'ops_co_cau.csv',
        'opr_opr.csv',
        'opr_oe.csv',
        'opr_raw.csv',
        'aging_raw.csv',
        'treo_stuck.csv',
        'buu_cuc_bat_on.csv',
        'off_tuyen_spe.csv',
        'vols_tao_don.csv',
        'co_cau_ntb.csv'
    ]
    if filename not in allowed_files:
        return jsonify({"error": "Tên file không hợp lệ."}), 400
        
    try:
        filepath = resolve_path(filename, write=False)
        if not os.path.exists(filepath):
            headers_map = {
                'ops_gtc.csv': 'Cấp Quản Lý,Chi tiết,Loại Hàng,Time,Volume,% Gán,% GTC,% Chuyển trả,Leadtime,AM,Tỉnh',
                'ops_ltc.csv': 'Cấp quản lý,Chi tiết,Ca,Time,Volume,%LTC,%LC,%Gán,Leadtime',
                'ops_co_cau.csv': 'BC,Bưu cục,AM,Tỉnh',
                'opr_opr.csv': 'NgayLTC,vol_ltc,ot,khung_gio_tao_don,ly_do_tre_12h,AM',
                'opr_oe.csv': 'tutinh,kholay,sellername,khung_gio_tao_don,ly_do_tre_12h,madh,AM',
                'opr_raw.csv': 'NgayLTC,vol_ltc,ot,khung_gio_tao_don,ly_do_tre_12h,AM',
                'aging_raw.csv': 'Tỉnh,BC,Mã đơn,Tệp khách,Số ngày đã nhập BC,Số lần giao,am_name,Trạng thái',
                'treo_stuck.csv': 'warehouse_name,Mã đơn hàng,Loại đơn,Thời gian tồn đọng,am_name,province_name,Trạng thái',
                'buu_cuc_bat_on.csv': 'kho_giao_id,kho_giao_name,tinh_giao,BL LM,BL LM >5 ngay,%BL LM >5 ngay,BL KTC,BL KTC cung tinh,%BL KTC cung tinh,du_kien_clear_ton,ly_do_bat_on,trang_thai',
                'off_tuyen_spe.csv': 'tỉnh,huyện,phường,bưu cục,kết quả,cap,% time off,tg mở,ghi chú',
                'vols_tao_don.csv': 'Date,Province,buu_cuc,bat_on,Vol_Shopee_Tiktok,Vol_LTC_Ontime_Shopee_Tiktok,OPR_Shopee_Tiktok',
                'co_cau_ntb.csv': 'warehouse_id,Bưu cục,AM,Tỉnh'
            }
            if filename in headers_map:
                os.makedirs(os.path.dirname(filepath), exist_ok=True)
                with open(filepath, 'w', encoding='utf-8-sig') as f:
                    f.write(headers_map[filename] + "\n")
        
        directory = os.path.dirname(filepath)
        name = os.path.basename(filepath)
        return send_from_directory(directory, name, as_attachment=True)
    except Exception as e:
        return jsonify({"error": f"Không thể tải file: {str(e)}"}), 500

@app.route('/api/upload', methods=['POST'])
@requires_permission('tab-sync')
def upload_file():
    global OPERATIONAL_CACHE, OPR_CACHE, BACKLOG_CACHE_RAW
    if 'file' not in request.files:
        return jsonify({"error": "Không tìm thấy file trong yêu cầu."}), 400
        
    file = request.files['file']
    filename = request.form.get('filename', '').strip()
    
    if not filename:
        return jsonify({"error": "Không xác định được tên file đích."}), 400
        
    allowed_files = [
        'ops_gtc.csv',
        'ops_ltc.csv',
        'ops_co_cau.csv',
        'opr_opr.csv',
        'opr_oe.csv',
        'opr_raw.csv',
        'aging_raw.csv',
        'treo_stuck.csv',
        'buu_cuc_bat_on.csv',
        'off_tuyen_spe.csv',
        'vols_tao_don.csv',
        'co_cau_ntb.csv'
    ]
    
    if filename not in allowed_files:
        return jsonify({"error": "Tên file không hợp lệ."}), 400
        
    if not file.filename.endswith('.csv'):
        return jsonify({"error": "Định dạng file không hợp lệ. Chỉ hỗ trợ file CSV (.csv)."}), 400
        
    try:
        filepath = resolve_path(filename, write=True)
        file.save(filepath)
        
        update_all_caches()
        if BACKLOG_CACHE_RAW and "error" not in BACKLOG_CACHE_RAW["aging"] and "error" not in BACKLOG_CACHE_RAW["treo"]:
            add_to_history(BACKLOG_CACHE_RAW["aging"], BACKLOG_CACHE_RAW["treo"])
                
        return jsonify({"success": True, "message": f"Tải lên và đồng bộ file {filename} thành công!"})
    except Exception as e:
        return jsonify({"error": f"Lỗi lưu file hoặc cập nhật cache: {str(e)}"}), 500

SYNC_STATUS = {
    "status": "idle",
    "progress": "",
    "error": None,
    "timestamp": None
}

def split_excel_to_csvs(xlsx_path):
    print(f"Splitting {xlsx_path} to CSV files using optimized openpyxl...")
    import openpyxl
    import csv
    try:
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        sheet_names = wb.sheetnames
        sheet_names_lower = [s.strip().lower() for s in sheet_names]
        
        sheet_mappings = [
            (["data"], "ops_gtc.csv"),
            (["dataltc", "rawltc", "data ltc"], "ops_ltc.csv"),
            (["cocauvung", "cơ cấu", "co_cau", "co cau"], "ops_co_cau.csv"),
            (["cocauvung", "cơ cấu", "co_cau", "co cau"], "co_cau_ntb.csv"),
            (["tts"], "ops_tts.csv"),
            (["opr"], "opr_opr.csv"),
            (["raw n-1", "oe_madh", "raw_n-1", "raw n - 1", "oe madh"], "opr_oe.csv"),
            (["rawopr"], "opr_raw.csv"),
            (["aging trên 5 ngày", "aging tren 5 ngay", "đơn giao aging trên 5 ngày", "don giao aging tren 5 ngay"], "aging_raw.csv"),
            (["treo lc", "stuck", "treo luân chuyển", "treo luan chuyen"], "treo_stuck.csv"),
            (["ntb", "bất ổn", "bat_on", "cảnh báo"], "buu_cuc_bat_on.csv"),
            (["đang off", "dang off", "off", "off_tuyen", "off tuyến"], "off_tuyen_spe.csv"),
            (["shopee_tiktok", "tao_don", "tạo đơn"], "vols_tao_don.csv"),
            (["odr tts", "odr_tts"], "ODR TTS.csv"),
            (["nhân sự", "nhan su"], "ops_nhan_su.csv")
        ]
        
        for candidates, target_csv in sheet_mappings:
            matched_sheet = None
            for c in candidates:
                if c in sheet_names_lower:
                    idx = sheet_names_lower.index(c)
                    matched_sheet = sheet_names[idx]
                    break
            
            if matched_sheet:
                safe_name = matched_sheet.encode('ascii', 'backslashreplace').decode('ascii')
                print(f"Extracting '{safe_name}' -> {target_csv}...")
                sheet = wb[matched_sheet]
                csv_path = resolve_path(target_csv, write=True)
                with open(csv_path, 'w', encoding='utf-8-sig', newline='') as f:
                    writer = csv.writer(f)
                    for row in sheet.iter_rows(values_only=True):
                        # Filter out completely empty rows
                        if not any(x is not None for x in row):
                            continue
                        writer.writerow(row)
                try:
                    df = pd.read_csv(csv_path)
                    save_df_to_db(df, target_csv)
                except Exception as e:
                    print(f"Error saving split sheet {target_csv} to DB: {e}")
            else:
                print(f"Warning: sheet for {target_csv} not found in Excel workbook.")
        wb.close()
        return True
    except Exception as e:
        print(f"Error splitting Excel to CSVs: {e}")
        import traceback
        traceback.print_exc()
        return False

def sync_sheets_directly_as_csv(url):
    import urllib.request
    import re
    import concurrent.futures
    import time
    
    if not url or not url.strip():
        return False, "Không có link Google Sheet."
        
    url = url.strip()
    if not is_allowed_spreadsheet_url(url):
        return False, "Spreadsheet ID không nằm trong danh sách được cho phép (SSRF Prevention)."
        
    match = re.search(r'/spreadsheets/d/([a-zA-Z0-9-_]+)', url)
    if not match:
        return False, "Không tìm thấy Spreadsheet ID hợp lệ trong link."
    spreadsheet_id = match.group(1)
    
    edit_url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/edit"
    user_agent = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko)'
    
    print(f"Fetching edit HTML to extract GIDs from: {edit_url}")
    req = urllib.request.Request(edit_url, headers={'User-Agent': user_agent})
    try:
        with urllib.request.urlopen(req, timeout=20) as r:
            html = r.read().decode('utf-8')
    except Exception as e:
        return False, f"Lỗi truy cập link Google Sheet: {str(e)}"
        
    pattern = r'\[\s*\d+\s*,\s*0\s*,\s*\\"?(\d+)\\"?\s*,\s*\[\s*\{\s*\\"?1\\"?\s*:\s*\[\s*\[\s*0\s*,\s*0\s*,\s*\\"?([^\\"\(\]]+)\\"?'
    matches = re.findall(pattern, html)
    if not matches:
        return False, "Không tìm thấy danh sách sheet (GIDs) trong HTML."
        
    gid_map = {}
    for gid, name in matches:
        gid_map[name.strip().lower()] = gid
        
    sheet_mappings = [
        (["data"], "ops_gtc.csv"),
        (["dataltc", "rawltc", "data ltc"], "ops_ltc.csv"),
        (["cocauvung", "cơ cấu", "co_cau", "co cau"], "ops_co_cau.csv"),
        (["cocauvung", "cơ cấu", "co_cau", "co cau"], "co_cau_ntb.csv"),
        (["tts"], "ops_tts.csv"),
        (["opr"], "opr_opr.csv"),
        (["raw n-1", "oe_madh", "raw_n-1", "raw n - 1", "oe madh"], "opr_oe.csv"),
        (["rawopr"], "opr_raw.csv"),
        (["aging trên 5 ngày", "aging tren 5 ngay", "đơn giao aging trên 5 ngày", "don giao aging tren 5 ngay"], "aging_raw.csv"),
        (["treo lc", "stuck", "treo luân chuyển", "treo luan chuyen"], "treo_stuck.csv"),
        (["ntb", "bất ổn", "bat_on", "cảnh báo"], "buu_cuc_bat_on.csv"),
        (["đang off", "dang off", "off", "off_tuyen", "off tuyến"], "off_tuyen_spe.csv"),
        (["shopee_tiktok", "tao_don", "tạo đơn"], "vols_tao_don.csv"),
        (["odr tts", "odr_tts"], "ODR TTS.csv"),
        (["nhân sự", "nhan su"], "ops_nhan_su.csv")
    ]
    
    gid_to_filenames = {}
    for candidates, target_csv in sheet_mappings:
        matched_gid = None
        for cand in candidates:
            cand_clean = cand.strip().lower()
            if cand_clean in gid_map:
                matched_gid = gid_map[cand_clean]
                break
        if matched_gid is not None:
            if matched_gid not in gid_to_filenames:
                gid_to_filenames[matched_gid] = []
            gid_to_filenames[matched_gid].append(target_csv)
            
    if not gid_to_filenames:
        return False, "Không tìm thấy sheet nào khớp với cấu hình."
        
    print(f"Found {len(gid_to_filenames)} unique GIDs to download.")
    
    def download_gid(gid, filenames):
        csv_url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/export?format=csv&gid={gid}"
        req_csv = urllib.request.Request(csv_url, headers={'User-Agent': user_agent})
        try:
            with urllib.request.urlopen(req_csv, timeout=30) as response:
                content = response.read()
            for filename in filenames:
                csv_path = resolve_path(filename, write=True)
                with open(csv_path, 'wb') as f:
                    f.write(content)
                try:
                    for encoding in ['utf-8-sig', 'utf-8', 'latin-1', 'cp1252']:
                        try:
                            import io
                            df = pd.read_csv(io.BytesIO(content), encoding=encoding)
                            save_df_to_db(df, filename)
                            break
                        except Exception:
                            continue
                except Exception as e:
                    print(f"Error parsing synced CSV {filename} for DB: {e}")
            return True
        except Exception as e:
            print(f"Error downloading GID {gid} for {filenames}: {e}")
            return False
            
    success_count = 0
    with concurrent.futures.ThreadPoolExecutor(max_workers=len(gid_to_filenames)) as executor:
        futures = {executor.submit(download_gid, gid, fnames): gid for gid, fnames in gid_to_filenames.items()}
        for future in concurrent.futures.as_completed(futures):
            gid = futures[future]
            try:
                if future.result():
                    success_count += len(gid_to_filenames[gid])
            except Exception as e:
                print(f"Exception downloading GID {gid}: {e}")
                
    if success_count == 0:
        return False, "Không tải được file CSV nào từ Google Sheets."
        
    return True, f"Đã tải thành công {success_count} files."

def async_sync_task(is_admin_flag):
    global OPERATIONAL_CACHE, OPR_CACHE, BACKLOG_CACHE_RAW, UNSTABLE_PO_CACHE, OFF_SPE_CACHE, SYNC_STATUS
    try:
        if is_admin_flag:
            config = load_config()
            url = config.get("consolidated_url", "")
            if url and url.strip():
                # 1. Try direct CSV sync first
                SYNC_STATUS["progress"] = "Đang tải dữ liệu trực tiếp dưới dạng CSV (Tối ưu)..."
                direct_success, direct_msg = sync_sheets_directly_as_csv(url)
                
                if direct_success:
                    print("Direct CSV sync succeeded:", direct_msg)
                else:
                    print("Direct CSV sync failed, falling back to XLSX split:", direct_msg)
                    # 2. Fallback to Excel download and split
                    SYNC_STATUS["progress"] = "Đang tải file Google Sheet tổng hợp (Dự phòng XLSX)..."
                    temp_xlsx = resolve_path('downloaded_consolidated_sheet.xlsx', write=True)
                    success, msg = download_google_sheet(url, temp_xlsx)
                    if not success:
                        SYNC_STATUS["status"] = "error"
                        SYNC_STATUS["error"] = f"Lỗi tải Google Sheet: {msg}"
                        return
                    
                    SYNC_STATUS["progress"] = "Đang trích xuất các sheet thành CSV (Dự phòng XLSX)..."
                    if not split_excel_to_csvs(temp_xlsx):
                        SYNC_STATUS["status"] = "error"
                        SYNC_STATUS["error"] = "Lỗi khi trích xuất dữ liệu từ Google Sheet sang CSV."
                        return
            else:
                print("No consolidated URL configured, using existing CSV files.")
                
        SYNC_STATUS["progress"] = "Đang xử lý dữ liệu và cập nhật bộ nhớ đệm..."
        
        update_all_caches()
        
        if OPERATIONAL_CACHE and "error" in OPERATIONAL_CACHE:
            SYNC_STATUS["status"] = "error"
            SYNC_STATUS["error"] = OPERATIONAL_CACHE["error"]
            return
        if OPR_CACHE and "error" in OPR_CACHE:
            SYNC_STATUS["status"] = "error"
            SYNC_STATUS["error"] = OPR_CACHE["error"]
            return
        if BACKLOG_CACHE_RAW:
            if "error" in BACKLOG_CACHE_RAW["aging"]:
                SYNC_STATUS["status"] = "error"
                SYNC_STATUS["error"] = BACKLOG_CACHE_RAW["aging"]["error"]
                return
            if "error" in BACKLOG_CACHE_RAW["treo"]:
                SYNC_STATUS["status"] = "error"
                SYNC_STATUS["error"] = BACKLOG_CACHE_RAW["treo"]["error"]
                return
        
        ts = add_to_history(BACKLOG_CACHE_RAW["aging"], BACKLOG_CACHE_RAW["treo"])
            
        SYNC_STATUS["status"] = "success"
        SYNC_STATUS["timestamp"] = ts
        SYNC_STATUS["progress"] = "Đồng bộ hoàn tất!"
        
    except Exception as e:
        SYNC_STATUS["status"] = "error"
        SYNC_STATUS["error"] = f"Lỗi hệ thống khi đồng bộ: {str(e)}"

@app.route('/api/sync', methods=['POST'])
@requires_permission('tab-sync')
def trigger_sync():
    global SYNC_STATUS
    ip = request.remote_addr
    is_ok, retry_after = check_rate_limit(f"{ip}_sync", limit=3, period=60)
    if not is_ok:
        return jsonify({"error": f"Yêu cầu đồng bộ quá nhanh. Vui lòng thử lại sau {retry_after} giây."}), 429
        
    if SYNC_STATUS["status"] == "processing":
        return jsonify({"error": "Đang có quá trình đồng bộ đang chạy."}), 400
        
    SYNC_STATUS["status"] = "processing"
    SYNC_STATUS["error"] = None
    SYNC_STATUS["progress"] = "Đang khởi tạo đồng bộ..."
    
    admin_flag = is_admin()
    
    # On Vercel, run synchronously because background threads are killed immediately when the HTTP response returns.
    is_vercel = os.environ.get("VERCEL") or (request.host and "vercel.app" in request.host) or (request.headers.get("Host") and "vercel.app" in request.headers.get("Host"))
    if is_vercel:
        async_sync_task(admin_flag)
        if SYNC_STATUS["status"] == "error":
            return jsonify({"error": SYNC_STATUS["error"]}), 500
        return jsonify({"success": True, "status": "success", "message": "Đồng bộ hoàn tất!"})
    else:
        threading.Thread(target=async_sync_task, args=(admin_flag,), daemon=True).start()
        return jsonify({"success": True, "status": "started"})

@app.route('/api/sync/status', methods=['GET'])
@requires_permission('tab-sync')
def get_sync_status():
    global SYNC_STATUS
    return jsonify(SYNC_STATUS)


@app.route('/api/nhan-su', methods=['GET'])
@requires_permission('tab-nhan-su')
def api_nhan_su():
    try:
        ns_path = resolve_path('ops_nhan_su.csv', write=False)
        df_ns = safe_read_csv(ns_path)
        if df_ns is None or df_ns.empty:
            return jsonify({"error": "Không tìm thấy dữ liệu nhân sự (ops_nhan_su.csv). Vui lòng thực hiện Đồng bộ dữ liệu."}), 404
            
        df_ns.columns = [c.strip() for c in df_ns.columns]
        
        def check_sp_team(x):
            val_str = str(x).strip().lower()
            return val_str in ['true', '1', 'yes', 't']
            
        active_mask = (
            (df_ns['Trạng thái'].astype(str).str.strip() == "Đang làm việc") &
            (df_ns['Vùng'].astype(str).str.strip() == "NTB") &
            (df_ns['Chức vụ'].astype(str).str.strip() == "Business Development Field Executive") &
            (~df_ns['SP Team?'].apply(check_sp_team))
        )
        active_df = df_ns[active_mask].copy()
        
        resigned_mask = (
            (df_ns['Trạng thái'].astype(str).str.strip() == "Đã nghỉ") &
            (df_ns['Vùng'].astype(str).str.strip() == "NTB")
        )
        resigned_count = len(df_ns[resigned_mask])
        
        def parse_warehouse_id(bc_str):
            if pd.isna(bc_str):
                return None
            parts = str(bc_str).split(" - ", 1)
            if len(parts) == 2:
                try:
                    return int(parts[0].strip())
                except:
                    return parts[0].strip()
            return None
            
        active_df['warehouse_id'] = active_df['Bưu cục'].apply(parse_warehouse_id)
        
        cc_path = resolve_path('ops_co_cau.csv', write=False)
        df_cc = safe_read_csv(cc_path)
        if df_cc is None or df_cc.empty:
            return jsonify({"error": "Không tìm thấy dữ liệu cơ cấu (ops_co_cau.csv)."}), 404
            
        df_cc.columns = [c.strip() for c in df_cc.columns]
        
        def clean_id(x):
            try:
                return int(float(x))
            except:
                return str(x).strip()
        df_cc['warehouse_id_clean'] = df_cc['warehouse_id'].apply(clean_id)
        active_df['warehouse_id_clean'] = active_df['warehouse_id'].apply(clean_id)
        
        vol_path = resolve_path('vols_tao_don.csv', write=False)
        po_vols = {}
        unique_dates_count = 1
        df_vol = safe_read_csv(vol_path)
        if df_vol is not None and not df_vol.empty:
            df_vol.columns = [c.strip() for c in df_vol.columns]
            df_vol_ntb = df_vol[df_vol['Vùng'].astype(str).str.strip() == 'NTB'].copy()
            unique_dates_count = len(df_vol_ntb['Date'].unique()) if 'Date' in df_vol_ntb.columns else 1
            if unique_dates_count == 0:
                unique_dates_count = 1
            df_vol_ntb['warehouse_id_clean'] = df_vol_ntb['warehouse_id'].apply(clean_id)
            vol_sums = df_vol_ntb.groupby('warehouse_id_clean')['Volume'].sum()
            for wh_id, v_sum in vol_sums.items():
                po_vols[wh_id] = round(v_sum / unique_dates_count)
                    
        hc_counts = active_df.groupby('warehouse_id_clean').size().to_dict()
        
        po_list = []
        for _, r in df_cc.iterrows():
            wh_id = r.get('warehouse_id')
            wh_id_clean = clean_id(wh_id)
            bc_name = r.get('Bưu cục', '')
            province = r.get('Tỉnh', '')
            am = r.get('AM', '')
            
            hc = hc_counts.get(wh_id_clean, 0)
            vol = po_vols.get(wh_id_clean, 0)
            
            po_list.append({
                "warehouse_id": str(wh_id),
                "name": str(bc_name),
                "province": str(province),
                "am": str(am),
                "hc_current": hc,
                "vol_2025": vol
            })
            
        active_pos_count = len([x for x in po_list if x["hc_current"] > 0 or x["vol_2025"] > 0])
        
        return jsonify({
            "active_headcount": len(active_df),
            "resigned_headcount": resigned_count,
            "po_count": active_pos_count,
            "pos": po_list
        })
    except Exception as e:
        print(f"Error in api_nhan_su: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Lỗi hệ thống: {str(e)}"}), 500


@app.route('/api/summary-dashboard')
@requires_permission('tab-dashboard')
def api_summary_dashboard():
    try:
        df_gtc, df_ltc, df_aging, df_treo = get_dataframes()
        
        # Filters from query params
        am = request.args.get('am')
        province = request.args.get('province')
        post_office = request.args.get('post_office')
        selected_date = request.args.get('date')
        
        # Find latest date dynamically
        all_times = set(df_gtc['Time'].dropna().unique()).union(set(df_ltc['Time'].dropna().unique()))
        dates_sorted = sorted(list(all_times), key=lambda x: pd.to_datetime(x.split(' - ')[0]))
        if not dates_sorted:
            return jsonify({"error": "Không có dữ liệu ngày trong Datagtc hoặc Dataltc"})
            
        if selected_date and selected_date in dates_sorted:
            latest_date = selected_date
        else:
            latest_date = dates_sorted[-1]
        
        prev_date = dates_sorted[-2] if len(dates_sorted) >= 2 else None
        
        latest_idx = dates_sorted.index(latest_date) if latest_date in dates_sorted else len(dates_sorted) - 1
        n3_date = dates_sorted[latest_idx - 3] if latest_idx >= 3 else None
        
        # Find same day of week last week (7 days ago)
        latest_dt = pd.to_datetime(latest_date.split(' - ')[0])
        last_week_date = None
        for d in dates_sorted:
            dt = pd.to_datetime(d.split(' - ')[0])
            if (latest_dt - dt).days == 7:
                last_week_date = d
                break
                
        def compute_kpi_for_date(date_str, filter_am=None, filter_prov=None, filter_po=None):
            if not date_str:
                return None
            gtc_d = df_gtc[df_gtc['Time'] == date_str]
            
            # Fallback if date is not present in df_ltc
            if date_str in df_ltc['Time'].values:
                ltc_d = df_ltc[df_ltc['Time'] == date_str]
            else:
                ltc_times = df_ltc['Time'].dropna().unique()
                if len(ltc_times) > 0:
                    ltc_dates_sorted = sorted(list(ltc_times), key=lambda x: pd.to_datetime(x.split(' - ')[0]))
                    ltc_d = df_ltc[df_ltc['Time'] == ltc_dates_sorted[-1]]
                else:
                    ltc_d = df_ltc[df_ltc['Time'] == date_str]
            
            gtc_d = apply_filters(gtc_d, filter_am, filter_prov, filter_po)
            ltc_d = apply_filters(ltc_d, filter_am, filter_prov, filter_po)
            
            ltc_vol = ltc_d['Volume'].sum()
            ltc_done = ltc_d['ltc_vol'].sum()
            gtc_vol = gtc_d['Volume'].sum()
            gtc_done = gtc_d['delivered_vol'].sum()
            ttc_done = gtc_d['ttc_vol'].sum()
            
            return {
                'ltc': round((ltc_done / ltc_vol * 100), 2) if ltc_vol > 0 else 0.0,
                'gtc': round((gtc_done / gtc_vol * 100), 2) if gtc_vol > 0 else 0.0,
                'ttc': round((ttc_done / gtc_vol * 100), 2) if gtc_vol > 0 else 0.0,
                'vol_ltc': int(ltc_done),
                'vol_gtc': int(gtc_vol)
            }
        
        # Filter for latest date
        today_gtc = df_gtc[df_gtc['Time'] == latest_date]
        
        # Fallback for today_ltc if latest_date not in df_ltc['Time'].values
        if latest_date in df_ltc['Time'].values:
            today_ltc = df_ltc[df_ltc['Time'] == latest_date]
        else:
            ltc_times = df_ltc['Time'].dropna().unique()
            if len(ltc_times) > 0:
                ltc_dates_sorted = sorted(list(ltc_times), key=lambda x: pd.to_datetime(x.split(' - ')[0]))
                today_ltc = df_ltc[df_ltc['Time'] == ltc_dates_sorted[-1]]
            else:
                today_ltc = df_ltc[df_ltc['Time'] == latest_date]
        
        # Apply filters to today's data
        today_gtc = apply_filters(today_gtc, am, province, post_office)
        today_ltc = apply_filters(today_ltc, am, province, post_office)
        
        # Apply filters to backlog datasets
        curr_aging = apply_filters(df_aging, am, province, post_office)
        curr_treo = apply_filters(df_treo, am, province, post_office)
        
        # 1. Completed Volume Grouped by Province
        gtc_prov = today_gtc.groupby('mapped_prov').agg({'Volume': 'sum', 'delivered_vol': 'sum', 'ttc_vol': 'sum'}).reset_index()
        ltc_prov = today_ltc.groupby('mapped_prov').agg({'Volume': 'sum', 'ltc_vol': 'sum'}).reset_index()
        
        merged_prov = pd.merge(gtc_prov, ltc_prov, on='mapped_prov', suffixes=('_gtc', '_ltc'), how='outer').fillna(0)
        completed_vols = []
        for _, r in merged_prov.iterrows():
            prov_name = r['mapped_prov']
            completed_vols.append({
                'province': prov_name,
                'LTC': int(r['ltc_vol']),
                'GTC': int(r['delivered_vol']),
                'TTC': int(r['ttc_vol']),
                'LTC_pct': round((r['ltc_vol'] / r['Volume_ltc'] * 100), 2) if r['Volume_ltc'] > 0 else 0.0,
                'GTC_pct': round((r['delivered_vol'] / r['Volume_gtc'] * 100), 2) if r['Volume_gtc'] > 0 else 0.0,
                'TTC_pct': round((r['ttc_vol'] / r['Volume_gtc'] * 100), 2) if r['Volume_gtc'] > 0 else 0.0,
                'Volume_gtc': int(r['Volume_gtc']),
                'Volume_ltc': int(r['ltc_vol'])
            })
            
        # 2. Backlog by AM
        treo_giao = curr_treo[curr_treo['Loại đơn'] == 'Luân chuyển giao']
        treo_tra = curr_treo[curr_treo['Loại đơn'] == 'Luân chuyển trả']
        
        aging_am = curr_aging.groupby('mapped_am').size().reset_index(name='chua_giao')
        treo_giao_am = treo_giao.groupby('mapped_am').size().reset_index(name='chua_lay')
        treo_tra_am = treo_tra.groupby('mapped_am').size().reset_index(name='chua_tra')
        
        backlog_am = pd.merge(treo_giao_am, aging_am, on='mapped_am', how='outer')
        backlog_am = pd.merge(backlog_am, treo_tra_am, on='mapped_am', how='outer').fillna(0)
        
        backlog_by_am = []
        for _, r in backlog_am.iterrows():
            backlog_by_am.append({
                'am': r['mapped_am'],
                'chua_lay': int(r['chua_lay']),
                'chua_giao': int(r['chua_giao']),
                'chua_tra': int(r['chua_tra'])
            })
            
        # 3. Overall & Province-specific KPIs
        kpis = {}
        
        def format_kpi_dict(latest_kpi, prev_kpi, lw_kpi, n3_kpi=None):
            res = {
                'ltc': latest_kpi['ltc'] if latest_kpi else 0.0,
                'gtc': latest_kpi['gtc'] if latest_kpi else 0.0,
                'ttc': latest_kpi['ttc'] if latest_kpi else 0.0,
                'vol_ltc': latest_kpi['vol_ltc'] if latest_kpi else 0,
                'vol_gtc': latest_kpi['vol_gtc'] if latest_kpi else 0
            }
            if latest_kpi:
                if prev_kpi:
                    res['dod'] = {
                        'ltc': round(latest_kpi['ltc'] - prev_kpi['ltc'], 2),
                        'gtc': round(latest_kpi['gtc'] - prev_kpi['gtc'], 2),
                        'ttc': round(latest_kpi['ttc'] - prev_kpi['ttc'], 2)
                    }
                else:
                    res['dod'] = None
                
                if lw_kpi:
                    res['wow'] = {
                        'ltc': round(latest_kpi['ltc'] - lw_kpi['ltc'], 2),
                        'gtc': round(latest_kpi['gtc'] - lw_kpi['gtc'], 2),
                        'ttc': round(latest_kpi['ttc'] - lw_kpi['ttc'], 2)
                    }
                else:
                    res['wow'] = None

                if n3_kpi:
                    res['n3'] = {
                        'ltc': round(latest_kpi['ltc'] - n3_kpi['ltc'], 2),
                        'gtc': round(latest_kpi['gtc'] - n3_kpi['gtc'], 2),
                        'ttc': round(latest_kpi['ttc'] - n3_kpi['ttc'], 2)
                    }
                else:
                    res['n3'] = None
            else:
                res['dod'] = None
                res['wow'] = None
                res['n3'] = None
            return res
            
        overall_latest = compute_kpi_for_date(latest_date, am, province, post_office)
        overall_prev = compute_kpi_for_date(prev_date, am, province, post_office) if prev_date else None
        overall_lw = compute_kpi_for_date(last_week_date, am, province, post_office) if last_week_date else None
        overall_n3 = compute_kpi_for_date(n3_date, am, province, post_office) if n3_date else None
        
        kpis['overall'] = format_kpi_dict(overall_latest, overall_prev, overall_lw, overall_n3)
        
        # Compute for each province
        all_provinces = set(df_ltc['mapped_prov'].unique()).union(set(df_gtc['mapped_prov'].unique()))
        for prov_name in all_provinces:
            if prov_name == "Không xác định":
                continue
            prov_latest = compute_kpi_for_date(latest_date, am, prov_name, post_office)
            prov_prev = compute_kpi_for_date(prev_date, am, prov_name, post_office) if prev_date else None
            prov_lw = compute_kpi_for_date(last_week_date, am, prov_name, post_office) if last_week_date else None
            prov_n3 = compute_kpi_for_date(n3_date, am, prov_name, post_office) if n3_date else None
            
            clean_key = clean_str_key(prov_name)
            kpis[clean_key] = format_kpi_dict(prov_latest, prov_prev, prov_lw, prov_n3)
            kpis[clean_key]['name'] = prov_name
            
        return jsonify(clean_nan({
            'latest_date': latest_date,
            'all_dates': dates_sorted[::-1],
            'completed_vols': completed_vols,
            'backlog_by_am': backlog_by_am,
            'kpis': kpis
        }))
        
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return jsonify({"error": f"Lỗi tính toán summary: {str(e)}"}), 500

@app.route('/api/trends-dashboard')
@requires_permission('tab-dashboard')
def api_trends_dashboard():
    try:
        df_gtc, df_ltc, _, _ = get_dataframes()
        
        am = request.args.get('am')
        province = request.args.get('province')
        post_office = request.args.get('post_office')
        
        # Apply filters
        df_gtc = apply_filters(df_gtc, am, province, post_office)
        df_ltc = apply_filters(df_ltc, am, province, post_office)
        
        # Find last 7 dates sorted chronologically
        all_times = set(df_gtc['Time'].dropna().unique()).union(set(df_ltc['Time'].dropna().unique()))
        dates_sorted = sorted(list(all_times), key=lambda x: pd.to_datetime(x.split(' - ')[0]))
        last_7_dates = dates_sorted[-7:]
        
        # Filter both dataframes to last 7 dates
        df_gtc_7 = df_gtc[df_gtc['Time'].isin(last_7_dates)]
        df_ltc_7 = df_ltc[df_ltc['Time'].isin(last_7_dates)]
        
        # Group by mapped_prov and Time
        gtc_grouped = df_gtc_7.groupby(['mapped_prov', 'Time']).agg({'Volume': 'sum', 'delivered_vol': 'sum'}).reset_index()
        ltc_grouped = df_ltc_7.groupby(['mapped_prov', 'Time']).agg({'Volume': 'sum', 'ltc_vol': 'sum'}).reset_index()
        
        # Overall trends
        gtc_overall = df_gtc_7.groupby('Time').agg({'Volume': 'sum', 'delivered_vol': 'sum'}).reset_index()
        ltc_overall = df_ltc_7.groupby('Time').agg({'Volume': 'sum', 'ltc_vol': 'sum'}).reset_index()
        
        res_dates = last_7_dates
        
        trends = {}
        all_provinces = set(df_ltc_7['mapped_prov'].unique()).union(set(df_gtc_7['mapped_prov'].unique()))
        
        for prov in all_provinces:
            if prov == "Không xác định":
                continue
            prov_ltc = []
            prov_gtc = []
            
            p_ltc_df = ltc_grouped[ltc_grouped['mapped_prov'] == prov]
            p_gtc_df = gtc_grouped[gtc_grouped['mapped_prov'] == prov]
            
            for d in res_dates:
                # LTC
                d_ltc = p_ltc_df[p_ltc_df['Time'] == d]
                if len(d_ltc) > 0 and d_ltc['Volume'].iloc[0] > 0:
                    prov_ltc.append(round((d_ltc['ltc_vol'].iloc[0] / d_ltc['Volume'].iloc[0] * 100), 2))
                else:
                    prov_ltc.append(None)
                # GTC
                d_gtc = p_gtc_df[p_gtc_df['Time'] == d]
                if len(d_gtc) > 0 and d_gtc['Volume'].iloc[0] > 0:
                    prov_gtc.append(round((d_gtc['delivered_vol'].iloc[0] / d_gtc['Volume'].iloc[0] * 100), 2))
                else:
                    prov_gtc.append(None)
                    
            # Normalize key to lowercase with underscores to match frontend (e.g. 'lâm_đồng', 'bình_thuận')
            trend_key = str(prov).strip().lower().replace(" ", "_")
            trends[trend_key] = {
                'ltc': prov_ltc,
                'gtc': prov_gtc
            }
            
        # Overall
        overall_ltc = []
        overall_gtc = []
        for d in res_dates:
            d_ltc = ltc_overall[ltc_overall['Time'] == d]
            if len(d_ltc) > 0 and d_ltc['Volume'].iloc[0] > 0:
                overall_ltc.append(round((d_ltc['ltc_vol'].iloc[0] / d_ltc['Volume'].iloc[0] * 100), 2))
            else:
                overall_ltc.append(None)
                
            d_gtc = gtc_overall[gtc_overall['Time'] == d]
            if len(d_gtc) > 0 and d_gtc['Volume'].iloc[0] > 0:
                overall_gtc.append(round((d_gtc['delivered_vol'].iloc[0] / d_gtc['Volume'].iloc[0] * 100), 2))
            else:
                overall_gtc.append(None)
                
        trends['overall'] = {
            'ltc': overall_ltc,
            'gtc': overall_gtc
        }
        
        return jsonify(clean_nan({
            'dates': res_dates,
            'trends': trends
        }))
        
    except Exception as e:
        return jsonify({"error": f"Lỗi tính toán xu hướng: {str(e)}"}), 500

@app.route('/api/matrix-tables')
@requires_permission('tab-volume-creation')
def api_matrix_tables():
    try:
        df_gtc, df_ltc, _, _ = get_dataframes()
        
        am_filter = request.args.get('am')
        province_filter = request.args.get('province')
        post_office_filter = request.args.get('post_office')
        
        # Apply filters
        df_gtc_f = apply_filters(df_gtc, am_filter, province_filter, post_office_filter)
        df_ltc_f = apply_filters(df_ltc, am_filter, province_filter, post_office_filter)
        
        # Find last 6 dates available
        all_times = set(df_gtc['Time'].dropna().unique()).union(set(df_ltc['Time'].dropna().unique()))
        dates_sorted = sorted(list(all_times), key=lambda x: pd.to_datetime(x.split(' - ')[0]))
        last_6_dates = dates_sorted[-6:][::-1]
        
        def build_matrix_data(df, vol_col, target_val_col):
            df_6 = df[df['Time'].isin(last_6_dates)].copy()
            if len(df_6) == 0:
                return {'dates': last_6_dates, 'rows': []}
                
            # Group by AM, Chi tiết, Time
            grouped = df_6.groupby(['mapped_am', 'Chi tiết', 'Time']).agg({'Volume': 'sum', target_val_col: 'sum'}).reset_index()
            
            # AM overall totals for each date
            am_totals = df_6.groupby(['mapped_am', 'Time']).agg({'Volume': 'sum', target_val_col: 'sum'}).reset_index()
            
            rows = []
            unique_ams = sorted(df_6['mapped_am'].unique())
            
            for am in unique_ams:
                if am == "Không xác định":
                    continue
                
                am_time_df = am_totals[am_totals['mapped_am'] == am]
                totals_dict = {}
                for d in last_6_dates:
                    d_row = am_time_df[am_time_df['Time'] == d]
                    if len(d_row) > 0:
                        vol = float(d_row['Volume'].iloc[0])
                        target_val = float(d_row[target_val_col].iloc[0])
                        pct = round((target_val / vol * 100), 2) if vol > 0 else 0.0
                        totals_dict[d] = {'vol': int(vol), 'pct': pct}
                    else:
                        totals_dict[d] = {'vol': 0, 'pct': 0.0}
                        
                am_details_df = grouped[grouped['mapped_am'] == am]
                unique_pos = sorted(am_details_df['Chi tiết'].unique())
                pos_list = []
                
                for po in unique_pos:
                    po_time_df = am_details_df[am_details_df['Chi tiết'] == po]
                    po_data_dict = {}
                    for d in last_6_dates:
                        d_row = po_time_df[po_time_df['Time'] == d]
                        if len(d_row) > 0:
                            vol = float(d_row['Volume'].iloc[0])
                            target_val = float(d_row[target_val_col].iloc[0])
                            pct = round((target_val / vol * 100), 2) if vol > 0 else 0.0
                            po_data_dict[d] = {'vol': int(vol), 'pct': pct}
                        else:
                            po_data_dict[d] = {'vol': 0, 'pct': 0.0}
                    pos_list.append({
                        'bc': po,
                        'data': po_data_dict
                    })
                    
                rows.append({
                    'am': am,
                    'is_header': True,
                    'totals': totals_dict,
                    'pos': pos_list
                })
                
            return {
                'dates': last_6_dates,
                'rows': rows
            }
            
        ltc_matrix = build_matrix_data(df_ltc_f, 'Volume', 'ltc_vol')
        gtc_matrix = build_matrix_data(df_gtc_f, 'Volume', 'delivered_vol')
        
        return jsonify(clean_nan({
            'ltc': ltc_matrix,
            'gtc': gtc_matrix
        }))
        
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return jsonify({"error": f"Lỗi tính toán matrix tables: {str(e)}"}), 500

@app.route('/api/chat', methods=['POST'])
@requires_auth
def api_chat():
    try:
        data = request.json or {}
        message = data.get('message', '').strip().lower()
        
        if not message:
            return jsonify({"reply": "Chào bạn! Mình có thể giúp gì cho bạn hôm nay?"})
            
        global OPERATIONAL_CACHE, OPR_CACHE
        # Ensure caches are loaded
        with CACHE_LOCK:
            if OPERATIONAL_CACHE is None:
                try:
                    OPERATIONAL_CACHE = process_operational_report()
                except Exception as e:
                    print(f"Error loading ops cache for chat: {e}")
            if OPR_CACHE is None:
                try:
                    OPR_CACHE = process_opr_report()
                except Exception as e:
                    print(f"Error loading opr cache for chat: {e}")
                    
        # Parse queries
        # 1. Bưu cục tệ nhất (Worst Post Office)
        if 'tệ nhất' in message or 'te nhat' in message or 'kém nhất' in message or 'kem nhat' in message:
            reply = "Dựa trên dữ liệu vận hành mới nhất:<br>"
            # GTC Worst
            if OPERATIONAL_CACHE and 'worst_10_gtc' in OPERATIONAL_CACHE and OPERATIONAL_CACHE['worst_10_gtc']:
                worst_gtc = OPERATIONAL_CACHE['worst_10_gtc'][0]
                reply += f"📍 **Bưu cục có tỷ lệ Giao thành công (GTC) thấp nhất**: <br><strong>{worst_gtc['Chi tiết']}</strong> với tỷ lệ GTC chỉ đạt <strong>{worst_gtc['% GTC']:.2f}%</strong> (Sản lượng: {int(worst_gtc['Volume'])} đơn).<br>"
            else:
                reply += "📍 Không tìm thấy dữ liệu bưu cục GTC.<br>"
                
            # LTC Worst
            if OPERATIONAL_CACHE and 'worst_10_ltc' in OPERATIONAL_CACHE and OPERATIONAL_CACHE['worst_10_ltc']:
                worst_ltc = OPERATIONAL_CACHE['worst_10_ltc'][0]
                reply += f"📍 **Bưu cục có tỷ lệ Giao lần đầu (LTC) thấp nhất**: <br><strong>{worst_ltc['Chi tiết']}</strong> với tỷ lệ LTC đạt <strong>{worst_ltc['% LTC']:.2f}%</strong> (Sản lượng: {int(worst_ltc['Volume'])} đơn).<br>"
                
            # Worst ODR (Ontime Delivery) from ODR TTS.csv
            try:
                odr_path = resolve_path('ODR TTS.csv', write=False)
                if os.path.exists(odr_path):
                    df_odr = pd.read_csv(odr_path)
                    df_odr['GTC'] = pd.to_numeric(df_odr['GTC'], errors='coerce')
                    df_odr['%Ontime'] = normalize_pct_col(df_odr['%Ontime'])
                    latest_date = df_odr['Time'].max()
                    df_latest = df_odr[(df_odr['Time'] == latest_date) & (df_odr['GTC'] >= 10)] # filter at least 10 orders
                    if not df_latest.empty:
                        worst_odr_row = df_latest.sort_values(by='%Ontime', ascending=True).iloc[0]
                        reply += f"📍 **Bưu cục trễ hẹn (ODR thấp nhất) ngày {latest_date}**: <br><strong>{worst_odr_row['Chi tiết']}</strong> với tỷ lệ Đúng giờ chỉ đạt <strong>{worst_odr_row['%Ontime']*100:.2f}%</strong> (Sản lượng: {int(worst_odr_row['GTC'])} đơn)."
            except Exception as e:
                print(f"Error computing worst ODR for chat: {e}")
                
            return jsonify({"reply": reply})

        # 2. OPR (Operational Performance Rate)
        elif 'opr' in message:
            if OPR_CACHE and 'overall_opr' in OPR_CACHE:
                opr_val = OPR_CACHE['overall_opr']
                reply = f"📊 **Tỷ lệ hiệu suất OPR ngày hôm nay**: đạt <strong>{opr_val}%</strong>.<br>"
                try:
                    val_float = float(str(opr_val).replace('%', ''))
                    if val_float >= 80:
                        reply += "✅ Hiệu suất OPR đạt mục tiêu tối thiểu (Target: >= 80%). Các bưu cục đang vận hành khá tốt!"
                    else:
                        reply += "⚠️ Hiệu suất OPR hiện tại dưới mục tiêu tối thiểu (Target: >= 80%). Cần tập trung cải thiện thời gian xử lý đơn hàng."
                except:
                    pass
            else:
                reply = "📊 Hiện tại hệ thống chưa cập nhật đủ dữ liệu OPR hôm nay."
            return jsonify({"reply": reply})
            
        # 3. ODR (Ontime Delivery Rate)
        elif 'odr' in message or 'đúng giờ' in message or 'trễ' in message or 'tre' in message or 'dung gio' in message:
            try:
                odr_path = resolve_path('ODR TTS.csv', write=False)
                if os.path.exists(odr_path):
                    df_odr = pd.read_csv(odr_path)
                    df_odr['GTC'] = pd.to_numeric(df_odr['GTC'], errors='coerce')
                    df_odr['%Ontime'] = normalize_pct_col(df_odr['%Ontime'])
                    df_odr['ontime_vol'] = df_odr['GTC'] * df_odr['%Ontime']
                    latest_date = df_odr['Time'].max()
                    df_latest = df_odr[df_odr['Time'] == latest_date]
                    if not df_latest.empty:
                        overall_odr = (df_latest['ontime_vol'].sum() / df_latest['GTC'].sum()) * 100
                        total_vol = df_latest['GTC'].sum()
                        reply = f"🚚 **Tỷ lệ Giao hàng đúng giờ (ODR) ngày {latest_date}**: đạt <strong>{overall_odr:.2f}%</strong> (Trên tổng sản lượng {int(total_vol):,} đơn giao).<br>"
                        if overall_odr >= 90:
                            reply += "🎉 Tỷ lệ giao hàng đúng giờ duy trì ở mức cao và đạt chuẩn chất lượng dịch vụ của GHN!"
                        else:
                            reply += "⚠️ Tỷ lệ giao đúng giờ đang giảm, vui lòng kiểm tra các tuyến luân chuyển bị nghẽn."
                    else:
                        reply = "🚚 Chưa có dữ liệu ODR cho ngày hôm nay."
                else:
                    reply = "🚚 Không tìm thấy file dữ liệu ODR TTS.csv để tính toán."
            except Exception as e:
                reply = f"🚚 Lỗi tính toán tỷ lệ ODR: {str(e)}"
            return jsonify({"reply": reply})

        # 4. GTC / LTC / Tổng quan
        elif 'gtc' in message or 'giao thành công' in message:
            if OPERATIONAL_CACHE and 'overall_gtc' in OPERATIONAL_CACHE:
                gtc_val = OPERATIONAL_CACHE['overall_gtc']
                reply = f"✅ **Tỷ lệ Giao thành công (GTC) hôm nay**: đạt <strong>{gtc_val}%</strong>.<br>(Target đề ra: >= 60.0%)"
            else:
                reply = "✅ Chưa có dữ liệu Giao thành công hôm nay."
            return jsonify({"reply": reply})

        elif 'ltc' in message or 'giao lần đầu' in message:
            if OPERATIONAL_CACHE and 'overall_ltc' in OPERATIONAL_CACHE:
                ltc_val = OPERATIONAL_CACHE['overall_ltc']
                reply = f"✈️ **Tỷ lệ Giao lần đầu thành công (LTC) hôm nay**: đạt <strong>{ltc_val}%</strong>."
            else:
                reply = "✈️ Chưa có dữ liệu Giao lần đầu thành công hôm nay."
            return jsonify({"reply": reply})

        # 5. Backlog / Tồn đọng
        elif 'backlog' in message or 'tồn' in message or 'chưa giao' in message:
            try:
                df_gtc, df_ltc, df_aging, df_treo = get_dataframes()
                total_aging = len(df_aging) if df_aging is not None else 0
                total_treo = len(df_treo) if df_treo is not None else 0
                reply = f"📦 **Tổng đơn tồn đọng (Backlog) hiện tại**: <strong>{total_aging + total_treo:,} đơn</strong>.<br>"
                reply += f"- Đơn giao trễ (Aging > 5 ngày): <strong>{total_aging:,} đơn</strong>.<br>"
                reply += f"- Đơn bị treo luân chuyển stuck: <strong>{total_treo:,} đơn</strong>."
            except Exception as e:
                reply = f"📦 Lỗi tính toán đơn backlog: {str(e)}"
            return jsonify({"reply": reply})
            
        else:
            # Help reply
            reply = "🤖 **Trợ lý NTB** xin chào!<br>Mình hỗ trợ trả lời nhanh các câu hỏi vận hành. Bạn có thể hỏi:<br>"
            reply += "- 📍 Bưu cục nào vận hành **tệ nhất**?<br>"
            reply += "- 📊 Tỷ lệ hiệu suất **OPR** hôm nay?<br>"
            reply += "- 🚚 Tỷ lệ giao hàng đúng giờ **ODR**?<br>"
            reply += "- 📦 Tổng số đơn **backlog** tồn đọng?<br>"
            reply += "- ✅ Tỷ lệ **GTC** hoặc **LTC** hiện tại?"
            return jsonify({"reply": reply})
            
    except Exception as ex:
        return jsonify({"reply": f"🤖 Đã xảy ra lỗi hệ thống: {str(ex)}"}), 500

@app.route('/api/ntb-structure')
@requires_permission('tab-introduction')
def get_ntb_structure():
    try:
        path = resolve_path('co_cau_ntb.csv', write=False)
        df = safe_read_csv(path)
        if df is None or df.empty:
            return jsonify({"error": "Không tìm thấy dữ liệu cơ cấu vùng (co_cau_ntb.csv)."}), 404
            
        df['Tỉnh'] = df['Tỉnh'].replace({'Khánh Hoà': 'Khánh Hòa', 'Bình Phước': 'Lâm Đồng'}).fillna("Chưa xác định").str.strip()
        df['AM'] = df['AM'].fillna("Chưa có AM").str.strip()
        df['Bưu cục'] = df['Bưu cục'].fillna("Chưa xác định").str.strip()
        
        structure = {}
        for prov, p_df in df.groupby('Tỉnh'):
            structure[prov] = {}
            for am, a_df in p_df.groupby('AM'):
                structure[prov][am] = sorted(list(a_df['Bưu cục'].unique()))
                
        return jsonify(structure)
    except Exception as e:
        return jsonify({"error": f"Lỗi đọc file cơ cấu: {str(e)}"}), 500

@app.route('/api/files-status')
@requires_permission('tab-sync')
def get_files_status():
    files = [
        'ops_gtc.csv',
        'ops_ltc.csv',
        'ops_co_cau.csv',
        'opr_opr.csv',
        'opr_oe.csv',
        'opr_raw.csv',
        'aging_raw.csv',
        'treo_stuck.csv',
        'buu_cuc_bat_on.csv',
        'off_tuyen_spe.csv',
        'vols_tao_don.csv',
        'co_cau_ntb.csv'
    ]
    status = []
    for f in files:
        path = resolve_path(f, write=False)
        if os.path.exists(path):
            stat = os.stat(path)
            mtime = datetime.datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M:%S")
            size_mb = round(stat.st_size / (1024 * 1024), 2)
            status.append({"name": f, "exists": True, "mtime": mtime, "size_mb": size_mb})
        else:
            status.append({"name": f, "exists": False, "mtime": "-", "size_mb": 0})
    return jsonify(status)
@app.route('/api/volume-creation')
@requires_permission('tab-volume-creation')
def get_volume_creation():
    global DF_TAO_DON_CACHE, DF_BUU_CUC_TYPE_MAP
    
    with CACHE_LOCK:
        if DF_TAO_DON_CACHE is None:
            DF_TAO_DON_CACHE = load_vols_tao_don_df()
        if DF_BUU_CUC_TYPE_MAP is None:
            DF_BUU_CUC_TYPE_MAP = load_buu_cuc_type_map()
            
    if DF_TAO_DON_CACHE is None:
        return jsonify({"error": "Không tìm thấy file vols_tao_don.csv. Vui lòng cấu hình Google Sheet và đồng bộ!"}), 404
        
    try:
        df = DF_TAO_DON_CACHE.copy()
        
        # Get query parameters
        province = request.args.get('province')
        district = request.args.get('district')
        ward = request.args.get('ward')
        post_office = request.args.get('post_office')
        customer = request.args.get('customer')
        po_type = request.args.get('po_type')
        date_range = request.args.get('date_range', '7d')
        
        # 1. Compute dynamic dropdown options based on cascaded filtering
        provinces_opt = sorted([str(x) for x in df['Tỉnh'].dropna().unique()])
        
        df_dist = df
        if province:
            df_dist = df_dist[df_dist['Tỉnh'] == province]
        districts_opt = sorted([str(x) for x in df_dist['Quận/Huyện'].dropna().unique()])
        
        df_ward = df_dist
        if district:
            df_ward = df_ward[df_ward['Quận/Huyện'] == district]
        wards_opt = sorted([str(x) for x in df_ward['Phường/Xã'].dropna().unique()])
        
        df_po = df_ward
        if ward:
            df_po = df_po[df_po['Phường/Xã'] == ward]
        post_offices_opt = sorted([str(x) for x in df_po['Bưu cục'].dropna().unique()])
        
        customers_opt = sorted([str(x) for x in df['Khách hàng'].dropna().unique()])
        po_types_opt = ['BC', 'GXT']
        
        # 2. Apply dropdown filters to our dataframe copies
        df_filtered = df.copy()
        if province:
            df_filtered = df_filtered[df_filtered['Tỉnh'] == province]
        if district:
            df_filtered = df_filtered[df_filtered['Quận/Huyện'] == district]
        if ward:
            df_filtered = df_filtered[df_filtered['Phường/Xã'] == ward]
        if post_office:
            df_filtered = df_filtered[df_filtered['Bưu cục'] == post_office]
        if customer:
            df_filtered = df_filtered[df_filtered['Khách hàng'] == customer]
        if po_type:
            if DF_BUU_CUC_TYPE_MAP:
                df_filtered['po_type_mapped'] = df_filtered['warehouse_id'].map(DF_BUU_CUC_TYPE_MAP).fillna('BC')
                df_filtered = df_filtered[df_filtered['po_type_mapped'] == po_type]
            else:
                # Fallback if map not loaded
                pass
                
        latest_dt = df_filtered['Date'].max()
        if pd.isna(latest_dt):
            return jsonify({"error": "Không tìm thấy dữ liệu ngày hợp lệ sau khi lọc."}), 400
            
        # WTD & comparison calculations based on filtered dataset (before date range filter)
        weekday = latest_dt.weekday()
        wtd_start = latest_dt - pd.Timedelta(days=weekday)
        wtd_end = latest_dt
        
        wtd1_start = wtd_start - pd.Timedelta(days=7)
        wtd1_end = latest_dt - pd.Timedelta(days=7)
        
        wtd2_start = wtd_start - pd.Timedelta(days=14)
        wtd2_end = latest_dt - pd.Timedelta(days=14)
        
        def sum_range(df_sub, start, end):
            mask = (df_sub['Date'] >= start) & (df_sub['Date'] <= end)
            return df_sub[mask].groupby('Tỉnh')['Volume'].sum()
            
        wtd_sums = sum_range(df_filtered, wtd_start, wtd_end)
        wtd1_sums = sum_range(df_filtered, wtd1_start, wtd1_end)
        wtd2_sums = sum_range(df_filtered, wtd2_start, wtd2_end)
        
        d_dt = latest_dt
        d1_dt = latest_dt - pd.Timedelta(days=1)
        d2_dt = latest_dt - pd.Timedelta(days=2)
        d7_dt = latest_dt - pd.Timedelta(days=7)
        
        def sum_day(df_sub, dt):
            return df_sub[df_sub['Date'] == dt].groupby('Tỉnh')['Volume'].sum()
            
        d_sums = sum_day(df_filtered, d_dt)
        d1_sums = sum_day(df_filtered, d1_dt)
        d2_sums = sum_day(df_filtered, d2_dt)
        d7_sums = sum_day(df_filtered, d7_dt)
        
        provinces_list = ['Bình Thuận', 'Khánh Hòa', 'Lâm Đồng', 'Ninh Thuận', 'Đắk Nông']
        pivot_rows = []
        
        for p in provinces_list:
            wtd = float(wtd_sums.get(p, 0))
            wtd1 = float(wtd1_sums.get(p, 0))
            wtd2 = float(wtd2_sums.get(p, 0))
            
            d_val = float(d_sums.get(p, 0))
            d1_val = float(d1_sums.get(p, 0))
            d2_val = float(d2_sums.get(p, 0))
            d7_val = float(d7_sums.get(p, 0))
            
            wtd_wtd2 = ((wtd - wtd2) / wtd2 * 100) if wtd2 > 0 else 0.0
            wtd_wtd1 = ((wtd - wtd1) / wtd1 * 100) if wtd1 > 0 else 0.0
            d_d7 = ((d_val - d7_val) / d7_val * 100) if d7_val > 0 else 0.0
            d_d1 = ((d_val - d1_val) / d1_val * 100) if d1_val > 0 else 0.0
            
            pivot_rows.append({
                'province': p,
                'wtd2': int(wtd2),
                'wtd1': int(wtd1),
                'wtd': int(wtd),
                'wtd_wtd2': round(wtd_wtd2, 1),
                'wtd_wtd1': round(wtd_wtd1, 1),
                'd_d7': round(d_d7, 1),
                'd_d1': round(d_d1, 1),
                'd7': int(d7_val),
                'd2': int(d2_val),
                'd1': int(d1_val),
                'd': int(d_val)
            })
            
        total_wtd = sum(r['wtd'] for r in pivot_rows)
        total_wtd1 = sum(r['wtd1'] for r in pivot_rows)
        total_wtd2 = sum(r['wtd2'] for r in pivot_rows)
        total_d = sum(r['d'] for r in pivot_rows)
        total_d1 = sum(r['d1'] for r in pivot_rows)
        total_d2 = sum(r['d2'] for r in pivot_rows)
        total_d7 = sum(r['d7'] for r in pivot_rows)
        
        total_row = {
            'province': 'Tổng cộng',
            'wtd2': int(total_wtd2),
            'wtd1': int(total_wtd1),
            'wtd': int(total_wtd),
            'wtd_wtd2': round(((total_wtd - total_wtd2) / total_wtd2 * 100) if total_wtd2 > 0 else 0.0, 1),
            'wtd_wtd1': round(((total_wtd - total_wtd1) / total_wtd1 * 100) if total_wtd1 > 0 else 0.0, 1),
            'd_d7': round(((total_d - total_d7) / total_d7 * 100) if total_d7 > 0 else 0.0, 1),
            'd_d1': round(((total_d - total_d1) / total_d1 * 100) if total_d1 > 0 else 0.0, 1),
            'd7': int(total_d7),
            'd2': int(total_d2),
            'd1': int(total_d1),
            'd': int(total_d)
        }
        
        kpi_card_data = {
            'latest_date': latest_dt.strftime('%d-%m-%Y'),
            'today_vol': int(total_d),
            'd_d1': round(((total_d - total_d1) / total_d1 * 100) if total_d1 > 0 else 0.0, 1),
            'd_d7': round(((total_d - total_d7) / total_d7 * 100) if total_d7 > 0 else 0.0, 1)
        }
        
        # 3. Apply Date Range filter for the matrix and charts
        df_date_filtered = df_filtered.copy()
        if date_range == '7d':
            start_date = latest_dt - pd.Timedelta(days=6)
            df_date_filtered = df_date_filtered[(df_date_filtered['Date'] >= start_date) & (df_date_filtered['Date'] <= latest_dt)]
        elif date_range == '14d':
            start_date = latest_dt - pd.Timedelta(days=13)
            df_date_filtered = df_date_filtered[(df_date_filtered['Date'] >= start_date) & (df_date_filtered['Date'] <= latest_dt)]
        elif date_range == '30d':
            start_date = latest_dt - pd.Timedelta(days=29)
            df_date_filtered = df_date_filtered[(df_date_filtered['Date'] >= start_date) & (df_date_filtered['Date'] <= latest_dt)]
            
        # Get list of unique dates for matrix and charts
        matrix_dates = sorted(df_date_filtered['Date'].dt.strftime('%Y-%m-%d').unique())
        
        # 4. Compute Matrix Heatmap Data
        matrix_rows = []
        if len(matrix_dates) > 0 and len(df_date_filtered) > 0:
            prov_date_vol = df_date_filtered.groupby(['Tỉnh', df_date_filtered['Date'].dt.strftime('%Y-%m-%d')])['Volume'].sum().unstack(fill_value=0)
            po_date_vol = df_date_filtered.groupby(['Tỉnh', 'Bưu cục', df_date_filtered['Date'].dt.strftime('%Y-%m-%d')])['Volume'].sum().unstack(fill_value=0)
            
            # Ensure all columns exist for all dates
            for d in matrix_dates:
                if d not in prov_date_vol.columns:
                    prov_date_vol[d] = 0
                if d not in po_date_vol.columns:
                    po_date_vol[d] = 0
                    
            for p in sorted(df_date_filtered['Tỉnh'].dropna().unique()):
                p_vols = prov_date_vol.loc[p].to_dict() if p in prov_date_vol.index else {}
                p_total = int(sum(p_vols.values()))
                
                pos_list = []
                if p in po_date_vol.index:
                    p_po_df = po_date_vol.loc[p]
                    for po in sorted(p_po_df.index):
                        po_vols = p_po_df.loc[po].to_dict()
                        po_total = int(sum(po_vols.values()))
                        pos_list.append({
                            'bc': po,
                            'data': {d: int(po_vols.get(d, 0)) for d in matrix_dates},
                            'total': po_total
                        })
                
                matrix_rows.append({
                    'province': p,
                    'data': {d: int(p_vols.get(d, 0)) for d in matrix_dates},
                    'total': p_total,
                    'pos': pos_list
                })
                
        # 5. Customer Group Line Chart Series
        line_series = []
        if len(matrix_dates) > 0 and len(df_date_filtered) > 0:
            cust_daily = df_date_filtered.groupby([df_date_filtered['Date'].dt.strftime('%Y-%m-%d'), 'Khách hàng'])['Volume'].sum().unstack(fill_value=0)
            standard_customers = ['SME', 'Shopee-nhỏ', 'TTS-nhỏ', 'Shopee-Bulky', 'Khác']
            for c in standard_customers:
                if c in cust_daily.columns:
                    c_data = [int(x) for x in cust_daily[c].tolist()]
                else:
                    c_data = [0] * len(matrix_dates)
                line_series.append({
                    'name': c,
                    'data': c_data
                })
                
        # 6. Grouped Bar Chart by Province Series
        bar_series = []
        if len(matrix_dates) > 0 and len(df_date_filtered) > 0:
            prov_daily = df_date_filtered.groupby([df_date_filtered['Date'].dt.strftime('%Y-%m-%d'), 'Tỉnh'])['Volume'].sum().unstack(fill_value=0)
            for p in provinces_list:
                if p in prov_daily.columns:
                    p_data = [int(x) for x in prov_daily[p].tolist()]
                else:
                    p_data = [0] * len(matrix_dates)
                bar_series.append({
                    'name': p,
                    'data': p_data
                })
                
        # 7. Treemap Data sorted by Absolute 7D Growth
        treemap_list = []
        # We compute D vs D-7. D is the latest date in df_filtered, D-7 is latest_dt - 7 days.
        df_d = df_filtered[df_filtered['Date'] == latest_dt]
        df_d7 = df_filtered[df_filtered['Date'] == (latest_dt - pd.Timedelta(days=7))]
        
        if len(df_d) > 0:
            vol_d = df_d.groupby(['Tỉnh', 'Bưu cục'])['Volume'].sum().reset_index()
            vol_d7 = df_d7.groupby('Bưu cục')['Volume'].sum().reset_index()
            
            merged_growth = pd.merge(vol_d, vol_d7, on='Bưu cục', suffixes=('_d', '_d7'), how='left').fillna(0)
            merged_growth['growth_abs'] = merged_growth['Volume_d'] - merged_growth['Volume_d7']
            merged_growth['growth_pct'] = (merged_growth['growth_abs'] / merged_growth['Volume_d7'] * 100).replace(np.inf, 100.0).replace(-np.inf, -100.0).fillna(0)
            
            # Sort by absolute growth descending
            merged_growth = merged_growth.sort_values(by='growth_abs', ascending=False)
            
            for _, row in merged_growth.iterrows():
                treemap_list.append({
                    'name': row['Bưu cục'],
                    'value': int(row['Volume_d']),
                    'province': row['Tỉnh'],
                    'growth_abs': int(row['growth_abs']),
                    'growth_pct': round(float(row['growth_pct']), 1)
                })
                
        return jsonify(clean_nan({
            "kpi": kpi_card_data,
            "table": pivot_rows,
            "total": total_row,
            "filters": {
                "provinces": provinces_opt,
                "districts": districts_opt,
                "wards": wards_opt,
                "post_offices": post_offices_opt,
                "customers": customers_opt,
                "po_types": po_types_opt
            },
            "matrix": {
                "dates": matrix_dates,
                "rows": matrix_rows
            },
            "charts": {
                "line": {
                    "dates": matrix_dates,
                    "series": line_series
                },
                "bar": {
                    "dates": matrix_dates,
                    "series": bar_series
                },
                "treemap": treemap_list
            }
        }))
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Lỗi xử lý báo cáo volume tạo đơn: {str(e)}"}), 500

def startup_cache_init():
    update_all_caches()
    try:
        # Initialize history with a sync if it is empty
        history = load_history()
        if not history and BACKLOG_CACHE_RAW:
            aging = BACKLOG_CACHE_RAW["aging"]
            treo = BACKLOG_CACHE_RAW["treo"]
            if "error" not in aging and "error" not in treo:
                add_to_history(aging, treo)
    except Exception as e:
        print(f"Error initializing history: {e}")

# Optimize to run cache initialization on startup thread upon import (for Gunicorn)
# But in local debug mode, only run it once to prevent double execution
if not app.debug or os.environ.get('WERKZEUG_RUN_MAIN') == 'true':
    threading.Thread(target=startup_cache_init, daemon=True).start()

def get_spiking_post_offices():
    global DF_TAO_DON_CACHE
    with CACHE_LOCK:
        if DF_TAO_DON_CACHE is None:
            DF_TAO_DON_CACHE = load_vols_tao_don_df()
    if DF_TAO_DON_CACHE is None:
        return []
        
    df = DF_TAO_DON_CACHE.copy()
    latest_dt = df['Date'].max()
    
    df_d = df[df['Date'] == latest_dt]
    df_d7 = df[df['Date'] == (latest_dt - pd.Timedelta(days=7))]
    
    if len(df_d) == 0 or len(df_d7) == 0:
        return []
        
    vol_d = df_d.groupby(['Tỉnh', 'Bưu cục'])['Volume'].sum().reset_index()
    vol_d7 = df_d7.groupby('Bưu cục')['Volume'].sum().reset_index()
    
    merged_growth = pd.merge(vol_d, vol_d7, on='Bưu cục', suffixes=('_d', '_d7'), how='left').fillna(0)
    merged_growth['growth_abs'] = merged_growth['Volume_d'] - merged_growth['Volume_d7']
    merged_growth['growth_pct'] = (merged_growth['growth_abs'] / merged_growth['Volume_d7'] * 100).replace(np.inf, 100.0).replace(-np.inf, -100.0).fillna(0)
    
    # Filter spikes: growth_abs >= 50 AND growth_pct >= 30%
    spikes = merged_growth[(merged_growth['growth_abs'] >= 50) & (merged_growth['growth_pct'] >= 30.0)]
    spikes = spikes.sort_values(by='growth_abs', ascending=False)
    
    spike_list = []
    for _, row in spikes.iterrows():
        spike_list.append({
            'name': row['Bưu cục'],
            'value': int(row['Volume_d']),
            'province': row['Tỉnh'],
            'growth_abs': int(row['growth_abs']),
            'growth_pct': round(float(row['growth_pct']), 1)
        })
    return spike_list

@app.route('/api/send-telegram-warning', methods=['POST'])
@requires_permission('tab-volume-creation')
def send_telegram_warning():
    try:
        config = load_config()
        bot_token = config.get("telegram_bot_token", "").strip()
        chat_id = config.get("telegram_chat_id", "").strip()
        
        # Accept updated config values in body to allow saving & sending at the same time
        data = request.json or {}
        req_token = data.get("telegram_bot_token", "").strip()
        req_chat_id = data.get("telegram_chat_id", "").strip()
        
        if req_token and not ("*" in req_token or "hidden" in req_token.lower()):
            bot_token = req_token
            # Also save it permanently
            save_config({
                "consolidated_url": config.get("consolidated_url", ""),
                "telegram_bot_token": req_token,
                "telegram_chat_id": req_chat_id if req_chat_id else chat_id
            })
        elif req_chat_id and req_chat_id != chat_id:
            chat_id = req_chat_id
            save_config({
                "consolidated_url": config.get("consolidated_url", ""),
                "telegram_bot_token": config.get("telegram_bot_token", ""),
                "telegram_chat_id": req_chat_id
            })
            
        if not bot_token or not chat_id:
            return jsonify({"error": "Chưa cấu hình Telegram Bot Token hoặc Chat ID / Group ID."}), 400
            
        spikes = get_spiking_post_offices()
        
        if spikes:
            items_str = ""
            for idx, item in enumerate(spikes):
                items_str += f"{idx + 1}. *{item['name']}* ({item['province']})\n"
                items_str += f"   - Sản lượng hôm nay: *{item['value']:,}* đơn\n"
                items_str += f"   - Tăng trưởng 7D: *+{item['growth_abs']:,}* đơn (*+{item['growth_pct']}%*)\n\n"
                
            message = (
                f"🚨 *CẢNH BÁO SẢN LƯỢNG TẠO ĐƠN TĂNG ĐỘT BIẾN* 🚨\n"
                f"(So sánh sản lượng hôm nay với 7 ngày trước)\n\n"
                f"Danh sách bưu cục có sản lượng tăng đột biến (>30% và >50 đơn):\n\n"
                f"{items_str}"
                f"📊 *Tổng quan:* Phát hiện {len(spikes)} bưu cục tăng đột biến trong Vùng."
            )
        else:
            message = (
                f"ℹ️ *BÁO CÁO SẢN LƯỢNG TẠO ĐƠN VÙNG* ℹ️\n"
                f"Không phát hiện bưu cục nào có sản lượng tăng đột biến (>30% và >50 đơn) trong 7 ngày qua."
            )
            
        url = f"https://api.telegram.org/bot{bot_token}/sendMessage"
        payload = {
            "chat_id": chat_id,
            "text": message,
            "parse_mode": "Markdown"
        }
        res = requests.post(url, json=payload, timeout=10)
        
        if res.status_code == 200:
            return jsonify({"success": True, "message": "Gửi cảnh báo qua Telegram thành công!"})
        else:
            return jsonify({"error": f"Lỗi từ Telegram API: {res.text}"}), 500
            
    except Exception as e:
        return jsonify({"error": f"Lỗi gửi tin nhắn: {str(e)}"}), 500

def get_overall_metrics_summary():
    try:
        df_gtc, df_ltc, df_aging, df_treo = get_dataframes(force=True)
        if df_gtc is None or df_ltc is None:
            return None
            
        all_times = set(df_gtc['Time'].dropna().unique()).union(set(df_ltc['Time'].dropna().unique()))
        dates_sorted = sorted(list(all_times), key=lambda x: pd.to_datetime(x.split(' - ')[0]))
        if not dates_sorted:
            return None
            
        latest_date = dates_sorted[-1]
        latest_idx = dates_sorted.index(latest_date)
        n3_date = dates_sorted[latest_idx - 3] if latest_idx >= 3 else None
        
        latest_dt = pd.to_datetime(latest_date.split(' - ')[0])
        lw_date = None
        for d in dates_sorted:
            dt = pd.to_datetime(d.split(' - ')[0])
            if (latest_dt - dt).days == 7:
                lw_date = d
                break
        
        def get_kpi_for_day(date_str):
            if not date_str:
                return None
            gtc_d = df_gtc[df_gtc['Time'] == date_str]
            ltc_d = df_ltc[df_ltc['Time'] == date_str] if date_str in df_ltc['Time'].values else df_ltc[df_ltc['Time'] == df_ltc['Time'].max()]
            
            gtc_vol = gtc_d['Volume'].sum()
            gtc_del = gtc_d['delivered_vol'].sum()
            ltc_vol = ltc_d['Volume'].sum()
            ltc_ltc = ltc_d['ltc_vol'].sum()
            
            return {
                'gtc_pct': round((gtc_del / gtc_vol * 100), 2) if gtc_vol > 0 else 0.0,
                'ltc_pct': round((ltc_ltc / ltc_vol * 100), 2) if ltc_vol > 0 else 0.0,
                'gtc_vol': int(gtc_vol),
                'ltc_vol': int(ltc_vol)
            }
            
        today_kpi = get_kpi_for_day(latest_date)
        n3_kpi = get_kpi_for_day(n3_date) if n3_date else None
        lw_kpi = get_kpi_for_day(lw_date) if lw_date else None
        
        chua_giao = len(df_aging)
        chua_lay = len(df_treo[df_treo['Loại đơn'] == 'Luân chuyển giao'])
        chua_tra = len(df_treo[df_treo['Loại đơn'] == 'Luân chuyển trả'])
        backlog_total = chua_giao + chua_lay + chua_tra
        
        today_gtc_po = df_gtc[df_gtc['Time'] == latest_date].copy()
        today_gtc_po['% GTC'] = (today_gtc_po['delivered_vol'] / today_gtc_po['Volume'] * 100).fillna(0)
        worst_gtc_po = today_gtc_po[today_gtc_po['Volume'] >= 10].sort_values(by='% GTC').head(3)
        
        today_ltc_po = df_ltc[df_ltc['Time'] == latest_date].copy()
        today_ltc_po['% LTC'] = (today_ltc_po['ltc_vol'] / today_ltc_po['Volume'] * 100).fillna(0)
        worst_ltc_po = today_ltc_po[today_ltc_po['Volume'] >= 10].sort_values(by='% LTC').head(3)
        
        worst_gtc_list = []
        for _, r in worst_gtc_po.iterrows():
            worst_gtc_list.append({
                'name': r.get('Chi tiết', 'N/A'),
                'province': r.get('mapped_prov', 'N/A'),
                'vol': int(r.get('Volume', 0)),
                'pct': float(r.get('% GTC', 0.0))
            })
            
        worst_ltc_list = []
        for _, r in worst_ltc_po.iterrows():
            worst_ltc_list.append({
                'name': r.get('Chi tiết', 'N/A'),
                'province': r.get('mapped_prov', 'N/A'),
                'vol': int(r.get('Volume', 0)),
                'pct': float(r.get('% LTC', 0.0))
            })
            
        spikes = get_spiking_post_offices()
        
        return {
            'date': latest_date,
            'today': today_kpi,
            'n3': n3_kpi,
            'lw': lw_kpi,
            'backlog': {
                'total': backlog_total,
                'chua_giao': chua_giao,
                'chua_lay': chua_lay,
                'chua_tra': chua_tra
            },
            'worst_gtc': worst_gtc_list,
            'worst_ltc': worst_ltc_list,
            'spikes': spikes[:5]
        }
    except Exception as e:
        print(f"Error in get_overall_metrics_summary: {e}")
        return None

@app.route('/api/send-telegram-ai-briefing', methods=['POST'])
@requires_permission('tab-volume-creation')
def send_telegram_ai_briefing():
    try:
        config = load_config()
        bot_token = config.get("telegram_bot_token", "").strip()
        chat_id = config.get("telegram_chat_id", "").strip()
        gemini_api_key = config.get("gemini_api_key", "").strip()
        
        print(f"DEBUG IN SERVER BRIEFING: bot_token={bot_token[:10] if bot_token else None}... chat_id={chat_id} gemini_api_key={gemini_api_key[:10] if gemini_api_key else None}...")
        data = request.json or {}
        req_token = data.get("telegram_bot_token", "").strip()
        req_chat_id = data.get("telegram_chat_id", "").strip()
        req_gemini_key = data.get("gemini_api_key", "").strip()
        
        # Save config if changed and not masked
        should_save = False
        updated_config = {
            "consolidated_url": config.get("consolidated_url", ""),
            "telegram_bot_token": config.get("telegram_bot_token", ""),
            "telegram_chat_id": config.get("telegram_chat_id", ""),
            "gemini_api_key": config.get("gemini_api_key", "")
        }
        
        if req_token and not ("*" in req_token or "hidden" in req_token.lower()):
            bot_token = req_token
            updated_config["telegram_bot_token"] = req_token
            should_save = True
            
        if req_chat_id and req_chat_id != chat_id:
            chat_id = req_chat_id
            updated_config["telegram_chat_id"] = req_chat_id
            should_save = True
            
        if req_gemini_key and not ("*" in req_gemini_key or "hidden" in req_gemini_key.lower()):
            gemini_api_key = req_gemini_key
            updated_config["gemini_api_key"] = req_gemini_key
            should_save = True
            
        if should_save:
            save_config(updated_config)
            
        if not bot_token or not chat_id:
            return jsonify({"error": "Chưa cấu hình Telegram Bot Token hoặc Chat ID / Group ID."}), 400
            
        if not gemini_api_key:
            return jsonify({"error": "Chưa cấu hình Gemini API Key."}), 400
            
        summary = get_overall_metrics_summary()
        if not summary:
            return jsonify({"error": "Không thể tổng hợp chỉ số vận hành."}), 500
            
        # Format worst post offices strings
        worst_gtc_str = ", ".join([f"{item['name']} ({item['province']}): {item['pct']:.1f}% GTC" for item in summary['worst_gtc']])
        worst_ltc_str = ", ".join([f"{item['name']} ({item['province']}): {item['pct']:.1f}% LTC" for item in summary['worst_ltc']])
        
        # Format volume spikes
        if summary['spikes']:
            spikes_str = "\n".join([f"- {s['name']} ({s['province']}): +{s['growth_abs']:,} đơn (+{s['growth_pct']}%)" for s in summary['spikes']])
        else:
            spikes_str = "Không phát hiện bưu cục tăng đơn đột biến."
            
        # Construct the Prompt for Gemini
        prompt = (
            f"Bạn là một AI Trợ lý Vận hành cao cấp phụ trách Vùng NTB của hệ thống Giao Hàng Nhanh (GHN).\n"
            f"Hãy phân tích dữ liệu vận hành tổng thể ngày {summary['date']} dưới đây và viết một bản tin phân tích cảnh báo đầu ngày gửi cho Group Telegram của Ban điều hành.\n\n"
            f"Dữ liệu vận hành ngày {summary['date']}:\n"
            f"1. Các chỉ số chính (KPI) hôm nay:\n"
            f"   - Giao thành công (% GTC): {summary['today']['gtc_pct']}%\n"
            f"   - Lưu kho thành công (% LTC): {summary['today']['ltc_pct']}%\n"
            f"   - Tổng sản lượng GTC: {summary['today']['gtc_vol']:,} đơn\n"
            f"   - Tổng sản lượng LTC: {summary['today']['ltc_vol']:,} đơn\n\n"
            f"2. So sánh hiệu suất (Tỷ lệ tăng/giảm %):\n"
            f"   - So với 3 ngày trước:\n"
            f"     + GTC: {round(summary['today']['gtc_pct'] - (summary['n3']['gtc_pct'] if summary['n3'] else 0), 2):+}% \n"
            f"     + LTC: {round(summary['today']['ltc_pct'] - (summary['n3']['ltc_pct'] if summary['n3'] else 0), 2):+}% \n"
            f"   - So với tuần trước:\n"
            f"     + GTC: {round(summary['today']['gtc_pct'] - (summary['lw']['gtc_pct'] if summary['lw'] else 0), 2):+}% \n"
            f"     + LTC: {round(summary['today']['ltc_pct'] - (summary['lw']['ltc_pct'] if summary['lw'] else 0), 2):+}% \n\n"
            f"3. Đơn tồn đọng (Backlog):\n"
            f"   - Tổng tồn đọng: {summary['backlog']['total']:,} đơn\n"
            f"     + Chưa giao (Aging): {summary['backlog']['chua_giao']:,} đơn\n"
            f"     + Chưa lấy (Treo giao): {summary['backlog']['chua_lay']:,} đơn\n"
            f"     + Chưa trả (Treo trả): {summary['backlog']['chua_tra']:,} đơn\n\n"
            f"4. Bưu cục cần đặc biệt lưu ý hôm nay (Hiệu suất tệ nhất):\n"
            f"   - Top GTC thấp nhất: {worst_gtc_str}\n"
            f"   - Top LTC thấp nhất: {worst_ltc_str}\n\n"
            f"5. Cảnh báo sản lượng tạo đơn tăng đột biến:\n"
            f"   {spikes_str}\n\n"
            f"Yêu cầu định dạng bản tin gửi Telegram:\n"
            f"- Định dạng Markdown chuẩn của Telegram (sử dụng *chữ đậm*, `chữ code`, biểu tượng emoji thích hợp).\n"
            f"- Tránh sử dụng các ký tự đặc biệt làm lỗi bộ phân tích cú pháp Markdown của Telegram (như dấu ngoặc vuông hoặc dấu gạch dưới đơn lẻ không đóng cặp). Hãy giữ Markdown đơn giản bằng cách dùng dấu * cho chữ đậm và không lồng ghép định dạng.\n"
            f"- Phong cách viết: Chuyên nghiệp, nghiêm túc, tập trung vào hành động nhanh và chỉ rõ các điểm nóng cần khắc phục (đặc biệt lưu ý bưu cục hiệu suất thấp hoặc có sản lượng tăng đột biến).\n"
            f"- Hãy chia rõ 4 phần sau:\n"
            f"  * 📊 *BẢN TIN VẬN HÀNH ĐẦU NGÀY VÙNG NTB* ({summary['date']})\n"
            f"  * ⚠️ *CẢNH BÁO ĐIỂM NÓNG* (Bưu cục yếu kém và backlog cao)\n"
            f"  * 🚨 *CẢNH BÁO SẢN LƯỢNG ĐỘT BIẾN*\n"
            f"  * 💡 *HÀNH ĐỘNG KHUYẾN NGHỊ* (Gợi ý hành động thực tế cho AM và Bưu cục trưởng)"
        )
        
        # Call Gemini REST API
        url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-flash-latest:generateContent?key={gemini_api_key}"
        payload = {
            "contents": [{
                "parts": [{"text": prompt}]
            }]
        }
        
        res_gemini = None
        try:
            res_gemini = requests.post(url, json=payload, headers={"Content-Type": "application/json"}, timeout=15)
        except Exception as e:
            print(f"Primary Gemini model failed with exception: {e}. Trying fallback...")
            
        # Fallback to gemini-3.1-flash-lite if primary is unavailable or timed out
        if res_gemini is None or res_gemini.status_code != 200:
            url_fallback = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-3.1-flash-lite:generateContent?key={gemini_api_key}"
            try:
                res_gemini = requests.post(url_fallback, json=payload, headers={"Content-Type": "application/json"}, timeout=30)
            except Exception as e:
                return jsonify({"error": f"Lỗi từ Gemini API (Fallback cũng thất bại): {str(e)}"}), 500
                
        if res_gemini and res_gemini.status_code == 200:
            gemini_data = res_gemini.json()
            message = gemini_data['candidates'][0]['content']['parts'][0]['text']
        else:
            err_text = res_gemini.text if res_gemini else "No response"
            return jsonify({"error": f"Lỗi từ Gemini API: {err_text}"}), 500
            
        # Send message to Telegram
        url_tele = f"https://api.telegram.org/bot{bot_token}/sendMessage"
        payload_tele = {
            "chat_id": chat_id,
            "text": message,
            "parse_mode": "Markdown"
        }
        res_tele = requests.post(url_tele, json=payload_tele, timeout=10)
        
        if res_tele.status_code == 200:
            return jsonify({"success": True, "message": "Gửi bản tin phân tích AI đầu ngày thành công!"})
        else:
            # Fallback to plain text if Markdown fails
            payload_tele["parse_mode"] = ""
            res_tele_plain = requests.post(url_tele, json=payload_tele, timeout=10)
            if res_tele_plain.status_code == 200:
                return jsonify({"success": True, "message": "Gửi bản tin phân tích AI thành công (chế độ plain-text do lỗi định dạng Markdown)!"})
            return jsonify({"error": f"Lỗi từ Telegram API: {res_tele.text}"}), 500
            
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Lỗi gửi bản tin AI: {str(e)}"}), 500

if __name__ == '__main__':
    print("Dashboard server starts on http://0.0.0.0:5000/")
    app.run(debug=False, host='0.0.0.0', port=5000)
