import sys
import openpyxl
import os

sys.stdout.reconfigure(encoding='utf-8')

workspace_dir = r"c:\Users\lap4all\Desktop\New folder"
file_path = os.path.join(workspace_dir, "Copy o NTB - BÁO CÁO VẬN HÀNH.xlsx")

if os.path.exists(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    for name in wb.sheetnames:
        sheet = wb[name]
        for r in range(1, sheet.max_row + 1):
            for c in range(1, sheet.max_column + 1):
                val = sheet.cell(row=r, column=c).value
                if val is not None:
                    try:
                        f_val = float(val)
                        if (0.8776 <= f_val <= 0.8778) or (87.76 <= f_val <= 87.78):
                            print(f"Match found in Sheet '{name}' at Cell ({r}, {c}) = {f_val}")
                    except ValueError:
                        pass
else:
    print("File not found")
