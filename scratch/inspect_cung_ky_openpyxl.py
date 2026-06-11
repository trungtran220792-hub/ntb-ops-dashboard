import openpyxl
import os

workspace_dir = r"c:\Users\lap4all\Desktop\New folder"
file_path = os.path.join(workspace_dir, "Copy o NTB - BÁO CÁO VẬN HÀNH.xlsx")
output_path = os.path.join(workspace_dir, "scratch", "inspect_cung_ky_detailed.txt")

with open(output_path, "w", encoding="utf-8") as f:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb["DataLTC"]
    
    row1 = [cell.value for cell in ws[1]]
    f.write(f"Row 1 (Header):\n{row1}\n\n")
    
    row2 = [cell.value for cell in ws[2]]
    f.write(f"Row 2:\n{row2}\n\n")

print("Done inspect openpyxl.")
