import openpyxl
import os

workspace_dir = r"c:\Users\lap4all\Desktop\New folder"
file_path = os.path.join(workspace_dir, "Copy o NTB - BÁO CÁO VẬN HÀNH.xlsx")

wb = openpyxl.load_workbook(file_path, data_only=True)
ws = wb["DataLTC"]

print("Row 1 (Header):")
row1 = [cell.value for cell in ws[1]]
print(row1)

print("\nRow 2:")
row2 = [cell.value for cell in ws[2]]
print(row2)
