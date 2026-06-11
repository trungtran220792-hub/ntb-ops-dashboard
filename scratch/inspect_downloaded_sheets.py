import openpyxl
import pandas as pd
import sys

output_file = "scratch/inspection_output.txt"
with open(output_file, "w", encoding="utf-8") as out:
    def log(msg):
        out.write(str(msg) + "\n")

    log("Loading workbook...")
    try:
        wb_path = "scratch/downloaded_consolidated_sheet.xlsx"
        wb = openpyxl.load_workbook(wb_path, read_only=True)
        log(f"Workbook sheets: {wb.sheetnames}")
        
        for name in wb.sheetnames:
            log(f"\n=================== Sheet: {name} ===================")
            try:
                df = pd.read_excel(wb_path, sheet_name=name, nrows=5)
                log(f"Shape: {df.shape}")
                log(f"Columns: {list(df.columns)}")
                log("First row data:")
                if not df.empty:
                    log(str(df.iloc[0].to_dict()))
                else:
                    log("Empty sheet")
            except Exception as e:
                log(f"Error reading sheet {name}: {e}")
    except Exception as e:
        log(f"General error: {e}")
