import pandas as pd
import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

xls_path = "scratch/temp_download.xlsx"
wb = openpyxl.load_workbook(xls_path, data_only=True)

targets = [1852, 957, 1098, 783, 1299, 1164, 608, 44209, 44209.0]

print("Starting cell search in Excel...")
for name in wb.sheetnames:
    ws = wb[name]
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=r, column=c).value
            if val is not None:
                # Check if matches any target
                try:
                    val_num = float(val)
                    if any(abs(val_num - t) < 1.0 for t in targets):
                        print(f"FOUND match in sheet '{name}' at cell ({r}, {c}): val={val}")
                except (ValueError, TypeError):
                    # Check string contains
                    val_str = str(val)
                    if any(str(t) in val_str for t in targets):
                        print(f"FOUND string match in sheet '{name}' at cell ({r}, {c}): val={val}")

print("Search complete.")
